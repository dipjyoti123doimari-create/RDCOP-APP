"""
flask_app.py
============
Production Flask application for Batching Incentive & Deduction Calculator.
Replaces the Streamlit front-end; all Python business-logic modules are
kept exactly as-is (calculator.py, database.py, oracle_connector.py …).

Run with:
    python flask_app.py
Opens on http://localhost:2001
"""

from __future__ import annotations

import io
import json
import logging
import os
import subprocess
import sys
import threading
import traceback
import uuid
from datetime import date as _date, datetime as _dt, timedelta

# Load .env if present (for SECRET_KEY, ADMIN_USERNAME, etc.)
try:
    from dotenv import load_dotenv as _load_dotenv
    _load_dotenv()
except ImportError:
    pass  # python-dotenv not installed; rely on environment variables set externally

import pandas as pd
from apscheduler.schedulers.background import BackgroundScheduler
from apscheduler.triggers.cron import CronTrigger
from flask import (Flask, flash, g, jsonify, make_response, redirect, render_template,
                   request, send_file, url_for)

import auth
import cache_helpers
import calculator
import config
import data_loader
import database
import email_helper
import google_sheets
import oracle_connector
import report_generator
import btrtp_calculator
import ecmd_calculator
import tp_calculator
import validations
from modules.slow_loading_alert import scheduler as sla_scheduler

# ── File logging (survives minimised/closed terminal windows) ────────────────
_LOG_FILE = os.path.join(os.path.dirname(os.path.abspath(__file__)), "server.log")
logging.basicConfig(
    level=logging.WARNING,
    format="%(asctime)s [%(levelname)s] %(message)s",
    handlers=[
        logging.FileHandler(_LOG_FILE, encoding="utf-8"),
        logging.StreamHandler(sys.stdout),
    ],
)
_log = logging.getLogger(__name__)

def _handle_unhandled_exception(exc_type, exc_value, exc_tb):
    if issubclass(exc_type, KeyboardInterrupt):
        sys.__excepthook__(exc_type, exc_value, exc_tb)
        return
    _log.critical("Unhandled exception — server will exit:\n%s",
                  "".join(traceback.format_exception(exc_type, exc_value, exc_tb)))

sys.excepthook = _handle_unhandled_exception

# ── App setup ────────────────────────────────────────────────────────────────
app = Flask(__name__)
app.secret_key = os.environ.get("SECRET_KEY", "rdc-incentive-local-key-2025")
app.config["SESSION_PERMANENT"] = False
app.config["PERMANENT_SESSION_LIFETIME"] = timedelta(minutes=20)

# ── Single-user in-memory state (no session serialization overhead) ──────────
_S: dict = {}


def _s(key, default=None):
    return _S.get(key, default)


def _ss(key, val):
    _S[key] = val


# ── Module-scoped session state ───────────────────────────────────────────────
# New modules MUST use _ms/_mss instead of _s/_ss so keys never collide.
# Keys are stored as "{module}.{key}" inside the shared _S dict.
# Example:  _mss("tp", "results", df)  →  _S["tp.results"] = df
#           _ms("tp", "results")       →  _S.get("tp.results")
def _ms(module: str, key: str, default=None):
    return _S.get(f"{module}.{key}", default)


def _mss(module: str, key: str, val):
    _S[f"{module}.{key}"] = val


# ── Real-time operation progress (single-user, single active job at a time) ───
_progress: dict = {"pct": 0, "msg": ""}
_progress_lock = threading.Lock()


def _set_progress(pct: int, msg: str = ""):
    with _progress_lock:
        _progress["pct"] = pct
        _progress["msg"] = msg


# ── Boot ─────────────────────────────────────────────────────────────────────
database.init_db()
auth.bootstrap_admin(app)

# Apply the rolling Oracle-data retention once at startup (covers the case
# where the app was restarted after a month rolled over).
try:
    _purge = database.purge_old_oracle_data()
    if any(_purge["deleted"].values()):
        print(f"[retention] purged Oracle data before {_purge['cutoff']}: {_purge['deleted']}")
except Exception as _exc:
    print(f"[retention] startup purge skipped: {_exc}")

# ── Monthly email scheduler ───────────────────────────────────────────────────
_scheduler: BackgroundScheduler | None = None


def _retention_job():
    """Daily: enforce the shared rolling Oracle-data retention for all modules."""
    try:
        database.purge_old_oracle_data()
    except Exception:
        pass


def _scheduled_email_job():
    """Fires on 1st of each month — sends previous month's report."""
    if database.get_setting("email_schedule_enabled", "false") != "true":
        return
    today         = _date.today()
    first_this    = today.replace(day=1)
    last_prev     = first_this - timedelta(days=1)
    first_prev    = last_prev.replace(day=1)
    from_s, to_s  = str(first_prev), str(last_prev)
    try:
        res = calculator.run_calculation(
            month=first_prev.month, year=first_prev.year,
            start_date=from_s, end_date=to_s, persist=False,
        )
        if res["error"]:
            database.set_setting("email_schedule_last_status",
                f"FAIL {_dt.now():%Y-%m-%d %H:%M} — {res['error']}")
            return
        rows     = res.get("results_rows", [])
        unmapped = res.get("unmapped_rows", [])
        month_label = first_prev.strftime("%B %Y")
        meta    = _build_meta(from_s, to_s, rows, unmapped, "Scheduled monthly report")
        df_f    = pd.DataFrame(rows)    if rows     else pd.DataFrame(columns=RESULT_COLS)
        df_u    = pd.DataFrame(unmapped) if unmapped else pd.DataFrame()
        val_df  = database.read_table("validation_errors")
        xlsx    = report_generator.generate_excel_report(df_f, df_u, val_df, meta)
        fname   = f"incentive_report_{from_s}_to_{to_s}.xlsx"
        to_addr = database.get_setting("email_schedule_to", "")
        cc_addr = database.get_setting("email_schedule_cc", "")
        result  = email_helper.send_report_email(
            to_emails=to_addr, cc_emails=cc_addr,
            subject=email_helper.compose_report_subject(month_label),
            body=email_helper.compose_report_body(month_label),
            attachment_bytes=xlsx, attachment_name=fname,
        )
        status = (f"OK {_dt.now():%Y-%m-%d %H:%M} — sent to {to_addr}" if result["success"]
                  else f"FAIL {_dt.now():%Y-%m-%d %H:%M} — {result.get('error','')}")
        database.set_setting("email_schedule_last_status", status)
    except Exception as exc:
        database.set_setting("email_schedule_last_status",
            f"ERROR {_dt.now():%Y-%m-%d %H:%M} — {exc}")


def _build_tp_excel(plant_rows, location_rows) -> bytes:
    """Build the RDC-TP report workbook (Plant + Location sheets) as bytes."""
    plant_df = pd.DataFrame(plant_rows)[[
        "lookup_code","plant_name","exco_location","business_head",
        "plant_manager","mixer_theo_cap","total_quantity","total_time_min",
        "throughput_pct","batch_count"
    ]].rename(columns={
        "lookup_code":"Plant Code","plant_name":"Plant","exco_location":"Exco Location",
        "business_head":"Business Head","plant_manager":"Plant Manager",
        "mixer_theo_cap":"Mixer Capacity","total_quantity":"Total Qty",
        "total_time_min":"Total Time (min)","throughput_pct":"Throughput %",
        "batch_count":"Batches",
    })
    loc_df = pd.DataFrame(location_rows)[[
        "exco_location","plant_count","total_quantity","avg_throughput_pct"
    ]].rename(columns={
        "exco_location":"Exco Location","plant_count":"Plants",
        "total_quantity":"Total Qty","avg_throughput_pct":"Avg Throughput %",
    })
    buf = io.BytesIO()
    with pd.ExcelWriter(buf, engine="openpyxl") as writer:
        plant_df.to_excel(writer, sheet_name="Plant Throughput", index=False)
        loc_df.to_excel(writer, sheet_name="Location Throughput", index=False)
    buf.seek(0)
    return buf.read()


def _shared_oracle_fetch_job():
    """
    Single daily Oracle fetch for ALL modules (replaces 3 separate jobs).
    Fetches previous month (full) + current month (1st → today) in one query.
    Stores to oracle_raw_data — shared by I&D, TP, BTRTP calculators.
    """
    if not oracle_connector.is_configured():
        return
    import calendar as _cal
    today     = _date.today()
    cur_from  = today.replace(day=1)
    cur_to    = today
    if today.month == 1:
        prev_from = _date(today.year - 1, 12, 1)
        prev_to   = _date(today.year - 1, 12, _cal.monthrange(today.year - 1, 12)[1])
    else:
        prev_from = _date(today.year, today.month - 1, 1)
        prev_to   = _date(today.year, today.month - 1,
                          _cal.monthrange(today.year, today.month - 1)[1])
    try:
        total = 0
        for fd, td in [(str(prev_from), str(prev_to)), (str(cur_from), str(cur_to))]:
            raw_df, warnings = oracle_connector.fetch_oracle_raw_data(fd, td)
            for w in warnings:
                print(f"[oracle-fetch] {w}")
            if not raw_df.empty:
                saved = oracle_connector.save_oracle_raw_data(raw_df, replace=True)
                total += saved
        database.purge_old_oracle_data()
        print(f"[oracle-fetch] done — {total} rows saved "
              f"(prev: {prev_from}→{prev_to}, cur: {cur_from}→{cur_to})")
    except Exception as exc:
        print(f"[oracle-fetch] error: {exc}")



def _auto_calc_one_period(month, year, fd, td, label):
    """Calculate all 3 modules for a single period. Called for both prev and current month."""
    import calendar as _cal2
    try:
        result = calculator.run_calculation(month, year, start_date=fd, end_date=td, persist=True)
        if result.get("error"):
            print(f"[auto-calc] I&D {label} skipped — {result['error']}")
        elif result["total_employees"] == 0:
            print(f"[auto-calc] I&D {label} skipped — no production data")
        else:
            print(f"[auto-calc] I&D {label} done — {result['total_employees']} employees")
    except Exception as exc:
        print(f"[auto-calc] I&D {label} error: {exc}")
    try:
        plant_rows, loc_rows, warns = tp_calculator.run_tp_calculation(
            month, year, from_date=fd, to_date=td)
        if plant_rows:
            tp_calculator.save_tp_results(plant_rows, month, year)
            print(f"[auto-calc] TP {label} done — {len(plant_rows)} plants")
        else:
            print(f"[auto-calc] TP {label} skipped — {'; '.join(warns[:2])}")
    except Exception as exc:
        print(f"[auto-calc] TP {label} error: {exc}")
    try:
        batcher_rows, warns = btrtp_calculator.run_btrtp_calculation(
            month, year, from_date=fd, to_date=td)
        if batcher_rows:
            btrtp_calculator.save_btrtp_results(batcher_rows, month, year)
            print(f"[auto-calc] BTRTP {label} done — {len(batcher_rows)} batcher rows")
        else:
            print(f"[auto-calc] BTRTP {label} skipped — {'; '.join(warns[:2])}")
    except Exception as exc:
        print(f"[auto-calc] BTRTP {label} error: {exc}")


def _auto_calculate_current_month():
    """
    Runs daily at 00:30 — after Oracle data has been fetched (00:10).
    Calculates I&D, TP, and BTRTP for both previous month (full) and current month.
    Previous month is always recalculated with complete data.
    Current month uses data up to today.
    """
    import calendar as _cal2
    today = _date.today()

    # Previous month — full range
    if today.month == 1:
        prev_m, prev_y = 12, today.year - 1
    else:
        prev_m, prev_y = today.month - 1, today.year
    prev_fd = str(_date(prev_y, prev_m, 1))
    prev_td = str(_date(prev_y, prev_m, _cal2.monthrange(prev_y, prev_m)[1]))
    _auto_calc_one_period(prev_m, prev_y, prev_fd, prev_td,
                          f"{prev_y}-{prev_m:02d}(prev)")

    # Current month — 1st → today
    cur_fd = str(today.replace(day=1))
    cur_td = str(today)
    _auto_calc_one_period(today.month, today.year, cur_fd, cur_td,
                          f"{today.year}-{today.month:02d}(cur)")


def _startup_oracle_fetch():
    """
    Runs once 45 seconds after startup in a background thread.
    Seeds local Oracle cache for all modules if Oracle is reachable.
    This ensures data is fresh even when the midnight cron was missed
    (e.g. server was off, VPN wasn't connected at midnight).
    """
    import time as _time
    _time.sleep(45)
    if not oracle_connector.is_configured() or not oracle_connector.is_reachable():
        print("[startup-fetch] Oracle not reachable — skipping startup seed")
        return
    print("[startup-fetch] Oracle reachable — running shared fetch for all modules")
    _shared_oracle_fetch_job()
    _auto_calculate_current_month()
    print("[startup-fetch] done")


def _tp_scheduled_email_job():
    """Fires on 1st of each month — emails the previous month's TP report."""
    if database.get_module_setting("tp", "email_schedule_enabled", "false") != "true":
        return
    today      = _date.today()
    first_this = today.replace(day=1)
    last_prev  = first_this - timedelta(days=1)
    first_prev = last_prev.replace(day=1)
    fd, td     = str(first_prev), str(last_prev)
    try:
        month, year = first_prev.month, first_prev.year
        plant_rows, loc_rows, _w = tp_calculator.run_tp_calculation(
            month, year, from_date=fd, to_date=td)
        if not plant_rows:
            database.set_module_setting("tp", "email_schedule_last_status",
                f"FAIL {_dt.now():%Y-%m-%d %H:%M} — no data for {fd} → {td}")
            return
        import calendar
        label   = f"{calendar.month_name[month]} {year}"
        subject = f"RDC-TP Plant Throughput Report — {label}"
        body    = (f"Dear Team,\n\nPlease find attached the Plant Throughput Report "
                   f"for {label}.\n\nRegards,\nRDC Operations")
        res = email_helper.send_report_email(
            to_emails=database.get_module_setting("tp", "email_schedule_to", ""),
            cc_emails=database.get_module_setting("tp", "email_schedule_cc", ""),
            subject=subject, body=body,
            attachment_bytes=_build_tp_excel(plant_rows, loc_rows),
            attachment_name=f"RDC_TP_{year}_{month:02d}.xlsx",
        )
        status = (f"OK {_dt.now():%Y-%m-%d %H:%M} — sent" if res["success"]
                  else f"FAIL {_dt.now():%Y-%m-%d %H:%M} — {res.get('error','')}")
        database.set_module_setting("tp", "email_schedule_last_status", status)
    except Exception as exc:
        database.set_module_setting("tp", "email_schedule_last_status",
            f"ERROR {_dt.now():%Y-%m-%d %H:%M} — {exc}")


def _start_scheduler():
    global _scheduler
    sched_time = database.get_setting("email_schedule_time", "08:00")
    try:
        h, m = map(int, sched_time.split(":"))
    except Exception:
        h, m = 8, 0
    _scheduler = BackgroundScheduler(daemon=True)
    _scheduler.add_job(_scheduled_email_job,
                       CronTrigger(day=1, hour=h, minute=m),
                       id="monthly_report", replace_existing=True)
    # Daily rolling retention for all modules' Oracle data (runs 00:05 + on 1st).
    _scheduler.add_job(_retention_job,
                       CronTrigger(hour=0, minute=5),
                       id="oracle_retention", replace_existing=True)
    # RDC-TP monthly report (own schedule time).
    tp_time = database.get_module_setting("tp", "email_schedule_time", "08:00")
    try:
        th, tm = map(int, tp_time.split(":"))
    except Exception:
        th, tm = 8, 0
    _scheduler.add_job(_tp_scheduled_email_job,
                       CronTrigger(day=1, hour=th, minute=tm),
                       id="tp_monthly_email", replace_existing=True)
    # Shared daily Oracle fetch — one query feeds I&D, TP, and BTRTP.
    _scheduler.add_job(_shared_oracle_fetch_job,
                       CronTrigger(hour=0, minute=10),
                       id="oracle_daily_fetch", replace_existing=True)
    # Auto-calculate TP and BTRTP for current month daily (after Oracle fetch).
    _scheduler.add_job(_auto_calculate_current_month,
                       CronTrigger(hour=0, minute=30),
                       id="auto_calculate_current_month", replace_existing=True)
    # SLA — hourly slow loading alert.
    _scheduler.add_job(sla_scheduler.run_hourly_alert_job,
                       CronTrigger(minute=0),  # fires at top of every hour
                       id="sla_hourly_alert", replace_existing=True)
    # SLA — daily summary (default 07:00; re-read from DB each time).
    _sla_ds_time = database.get_module_setting("sla", "daily_summary_time", "07:00")
    try:
        _sla_h, _sla_m = map(int, _sla_ds_time.split(":"))
    except Exception:
        _sla_h, _sla_m = 7, 0
    _scheduler.add_job(sla_scheduler.run_daily_summary_job,
                       CronTrigger(hour=_sla_h, minute=_sla_m),
                       id="sla_daily_summary", replace_existing=True)
    _scheduler.start()


_start_scheduler()

# Seed Oracle cache on startup (45s delay — waits for network/VPN to settle)
import threading as _threading
_threading.Thread(target=_startup_oracle_fetch, daemon=True).start()

# ── Jinja filters / globals ───────────────────────────────────────────────────

@app.template_filter("format_int")
def _fmt_int(v):
    try:
        return f"{int(v):,}"
    except Exception:
        return v


@app.template_global()
def bg_auto():
    return database.get_setting("bg_auto_theme", "true") == "true"


@app.template_global()
def bg_animate():
    return database.get_setting("bg_animate", "true") == "true"


@app.template_global()
def bg_theme():
    return database.get_setting("bg_manual_theme", "Daytime")


# ── Global auth gate ─────────────────────────────────────────────────────────
# Enforce login for every request except /login, /logout and static assets.
# This is the "minimum disruption" approach — existing routes need no changes.

_AUTH_EXEMPT = {"/login", "/logout"}

@app.before_request
def _require_login():
    if request.path.startswith("/static/"):
        return
    if request.path in _AUTH_EXEMPT:
        return
    user = auth.get_current_user()
    if user is None:
        if request.path.startswith("/api/"):
            return jsonify({"error": "Not authenticated"}), 401
        from flask import flash as _flash
        _flash("Please log in to continue.", "warning")
        return redirect(url_for("page_login", next=request.path))
    # Inject into g so route handlers can read it without calling get_current_user() again
    g.current_user = user

    # ── Role-based path restrictions ─────────────────────────────────────────
    role = user.get("role", "")
    path = request.path

    # SUPER_ADMIN-only: settings, calculate, upload/sync actions, admin panel
    _admin_only_patterns = (
        "/settings", "/calculate", "/data-uploader",
        "/action/upload-", "/action/sync-", "/action/fetch-oracle",
        "/action/run-calculation", "/action/run-validation",
        "/action/save-smtp", "/action/save-oracle",
        "/action/add-employee", "/action/update-employee", "/action/delete-employee",
        "/action/add-waiver", "/action/delete-waiver",
        "/action/restart-server", "/action/save-email-schedule",
        "/action/toggle-email-schedule", "/action/clear-",
        "/action/assign-maintenance", "/action/delete-maintenance",
        "/action/toggle-auto-sync", "/action/save-",
        "/tp/action/add-plant", "/tp/action/update-plant", "/tp/action/delete-plant",
        "/btrtp/settings", "/tp/settings",
    )
    _ecmd_entry_roles = (auth.PLANT_USER, auth.REGIONAL_USER,
                         auth.HO_VIEWER, auth.FINANCE_VIEWER, auth.UEP_ADMIN)
    if role != auth.SUPER_ADMIN:
        # ECMD routes (including settings/admin actions) are fully open to UEP_ADMIN
        # and data-entry routes are open to PLANT_USER / REGIONAL_USER / HO / FINANCE
        if path.startswith("/ecmd/") and role in _ecmd_entry_roles:
            pass  # let ECMD routes handle their own auth
        elif path.startswith("/admin/users") and role == auth.UEP_ADMIN:
            pass  # UEP_ADMIN user-management — route handles its own scope guard
        else:
            for pattern in _admin_only_patterns:
                if pattern in path:
                    auth.log_activity(user, "ACCESS_DENIED",
                                      details={"path": path, "method": request.method})
                    return render_template("access_denied.html",
                                           current_user=user,
                                           required_roles=[auth.SUPER_ADMIN]), 403

    # Email send: PLANT_USER cannot send emails
    if role == auth.PLANT_USER and "/action/send-email" in path:
        return render_template("access_denied.html",
                               current_user=user,
                               required_roles=[auth.SUPER_ADMIN, auth.HO_VIEWER,
                                               auth.FINANCE_VIEWER, auth.REGIONAL_USER]), 403

    # Non-SUPER_ADMIN users must not enter I&D / TP / BTRTP module pages
    # PLANT_USER and REGIONAL_USER ARE allowed into ECMD (data entry + reports)
    _module_prefixes = ("/dashboard", "/id", "/tp/", "/btrtp/")
    _ecmd_allowed_roles = (auth.PLANT_USER, auth.REGIONAL_USER,
                           auth.HO_VIEWER, auth.FINANCE_VIEWER, auth.UEP_ADMIN)
    if role != auth.SUPER_ADMIN and request.method == "GET":
        if path.startswith("/ecmd/") or path == "/ecmd":
            if role not in _ecmd_allowed_roles:
                return redirect(url_for("page_home"))
        elif path.startswith("/sla/") or path == "/sla":
            pass  # SLA handles its own plant-filter per route
        else:
            for pfx in _module_prefixes:
                if path == pfx.rstrip("/") or path.startswith(pfx):
                    return redirect(url_for("page_home"))

    # SLA: block config/manual-run for non-SUPER_ADMIN via POST
    if path.startswith("/sla/") and role != auth.SUPER_ADMIN:
        _sla_admin_paths = (
            "/sla/configuration", "/sla/threshold/add", "/sla/threshold/delete",
            "/sla/settings/save", "/sla/oracle-cols/save", "/sla/sheet-config/save",
            "/sla/manual/run-hourly-send", "/sla/manual/run-daily-send",
            "/sla/manual/send-test-email",
        )
        if any(path.startswith(p) for p in _sla_admin_paths):
            return render_template("access_denied.html",
                                   current_user=user,
                                   required_roles=[auth.SUPER_ADMIN]), 403


# ── Context processor (sidebar active-page + bg settings + auth) ────────────

@app.context_processor
def _ctx():
    auth_ctx = auth.inject_auth_context()
    return dict(
        active_page=_s("active_page", ""),
        bg_auto=bg_auto(),
        bg_animate=bg_animate(),
        bg_theme=bg_theme(),
        now=_dt.now(),
        **auth_ctx,
    )


# ── Helpers ───────────────────────────────────────────────────────────────────

def _records(df: pd.DataFrame) -> list:
    if df is None or df.empty:
        return []
    return json.loads(df.to_json(orient="records", date_format="iso"))


def _row_cls(row: dict) -> str:
    if (row.get("deduction_amount") or 0) > 0:
        return "row-red"
    if (row.get("incentive_amount") or 0) > 0:
        return "row-green"
    return ""


def _sort_rows(rows: list) -> list:
    def _key(r):
        d = r.get("deduction_amount") or 0
        i = r.get("incentive_amount") or 0
        if d > 0:
            return (0, -d, -i)
        if i > 0:
            return (1, 0, -i)
        return (2, 0, 0)
    return sorted(rows, key=_key)


def _apply_filters(rows: list, cats, desigs, plants, elig, outcome, search):
    out = rows
    if cats:
        out = [r for r in out if r.get("category") in cats]
    if desigs:
        out = [r for r in out if r.get("designation") in desigs]
    if plants:
        out = [r for r in out if r.get("plant") in plants]
    if elig == "Yes":
        out = [r for r in out if r.get("incentive_eligible") == "Yes"]
    elif elig == "No":
        out = [r for r in out if r.get("incentive_eligible") == "No"]
    if outcome == "incentive":
        out = [r for r in out if (r.get("incentive_amount") or 0) > 0]
    elif outcome == "deduction":
        out = [r for r in out if (r.get("deduction_amount") or 0) > 0]
    elif outcome == "both":
        out = [r for r in out if (r.get("incentive_amount") or 0) > 0 and (r.get("deduction_amount") or 0) > 0]
    elif outcome == "neither":
        out = [r for r in out if (r.get("incentive_amount") or 0) == 0 and (r.get("deduction_amount") or 0) == 0]
    if search.strip():
        s = search.strip().lower()
        out = [r for r in out if s in str(r.get("employee_code", "")).lower()
               or s in str(r.get("employee_name", "")).lower()]
    return out


RESULT_COLS = [
    "employee_code", "employee_name", "designation",
    "plant", "plant_code",
    "total_quantity", "ytd_maintenance_cost",
    "incentive_eligible", "incentive_rate", "incentive_amount",
    "deduction_target", "shortfall_quantity", "deduction_amount",
    "remarks",
]
RESULT_LABELS = {
    "employee_code":        "Emp Code",
    "employee_name":        "Name",
    "designation":          "Designation",
    "plant":                "Plant",
    "plant_code":           "Plant Code",
    "total_quantity":       "Total Qty",
    "ytd_maintenance_cost": "R&M Cost YTD",
    "incentive_eligible":   "Eligible",
    "incentive_rate":       "Inc Rate",
    "incentive_amount":     "Incentive Amount (Rs)",
    "deduction_target":     "Ded Target",
    "shortfall_quantity":   "Shortfall",
    "deduction_amount":     "Deduction Amount (Rs)",
    "remarks":              "Remarks",
}

# Reports page gets two extra plant-aggregate columns
REPORT_COLS = RESULT_COLS + ["plant_total_incentive", "plant_total_deduction"]
REPORT_LABELS = {
    **RESULT_LABELS,
    "plant_total_incentive": "Plant Incentive (Rs)",
    "plant_total_deduction": "Plant Deduction (Rs)",
}
CAT_TABS = {
    "All Trainees":       ["Civil Trainee", "Non-Civil Trainee"],
    "Plant Mgr & PI":     ["PM & API"],
    "QCI":                ["QCI"],
    "All MO":             ["MO"],
    "SPE":                ["SPE"],
    "TL Employee":        ["TL BPO"],
    "Production Officer": ["Production Officer"],
    "Deduction NA":       ["Deduction NA"],
}


def _build_meta(from_date, to_date, filtered, unmapped, applied_filters):
    return {
        "generated_on":     _dt.now().isoformat(timespec="seconds"),
        "date_range":       f"{from_date} to {to_date}",
        "applied_filters":  applied_filters,
        "total_employees":  len(filtered),
        "total_quantity":   sum(r.get("total_quantity") or 0 for r in filtered),
        "total_incentive":  sum(r.get("incentive_amount") or 0 for r in filtered),
        "total_deduction":  sum(r.get("deduction_amount") or 0 for r in filtered),
        "unmapped_count":   len(unmapped),
        "validation_count": database.get_table_counts().get("validation_errors", 0),
    }


def _rows_to_df(rows: list, cols) -> pd.DataFrame:
    if not rows:
        return pd.DataFrame(columns=cols)
    return pd.DataFrame(rows)[list(cols)]


# ── PAGES ─────────────────────────────────────────────────────────────────────

# ── Auth routes ───────────────────────────────────────────────────────────────

@app.route("/login", methods=["GET", "POST"])
def page_login():
    return auth.login_view()


@app.route("/logout", methods=["POST"])
def page_logout():
    return auth.logout_view()


@app.route("/change-password", methods=["GET", "POST"])
@auth.login_required
def page_change_password():
    return auth.change_password_view()


# ── Admin / User Management routes (SUPER_ADMIN only) ─────────────────────────

@app.route("/admin/users")
@auth.uep_admin_required
def admin_users():
    caller = auth.get_current_user()
    users = database.get_all_users()
    # UEP_ADMIN sees only PLANT_USER accounts
    if caller and caller["role"] == auth.UEP_ADMIN:
        users = [u for u in users if u["role"] == auth.PLANT_USER]
    plant_map = {u["id"]: database.get_user_plant_access(u["id"]) for u in users}
    return render_template("admin/users.html",
                           active_page="users",
                           users=users,
                           plant_map=plant_map)


@app.route("/admin/users/create", methods=["GET", "POST"])
@auth.uep_admin_required
def admin_create_user():
    caller = auth.get_current_user()
    is_uep_admin = caller and caller["role"] == auth.UEP_ADMIN
    # UEP_ADMIN can only create PLANT_USER accounts
    allowed_roles = [auth.PLANT_USER] if is_uep_admin else auth.ALLOWED_ROLES
    all_plants = database.get_tp_plants()
    if request.method == "GET":
        return render_template("admin/user_form.html",
                               active_page="create",
                               edit_mode=False,
                               user={},
                               all_plants=all_plants,
                               assigned_plants=[],
                               assigned_plant_names=[],
                               allowed_roles=allowed_roles)
    # POST
    from werkzeug.security import generate_password_hash as _hash
    full_name  = request.form.get("full_name", "").strip()
    username   = request.form.get("username", "").strip()
    email      = request.form.get("email", "").strip()
    role       = request.form.get("role", "PLANT_USER")
    password   = request.form.get("password", "")
    is_active  = request.form.get("is_active", "1") == "1"
    must_chg   = request.form.get("must_change_password", "1") == "1"
    plant_names = request.form.getlist("plant_names")

    # Enforce UEP_ADMIN can only assign PLANT_USER role
    if is_uep_admin and role != auth.PLANT_USER:
        flash("You can only create PLANT_USER accounts.", "error")
        role = auth.PLANT_USER

    if not all([full_name, username, email, password]):
        flash("All required fields must be filled.", "error")
        return render_template("admin/user_form.html",
                               active_page="create",
                               edit_mode=False,
                               user=dict(full_name=full_name, username=username, email=email, role=role),
                               all_plants=all_plants,
                               assigned_plants=[],
                               assigned_plant_names=[],
                               allowed_roles=allowed_roles), 400
    if len(password) < 8:
        flash("Password must be at least 8 characters.", "error")
        return render_template("admin/user_form.html",
                               active_page="create",
                               edit_mode=False,
                               user=dict(full_name=full_name, username=username, email=email, role=role),
                               all_plants=all_plants,
                               assigned_plants=[],
                               assigned_plant_names=[],
                               allowed_roles=allowed_roles), 400
    try:
        user_id = database.create_user(
            full_name=full_name, email=email, username=username,
            password_hash=_hash(password),
            role=role, is_active=is_active, must_change_password=must_chg,
        )
        if role in auth.RESTRICTED_ROLES and plant_names:
            code_map = {p["plant_name"]: p["plant_code"] for p in all_plants}
            plant_list = [{"plant_name": n, "plant_code": code_map.get(n, "")} for n in plant_names]
            database.set_user_plants(user_id, plant_list)
        auth.log_activity(caller, "CREATE_USER",
                          details={"username": username, "role": role})
        flash(f"User '{username}' created successfully.", "success")
        return redirect(url_for("admin_users"))
    except Exception as exc:
        flash(f"Error creating user: {exc}", "error")
        return render_template("admin/user_form.html",
                               active_page="create",
                               edit_mode=False,
                               user=dict(full_name=full_name, username=username, email=email, role=role),
                               all_plants=all_plants,
                               assigned_plants=[],
                               assigned_plant_names=[],
                               allowed_roles=allowed_roles), 400


@app.route("/admin/users/<int:user_id>/edit", methods=["GET", "POST"])
@auth.uep_admin_required
def admin_edit_user(user_id):
    from werkzeug.security import generate_password_hash as _hash
    caller = auth.get_current_user()
    is_uep_admin = caller and caller["role"] == auth.UEP_ADMIN
    user = database.get_user_by_id(user_id)
    if not user:
        flash("User not found.", "error")
        return redirect(url_for("admin_users"))
    # UEP_ADMIN may only edit PLANT_USER accounts
    if is_uep_admin and user["role"] != auth.PLANT_USER:
        flash("You can only edit PLANT_USER accounts.", "error")
        return redirect(url_for("admin_users"))
    allowed_roles   = [auth.PLANT_USER] if is_uep_admin else auth.ALLOWED_ROLES
    all_plants      = database.get_tp_plants()
    assigned_plants = database.get_user_plant_access(user_id)
    assigned_names  = [p["plant_name"] for p in assigned_plants]

    if request.method == "GET":
        return render_template("admin/user_form.html",
                               active_page="users",
                               edit_mode=True,
                               user=user,
                               all_plants=all_plants,
                               assigned_plants=assigned_plants,
                               assigned_plant_names=assigned_names,
                               allowed_roles=allowed_roles)
    # POST
    full_name  = request.form.get("full_name", "").strip()
    email      = request.form.get("email", "").strip()
    role       = request.form.get("role", user["role"])
    is_active  = request.form.get("is_active", "1") == "1"
    must_chg   = request.form.get("must_change_password", "0") == "1"
    password   = request.form.get("password", "").strip()
    plant_names = request.form.getlist("plant_names")

    # Enforce UEP_ADMIN cannot change role away from PLANT_USER
    if is_uep_admin:
        role = auth.PLANT_USER

    database.update_user(user_id, full_name=full_name, email=email, role=role,
                         is_active=is_active, must_change_password=must_chg)
    if password:
        if len(password) < 8:
            flash("Password must be at least 8 characters.", "error")
            return redirect(url_for("admin_edit_user", user_id=user_id))
        database.update_user_password(user_id, _hash(password))

    code_map = {p["plant_name"]: p["plant_code"] for p in all_plants}
    plant_list = [{"plant_name": n, "plant_code": code_map.get(n, "")} for n in plant_names]
    database.set_user_plants(user_id, plant_list)

    auth.log_activity(caller, "EDIT_USER",
                      details={"user_id": user_id, "role": role})
    flash(f"User '{user['username']}' updated.", "success")
    return redirect(url_for("admin_users"))


@app.route("/admin/users/<int:user_id>/toggle", methods=["POST"])
@auth.uep_admin_required
def admin_toggle_user(user_id):
    current = auth.get_current_user()
    is_uep_admin = current and current["role"] == auth.UEP_ADMIN
    if current and current["id"] == user_id:
        flash("You cannot deactivate your own account.", "error")
        return redirect(url_for("admin_users"))
    user = database.get_user_by_id(user_id)
    if not user:
        flash("User not found.", "error")
        return redirect(url_for("admin_users"))
    if is_uep_admin and user["role"] != auth.PLANT_USER:
        flash("You can only manage PLANT_USER accounts.", "error")
        return redirect(url_for("admin_users"))
    database.update_user(user_id,
                         full_name=user["full_name"], email=user["email"],
                         role=user["role"],
                         is_active=not user["is_active"],
                         must_change_password=user["must_change_password"])
    status = "activated" if not user["is_active"] else "deactivated"
    auth.log_activity(current, "TOGGLE_USER",
                      details={"user_id": user_id, "new_status": status})
    flash(f"User '{user['username']}' {status}.", "success")
    return redirect(url_for("admin_users"))


@app.route("/admin/users/<int:user_id>/reset-password", methods=["GET", "POST"])
@auth.uep_admin_required
def admin_reset_password(user_id):
    from werkzeug.security import generate_password_hash as _hash
    caller = auth.get_current_user()
    is_uep_admin = caller and caller["role"] == auth.UEP_ADMIN
    target_user = database.get_user_by_id(user_id)
    if not target_user:
        flash("User not found.", "error")
        return redirect(url_for("admin_users"))
    if is_uep_admin and target_user["role"] != auth.PLANT_USER:
        flash("You can only reset passwords for PLANT_USER accounts.", "error")
        return redirect(url_for("admin_users"))
    if request.method == "GET":
        return render_template("admin/reset_password.html",
                               active_page="users",
                               target_user=target_user)
    new_pw  = request.form.get("new_password", "")
    conf_pw = request.form.get("confirm_password", "")
    must_chg = request.form.get("must_change", "1") == "1"
    if len(new_pw) < 8 or new_pw != conf_pw:
        flash("Passwords must match and be at least 8 characters.", "error")
        return redirect(url_for("admin_reset_password", user_id=user_id))
    database.update_user_password(user_id, _hash(new_pw), must_change_password=must_chg)
    auth.log_activity(auth.get_current_user(), "RESET_PASSWORD",
                      details={"target_user_id": user_id})
    flash(f"Password for '{target_user['username']}' has been reset.", "success")
    return redirect(url_for("admin_edit_user", user_id=user_id))


@app.route("/admin/audit-log")
@auth.admin_required
def admin_audit_log():
    rows = database.get_login_audit_log(limit=500)
    return render_template("admin/audit_log.html",
                           active_page="audit",
                           rows=rows)


@app.route("/admin/activity-log")
@auth.admin_required
def admin_activity_log():
    rows = database.get_user_activity_log(limit=500)
    return render_template("admin/activity_log.html",
                           active_page="activity",
                           rows=rows)


# ── Home page ─────────────────────────────────────────────────────────────────

@app.route("/")
@auth.login_required
def page_home():
    user = auth.get_current_user()
    # SUPER_ADMIN gets the module launcher as before
    if user and user.get("role") == "SUPER_ADMIN":
        resp = make_response(render_template("home.html"))
        resp.headers["Cache-Control"] = "no-cache, no-store, must-revalidate"
        resp.headers["Pragma"] = "no-cache"
        resp.headers["Expires"] = "0"
        return resp

    # All other roles get the unified user dashboard
    from datetime import date as _today
    MONTHS = ["January","February","March","April","May","June",
              "July","August","September","October","November","December"]
    def _mn(m): return MONTHS[m-1] if m else "—"

    import calendar as _cal
    now       = _today.today()
    cur_ym    = f"{now.year}-{now.month:02d}"
    # previous calendar month
    if now.month == 1:
        prev_m, prev_y = 12, now.year - 1
    else:
        prev_m, prev_y = now.month - 1, now.year
    prev_ym = f"{prev_y}-{prev_m:02d}"

    def _date_range(month, year):
        """Return (from_str, to_str) for a month. Current month ends today."""
        fd = f"{year}-{month:02d}-01"
        if month == now.month and year == now.year:
            td = now.strftime("%Y-%m-%d")
        else:
            last_day = _cal.monthrange(year, month)[1]
            td = f"{year}-{month:02d}-{last_day:02d}"
        return fd, td

    conn = database.get_connection()
    # defaults — overwritten below as each section runs
    tp_last_known_sync = bt_last_known_sync = None
    try:
        cur = conn.cursor()

        # Allowed plants for this user (empty list = all plants for global roles)
        allowed_plants  = auth.get_user_allowed_plants(user)
        user_plant_rows = database.get_user_plant_access(user["id"]) if allowed_plants else []
        is_restricted   = bool(allowed_plants)  # True for REGIONAL_USER / PLANT_USER
        _allowed_codes  = []  # tp_plant_data plant_codes for restricted user (filled below)

        def _plant_filter(rows, col="plant"):
            if not is_restricted:
                return rows
            al = [p.lower() for p in allowed_plants]
            return [r for r in rows if str(r.get(col, "")).lower() in al]

        # ── I&D — fetch up to 2 calculated months ────────────
        cur.execute("""SELECT month, year,
            COUNT(*) as total_emp,
            SUM(CASE WHEN incentive_eligible='Yes' THEN 1 ELSE 0 END) as inc_emp,
            SUM(CASE WHEN deduction_amount > 0   THEN 1 ELSE 0 END) as ded_emp,
            ROUND(SUM(COALESCE(incentive_amount,0)),0) as total_inc,
            ROUND(SUM(COALESCE(deduction_amount,0)),0) as total_ded
            FROM calculation_results GROUP BY year,month
            ORDER BY year DESC, month DESC LIMIT 2""")
        id_months = []
        for row in cur.fetchall():
            d = dict(row); d["month_name"] = _mn(d["month"])
            if not (d["month"] == now.month and d["year"] == now.year):
                d["from_date"], d["to_date"] = _date_range(d["month"], d["year"])
                id_months.append(d)
        id_last = id_months[0] if id_months else None

        # Detail rows for each I&D month
        id_months_rows = []
        for m in id_months:
            cur.execute("""SELECT employee_code, employee_name, designation, plant,
                total_quantity, incentive_eligible, incentive_rate,
                incentive_amount, deduction_target, shortfall_quantity, deduction_amount, remarks
                FROM calculation_results WHERE month=? AND year=?
                ORDER BY plant, employee_name""",
                (m["month"], m["year"]))
            cols = [d[0] for d in cur.description]
            rows_f = _plant_filter([dict(zip(cols, r)) for r in cur.fetchall()], col="plant")
            if is_restricted:
                m["total_emp"] = len(rows_f)
                m["inc_emp"]   = sum(1 for r in rows_f if r.get("incentive_amount") and r["incentive_amount"] > 0)
                m["ded_emp"]   = sum(1 for r in rows_f if r.get("deduction_amount") and r["deduction_amount"] > 0)
                m["total_inc"] = sum(r.get("incentive_amount") or 0 for r in rows_f)
            id_months_rows.append(rows_f)
        id_last_rows = id_months_rows[0] if id_months_rows else []

        # I&D current month — from calculation_results (same as previous month card)
        cur.execute("""SELECT COUNT(*) as total_emp,
            SUM(CASE WHEN incentive_eligible='Yes' THEN 1 ELSE 0 END) as inc_emp,
            SUM(CASE WHEN deduction_amount > 0   THEN 1 ELSE 0 END) as ded_emp,
            ROUND(SUM(COALESCE(incentive_amount,0)),0) as total_inc,
            ROUND(SUM(COALESCE(deduction_amount,0)),0) as total_ded
            FROM calculation_results WHERE month=? AND year=?""",
            (now.month, now.year))
        row = cur.fetchone()
        id_cur = dict(row) if (row and row[0]) else None
        if id_cur:
            id_cur["month_name"] = _mn(now.month)
            id_cur["month"] = now.month
            id_cur["year"]  = now.year
            id_cur["from_date"], id_cur["to_date"] = _date_range(now.month, now.year)
        # Detail rows for current month I&D
        id_cur_rows = []
        if id_cur:
            cur.execute("""SELECT employee_code, employee_name, designation, plant,
                total_quantity, incentive_eligible, incentive_rate,
                incentive_amount, deduction_target, shortfall_quantity, deduction_amount, remarks
                FROM calculation_results WHERE month=? AND year=?
                ORDER BY plant, employee_name""", (now.month, now.year))
            cols = [d[0] for d in cur.description]
            id_cur_rows = _plant_filter([dict(zip(cols, r)) for r in cur.fetchall()], col="plant")
            if is_restricted:
                id_cur["total_emp"] = len(id_cur_rows)
                id_cur["inc_emp"]   = sum(1 for r in id_cur_rows if r.get("incentive_amount") and r["incentive_amount"] > 0)
                id_cur["ded_emp"]   = sum(1 for r in id_cur_rows if r.get("deduction_amount") and r["deduction_amount"] > 0)
                id_cur["total_inc"] = sum(r.get("incentive_amount") or 0 for r in id_cur_rows)

        # ── TP — fetch up to 2 calculated months ─────────────
        cur.execute("""SELECT month, year,
            COUNT(DISTINCT lookup_code) as plants,
            ROUND(AVG(throughput_pct),1) as avg_tp,
            SUM(CASE WHEN throughput_pct >= 75 THEN 1 ELSE 0 END) as above_target,
            SUM(CASE WHEN throughput_pct <  75 THEN 1 ELSE 0 END) as below_target,
            ROUND(SUM(total_quantity),0) as total_qty
            FROM tp_results GROUP BY year,month
            ORDER BY year DESC, month DESC LIMIT 2""")
        tp_months = []
        for row in cur.fetchall():
            d = dict(row); d["month_name"] = _mn(d["month"])
            if not (d["month"] == now.month and d["year"] == now.year):
                d["from_date"], d["to_date"] = _date_range(d["month"], d["year"])
                tp_months.append(d)
        tp_last = tp_months[0] if tp_months else None

        tp_months_rows = []
        for m in tp_months:
            cur.execute("""SELECT plant_name, exco_location, business_head,
                plant_manager, total_quantity, total_time_hrs, mixer_theo_cap, throughput_pct, batch_count
                FROM tp_results WHERE month=? AND year=?
                ORDER BY throughput_pct DESC""",
                (m["month"], m["year"]))
            cols = [d[0] for d in cur.description]
            rows_f = _plant_filter([dict(zip(cols, r)) for r in cur.fetchall()], col="plant_name")
            if is_restricted:
                m["plants"]       = len(rows_f)
                m["above_target"] = sum(1 for r in rows_f if (r.get("throughput_pct") or 0) >= 75)
                m["below_target"] = sum(1 for r in rows_f if (r.get("throughput_pct") or 0) < 75)
                m["avg_tp"]       = round(sum(r.get("throughput_pct") or 0 for r in rows_f) / len(rows_f), 1) if rows_f else 0
            tp_months_rows.append(rows_f)
        tp_last_rows = tp_months_rows[0] if tp_months_rows else []

        # TP current month — from tp_results (same as previous month card)
        cur.execute("""SELECT COUNT(DISTINCT lookup_code) as plants,
            ROUND(AVG(throughput_pct),1) as avg_tp,
            SUM(CASE WHEN throughput_pct >= 75 THEN 1 ELSE 0 END) as above_target,
            SUM(CASE WHEN throughput_pct <  75 THEN 1 ELSE 0 END) as below_target,
            ROUND(SUM(total_quantity),0) as total_qty
            FROM tp_results WHERE month=? AND year=?""", (now.month, now.year))
        row = cur.fetchone()
        tp_cur = dict(row) if (row and row[0]) else None
        if tp_cur:
            tp_cur["month_name"] = _mn(now.month)
            tp_cur["month"] = now.month
            tp_cur["year"]  = now.year
            tp_cur["from_date"], tp_cur["to_date"] = _date_range(now.month, now.year)
        tp_last_known_sync = None
        # Detail rows for current month TP
        tp_cur_rows = []
        if tp_cur:
            cur.execute("""SELECT plant_name, exco_location, business_head,
                plant_manager, total_quantity, total_time_hrs, mixer_theo_cap, throughput_pct, batch_count
                FROM tp_results WHERE month=? AND year=?
                ORDER BY throughput_pct DESC""", (now.month, now.year))
            cols = [d[0] for d in cur.description]
            tp_cur_rows = _plant_filter([dict(zip(cols, r)) for r in cur.fetchall()], col="plant_name")
            if is_restricted:
                tp_cur["plants"]       = len(tp_cur_rows)
                tp_cur["above_target"] = sum(1 for r in tp_cur_rows if (r.get("throughput_pct") or 0) >= 75)
                tp_cur["below_target"] = sum(1 for r in tp_cur_rows if (r.get("throughput_pct") or 0) < 75)
                tp_cur["avg_tp"]       = round(sum(r.get("throughput_pct") or 0 for r in tp_cur_rows) / len(tp_cur_rows), 1) if tp_cur_rows else 0

        # ── BTRTP — fetch up to 2 calculated months ──────────
        cur.execute("""SELECT month, year,
            COUNT(*) as batchers,
            ROUND(SUM(throughput_pct * total_quantity) / NULLIF(SUM(total_quantity),0), 1) as avg_tp,
            SUM(CASE WHEN throughput_pct >= 75 THEN 1 ELSE 0 END) as above_target,
            SUM(CASE WHEN throughput_pct <  75 THEN 1 ELSE 0 END) as below_target
            FROM btrtp_results GROUP BY year,month
            ORDER BY year DESC, month DESC LIMIT 2""")
        bt_months = []
        for row in cur.fetchall():
            d = dict(row); d["month_name"] = _mn(d["month"])
            if not (d["month"] == now.month and d["year"] == now.year):
                d["from_date"], d["to_date"] = _date_range(d["month"], d["year"])
                bt_months.append(d)
        bt_last = bt_months[0] if bt_months else None

        bt_months_rows = []
        for m in bt_months:
            cur.execute("""SELECT batcher_name, batcher_id, plant_name, exco_location,
                total_quantity, total_time_hrs, mixer_theo_cap, throughput_pct, batch_count
                FROM btrtp_results WHERE month=? AND year=?
                ORDER BY throughput_pct DESC""",
                (m["month"], m["year"]))
            cols = [d[0] for d in cur.description]
            rows_f = _plant_filter([dict(zip(cols, r)) for r in cur.fetchall()], col="plant_name")
            if is_restricted:
                m["batchers"]     = len(rows_f)
                m["above_target"] = sum(1 for r in rows_f if (r.get("throughput_pct") or 0) >= 75)
                m["below_target"] = sum(1 for r in rows_f if (r.get("throughput_pct") or 0) < 75)
                total_qty_bt = sum(r.get("total_quantity") or 0 for r in rows_f)
                m["avg_tp"] = round(sum((r.get("throughput_pct") or 0) * (r.get("total_quantity") or 0) for r in rows_f) / total_qty_bt, 1) if total_qty_bt else 0
            bt_months_rows.append(rows_f)
        bt_last_rows = bt_months_rows[0] if bt_months_rows else []

        # BTRTP current month — from btrtp_results (same as previous month card)
        bt_last_known_sync = None
        cur.execute("""SELECT COUNT(*) as batchers,
            ROUND(SUM(throughput_pct * total_quantity) / NULLIF(SUM(total_quantity),0), 1) as avg_tp,
            SUM(CASE WHEN throughput_pct >= 75 THEN 1 ELSE 0 END) as above_target,
            SUM(CASE WHEN throughput_pct <  75 THEN 1 ELSE 0 END) as below_target
            FROM btrtp_results WHERE month=? AND year=?""", (now.month, now.year))
        row = cur.fetchone()
        bt_cur = dict(row) if (row and row[0]) else None
        if bt_cur:
            bt_cur["month_name"] = _mn(now.month)
            bt_cur["month"] = now.month
            bt_cur["year"]  = now.year
            bt_cur["from_date"], bt_cur["to_date"] = _date_range(now.month, now.year)
        # Detail rows for current month BTRTP
        bt_cur_rows = []
        if bt_cur:
            cur.execute("""SELECT batcher_name, batcher_id, plant_name, exco_location,
                total_quantity, total_time_hrs, mixer_theo_cap, throughput_pct, batch_count
                FROM btrtp_results WHERE month=? AND year=?
                ORDER BY throughput_pct DESC""", (now.month, now.year))
            cols = [d[0] for d in cur.description]
            bt_cur_rows = _plant_filter([dict(zip(cols, r)) for r in cur.fetchall()], col="plant_name")
            if is_restricted:
                bt_cur["batchers"]     = len(bt_cur_rows)
                bt_cur["above_target"] = sum(1 for r in bt_cur_rows if (r.get("throughput_pct") or 0) >= 75)
                bt_cur["below_target"] = sum(1 for r in bt_cur_rows if (r.get("throughput_pct") or 0) < 75)
                total_qty_btc = sum(r.get("total_quantity") or 0 for r in bt_cur_rows)
                bt_cur["avg_tp"] = round(sum((r.get("throughput_pct") or 0) * (r.get("total_quantity") or 0) for r in bt_cur_rows) / total_qty_btc, 1) if total_qty_btc else 0

        # ── ECMD — fetch up to 2 calculated months ───────────
        cur.execute("""SELECT month, year,
            COUNT(*) as plants,
            ROUND(AVG(energy_per_mt),2) as avg_energy,
            ROUND(AVG(mixer_dg_ratio),1) as avg_dg
            FROM ecmd_results GROUP BY year,month
            ORDER BY year DESC, month DESC LIMIT 2""")
        ec_months = []
        for row in cur.fetchall():
            d = dict(row); d["month_name"] = _mn(d["month"])
            if not (d["month"] == now.month and d["year"] == now.year):
                d["from_date"], d["to_date"] = _date_range(d["month"], d["year"])
                ec_months.append(d)
        ec_last = ec_months[0] if ec_months else None

        ec_months_rows = []
        for m in ec_months:
            cur.execute("""SELECT plant_name, exco_location, plant_manager,
                eb_kwh, dg_kwh, total_kwh, total_volume, energy_per_mt,
                mixer_dg_ratio, diesel_issued_ltrs
                FROM ecmd_results WHERE month=? AND year=?
                ORDER BY plant_name""",
                (m["month"], m["year"]))
            cols = [d[0] for d in cur.description]
            rows_f = _plant_filter([dict(zip(cols, r)) for r in cur.fetchall()], col="plant_name")
            if is_restricted:
                m["plants"] = len(rows_f)
            ec_months_rows.append(rows_f)
        ec_last_rows = ec_months_rows[0] if ec_months_rows else []

        # ECMD current month — from ecmd_results (same as previous month card)
        cur.execute("""SELECT COUNT(*) as plants,
            ROUND(AVG(energy_per_mt),2) as avg_energy,
            ROUND(AVG(mixer_dg_ratio),1) as avg_dg
            FROM ecmd_results WHERE month=? AND year=?""", (now.month, now.year))
        row = cur.fetchone()
        ec_cur = dict(row) if (row and row[0]) else None
        if ec_cur:
            ec_cur["month_name"] = _mn(now.month)
            ec_cur["month"] = now.month
            ec_cur["year"]  = now.year
            ec_cur["from_date"], ec_cur["to_date"] = _date_range(now.month, now.year)
        # Detail rows for current month ECMD
        ec_cur_rows = []
        if ec_cur:
            cur.execute("""SELECT plant_name, exco_location, plant_manager,
                eb_kwh, dg_kwh, total_kwh, total_volume, energy_per_mt,
                mixer_dg_ratio, diesel_issued_ltrs
                FROM ecmd_results WHERE month=? AND year=?
                ORDER BY plant_name""", (now.month, now.year))
            cols = [d[0] for d in cur.description]
            ec_cur_rows = _plant_filter([dict(zip(cols, r)) for r in cur.fetchall()], col="plant_name")
            if is_restricted:
                ec_cur["plants"] = len(ec_cur_rows)

    finally:
        conn.close()

    # ECMD entry window
    entry_month = int(database.get_setting("ecmd_entry_open_month", 0) or 0)
    entry_year  = int(database.get_setting("ecmd_entry_open_year",  0) or 0)
    entry_open  = bool(entry_month and entry_year)

    resp = make_response(render_template("user_dashboard.html",
        id_last=id_last, id_last_rows=id_last_rows, id_cur=id_cur, id_cur_rows=id_cur_rows,
        id_months=id_months, id_months_rows=id_months_rows,
        tp_last=tp_last, tp_last_rows=tp_last_rows, tp_cur=tp_cur, tp_cur_rows=tp_cur_rows,
        tp_months=tp_months, tp_months_rows=tp_months_rows,
        bt_last=bt_last, bt_last_rows=bt_last_rows, bt_cur=bt_cur, bt_cur_rows=bt_cur_rows,
        bt_months=bt_months, bt_months_rows=bt_months_rows,
        ec_last=ec_last, ec_last_rows=ec_last_rows, ec_cur=ec_cur, ec_cur_rows=ec_cur_rows,
        ec_months=ec_months, ec_months_rows=ec_months_rows,
        cur_month_name=_mn(now.month), cur_year=now.year,
        user_plant_rows=user_plant_rows,
        tp_last_known_sync=tp_last_known_sync,
        bt_last_known_sync=bt_last_known_sync,
        entry_open=entry_open,
        entry_month_name=_mn(entry_month) if entry_open else "",
        entry_year=entry_year,
        bg_auto=bg_auto(), bg_animate=bg_animate(), bg_theme=bg_theme(),
    ))
    resp.headers["Cache-Control"] = "no-cache, no-store, must-revalidate"
    resp.headers["Pragma"] = "no-cache"
    resp.headers["Expires"] = "0"
    return resp

@app.route("/id")
def page_id_dashboard():
    return redirect(url_for("page_dashboard"))

@app.route("/dashboard")
def page_dashboard():
    _ss("active_page", "dashboard")
    counts = database.get_table_counts()
    return render_template("dashboard.html",
                           counts=counts,
                           db_path=database.DB_PATH)

# ═══════════════════════════════════════════════════════════════════════════
# RDC-TP MODULE — Plant Throughput Calculator
# ═══════════════════════════════════════════════════════════════════════════

def _tp_ctx():
    """Base context dict for every TP page."""
    return dict(
        active_page="",
        bg_auto=bg_auto(), bg_animate=bg_animate(), bg_theme=bg_theme(),
    )


@app.route("/tp")
def page_tp():
    return redirect(url_for("tp_dashboard"))


# ── Dashboard ─────────────────────────────────────────────────────────────────
@app.route("/tp/dashboard")
def tp_dashboard():
    all_counts = database.get_table_counts()
    counts = {
        "tp_plant_data":  all_counts.get("tp_plant_data", 0),
        "tp_oracle_data": all_counts.get("oracle_raw_data", 0),
        "tp_results":     all_counts.get("tp_results", 0),
    }
    last_sync  = google_sheets.get_tp_last_sync_info()
    ora_ready  = oracle_connector.is_configured()
    ctx = _tp_ctx()
    ctx["active_page"] = "dashboard"
    return render_template("tp_dashboard.html", counts=counts,
                           last_sync=last_sync, ora_ready=ora_ready, **ctx)


# ── Data Uploader ─────────────────────────────────────────────────────────────
@app.route("/tp/data-uploader")
def tp_data_uploader():
    ora_ready    = oracle_connector.is_configured()
    last_sync    = google_sheets.get_tp_last_sync_info()
    sheet_id     = database.get_module_setting("tp", "gsheet_id",
                                               database.get_setting("gsheet_id", ""))
    tp_worksheet = database.get_module_setting("tp", "gsheet_worksheet", "Plant Data for TP")
    ora_counts   = database.get_table_counts().get("oracle_raw_data", 0)

    plant_rows  = database.get_tp_plants()
    plant_count = len(plant_rows)
    tp_codes    = database.get_tp_plant_codes()
    log_rows    = database.get_tp_plant_log()

    edit_code  = request.args.get("edit_code", "").strip()
    del_code   = request.args.get("del_code", "").strip()
    edit_plant = database.get_tp_plant(edit_code) if edit_code else None
    del_plant  = database.get_tp_plant(del_code)  if del_code  else None

    ctx = _tp_ctx()
    ctx["active_page"] = "data_uploader"
    return render_template("tp_data_uploader.html",
                           ora_ready=ora_ready, last_sync=last_sync,
                           sheet_id=sheet_id, tp_worksheet=tp_worksheet,
                           ora_counts=ora_counts,
                           plant_rows=plant_rows, plant_count=plant_count,
                           tp_codes=tp_codes, log_rows=log_rows,
                           edit_plant=edit_plant, del_plant=del_plant,
                           **ctx)


@app.route("/tp/action/add-plant", methods=["POST"])
def tp_add_plant():
    plant_data = {
        "plant_code":     request.form.get("plant_code", ""),
        "exco_location":  request.form.get("exco_location", ""),
        "plant_name":     request.form.get("plant_name", ""),
        "business_head":  request.form.get("business_head", ""),
        "plant_manager":  request.form.get("plant_manager", ""),
        "mixer_theo_cap": request.form.get("mixer_theo_cap", 0),
    }
    try:
        database.add_tp_plant(**plant_data)
        flash("✅ Plant added successfully.", "success")
        res = google_sheets.push_tp_plant_add(plant_data)
        if res["ok"]:
            flash(f"☁️ {res['message']}", "info")
        else:
            flash(f"⚠️ Saved locally but Google Sheet not updated: {res['message']}", "warning")
    except ValueError as exc:
        flash(str(exc), "error")
    return redirect(url_for("tp_data_uploader") + "?open_modal=plant&tab=add")


@app.route("/tp/action/update-plant/<code>", methods=["POST"])
def tp_update_plant(code):
    plant_data = {
        "plant_code":     code,
        "exco_location":  request.form.get("exco_location", ""),
        "plant_name":     request.form.get("plant_name", ""),
        "business_head":  request.form.get("business_head", ""),
        "plant_manager":  request.form.get("plant_manager", ""),
        "mixer_theo_cap": request.form.get("mixer_theo_cap", 0),
    }
    try:
        database.update_tp_plant(**plant_data)
        flash("✅ Plant updated successfully.", "success")
        res = google_sheets.push_tp_plant_update(plant_data)
        if res["ok"]:
            flash(f"☁️ {res['message']}", "info")
        else:
            flash(f"⚠️ Saved locally but Google Sheet not updated: {res['message']}", "warning")
    except ValueError as exc:
        flash(str(exc), "error")
    return redirect(url_for("tp_data_uploader") + "?open_modal=plant&tab=edit")


@app.route("/tp/action/delete-plant/<code>", methods=["POST"])
def tp_delete_plant(code):
    try:
        database.delete_tp_plant(code)
        flash(f"✅ Plant '{code}' deleted.", "success")
        res = google_sheets.push_tp_plant_delete(code)
        if res["ok"]:
            flash(f"☁️ {res['message']}", "info")
        else:
            flash(f"⚠️ Deleted locally but Google Sheet not updated: {res['message']}", "warning")
    except ValueError as exc:
        flash(str(exc), "error")
    return redirect(url_for("tp_data_uploader") + "?open_modal=plant&tab=delete")


@app.route("/tp/action/fetch-oracle", methods=["POST"])
def tp_fetch_oracle():
    from_date = request.form.get("from_date", "")
    to_date   = request.form.get("to_date", "")
    if not from_date or not to_date:
        flash("Please select both From and To dates.", "error")
        return jsonify({"ok": False, "redirect": url_for("tp_data_uploader")})
    try:
        _set_progress(20, "Fetching TP data from Oracle…")
        raw_df, ora_warnings = oracle_connector.fetch_tp_data(from_date, to_date)
        _set_progress(55, "Parsing plant records…")
        parsed, skip_log     = tp_calculator.parse_oracle_df(raw_df)
        _set_progress(75, "Saving to database…")
        oracle_connector.save_tp_oracle_data(raw_df, from_date, to_date, parsed, replace=True)
        _set_progress(90, "Cleaning old records…")
        database.purge_old_oracle_data()
        _mss("tp", "skip_log", skip_log)
        _mss("tp", "ora_from", from_date)
        _mss("tp", "ora_to",   to_date)
        for w in ora_warnings:
            flash(w, "warning")
        flash(f"✅ {len(parsed)} rows fetched & saved ({len(skip_log)} skipped).", "success")
    except Exception as exc:
        flash(f"Oracle error: {exc}", "error")
    _set_progress(100, "Complete")
    return jsonify({"ok": True, "redirect": url_for("tp_data_uploader")})


@app.route("/tp/action/sync-sheets", methods=["POST"])
def tp_sync_sheets():
    sheet_id    = (request.form.get("sheet_id", "").strip()
                   or database.get_module_setting("tp", "gsheet_id",
                                                   database.get_setting("gsheet_id", "")))
    worksheet   = request.form.get("worksheet", "Plant Data for TP").strip()
    _set_progress(20, "Syncing plant data from Google Sheets…")
    result      = google_sheets.sync_tp_plant_data(sheet_id, worksheet)
    _set_progress(85, "Saving settings…")
    if result["error"]:
        flash(f"Sync failed: {result['error']}", "error")
    else:
        database.set_module_setting("tp", "gsheet_id", google_sheets.extract_sheet_id(sheet_id))
        database.set_module_setting("tp", "gsheet_worksheet", worksheet)
        flash(f"✅ {result['rows_synced']} plant rows synced from Google Sheets ({result['mode']} mode).", "success")
    _set_progress(100, "Complete")
    if request.headers.get("X-Requested-With") == "XMLHttpRequest":
        return jsonify({"ok": True, "redirect": url_for("tp_data_uploader")})
    return redirect(url_for("sysconfig_page", open_modal="plant-data"))


# ── Calculate ─────────────────────────────────────────────────────────────────
@app.route("/tp/calculate", methods=["GET", "POST"])
def tp_calculate():
    # Unified page: calculate → redirect to reports with dates in query params
    today = _date.today()
    if request.method == "POST":
        from_s = request.form.get("from_date", str(today.replace(day=1)))
        to_s   = request.form.get("to_date",   str(today))
        return redirect(url_for("tp_reports", from_date=from_s, to_date=to_s))
    return redirect(url_for("tp_reports",
                            from_date=str(today.replace(day=1)),
                            to_date=str(today)))


# ── Reports ───────────────────────────────────────────────────────────────────
def _tp_band(pct):
    """Throughput colour band for a percentage."""
    if pct < 60:
        return "red"
    if pct < 75:
        return "yellow"
    return "green"


def _apply_tp_filters(plant_rows, excos, bheads, plants, band, search):
    """Filter plant rows by Exco Location, Business Head, Plant, throughput band
    and a free-text search over plant code / plant name."""
    out = []
    s = (search or "").strip().lower()
    for r in plant_rows:
        if excos and r.get("exco_location") not in excos:
            continue
        if bheads and r.get("business_head") not in bheads:
            continue
        if plants and r.get("plant_name") not in plants:
            continue
        if band and band != "All" and _tp_band(r.get("throughput_pct", 0)) != band:
            continue
        if s and s not in str(r.get("lookup_code", "")).lower() \
             and s not in str(r.get("plant_name", "")).lower():
            continue
        out.append(r)
    return out


@app.route("/tp/reports")
def tp_reports():
    today = _date.today()

    # Date range — default to 1st of current month → today.
    from_s = request.args.get("from_date", str(today.replace(day=1)))
    to_s   = request.args.get("to_date",   str(today))
    try:
        fd = _date.fromisoformat(from_s)
        td = _date.fromisoformat(to_s)
    except ValueError:
        fd, td = today.replace(day=1), today

    ora_note = None

    # Re-run the calculation for the chosen range when the user loads a report,
    # otherwise fall back to the rows from the last Calculate run.
    if "from_date" in request.args:
        month, year = fd.month, fd.year

        # Auto-fetch Oracle data for this range in-memory (no DB write, no purge)
        # so historical date ranges are not blocked by the 2-month retention window.
        ora_df_live = None
        if oracle_connector.is_configured():
            try:
                raw_df, ora_warns = oracle_connector.fetch_tp_data(str(fd), str(td))
                if not raw_df.empty:
                    parsed, skip_log = tp_calculator.parse_oracle_df(raw_df)
                    _mss("tp", "skip_log", skip_log)
                    import pandas as _pd
                    ora_df_live = _pd.DataFrame(parsed)
                    ora_note = f"Oracle: {len(raw_df):,} rows loaded for {fd} → {td}."
                else:
                    ora_note = "Oracle returned no rows for this range."
                    _mss("tp", "skip_log", [])
            except Exception as exc:
                ora_note = f"Oracle fetch failed: {exc}"

        all_plants, all_locs, calc_warns = tp_calculator.run_tp_calculation(
            month, year, from_date=str(fd), to_date=str(td), ora_df=ora_df_live)
        _mss("tp", "calc_warnings", calc_warns)
        _mss("tp", "plant_rows", all_plants)
        _mss("tp", "location_rows", all_locs)
        _mss("tp", "calc_from", str(fd))
        _mss("tp", "calc_to", str(td))
        _mss("tp", "calc_month", month)
        _mss("tp", "calc_year", year)
    else:
        all_plants = _ms("tp", "plant_rows", [])
        all_locs   = _ms("tp", "location_rows", [])

    # Filter inputs
    excos  = request.args.getlist("exco")
    bheads = request.args.getlist("bhead")
    plants = request.args.getlist("plant")
    band   = request.args.get("band", "All")
    search = request.args.get("search", "")

    # Distinct filter options (from the full, unfiltered set)
    unique_excos  = sorted({r.get("exco_location", "") for r in all_plants if r.get("exco_location")})
    unique_bheads = sorted({r.get("business_head", "") for r in all_plants if r.get("business_head")})
    unique_plants = sorted({r.get("plant_name", "")    for r in all_plants if r.get("plant_name")})

    # Apply filters → plant rows, then rebuild location rows from the filtered set
    plant_rows = _apply_tp_filters(all_plants, excos, bheads, plants, band, search)
    location_rows = tp_calculator.build_location_rows(plant_rows, fd.month, fd.year)

    # Keep a filtered snapshot for downloads / email
    _mss("tp", "report_plant_rows", plant_rows)
    _mss("tp", "report_location_rows", location_rows)

    smtp        = email_helper.get_smtp_config()
    email_ready = email_helper.is_configured()

    # TP-only email log (filter shared table by RDC_TP_ filename prefix)
    _elog = database.read_table("email_log", order_by="id DESC")
    if not _elog.empty and "report_file_name" in _elog.columns:
        _elog = _elog[_elog["report_file_name"].str.startswith("RDC_TP_", na=False)]
    tp_email_log = _records(_elog.drop(columns=["id"], errors="ignore").head(20)) \
        if not _elog.empty else []

    if (fd.year, fd.month) == (td.year, td.month):
        import calendar
        month_label = f"{calendar.month_name[fd.month]} {fd.year}"
    else:
        month_label = f"{fd:%d %b %Y} → {td:%d %b %Y}"

    # TP-specific email defaults (set in TP Settings → Email)
    default_to      = database.get_module_setting("tp", "email_default_to", "") or smtp.get("default_to", "")
    default_cc      = database.get_module_setting("tp", "email_default_cc", "") or smtp.get("default_cc", "")
    default_subject = (database.get_module_setting("tp", "email_default_subject", "")
                       or f"RDC-TP Plant Throughput Report — {month_label}")
    default_body    = (database.get_module_setting("tp", "email_default_body", "")
                       or f"Dear Team,\n\nPlease find attached the Plant Throughput "
                          f"Report for {month_label}.\n\nRegards,\nRDC Operations")

    ctx = _tp_ctx()
    ctx["active_page"] = "reports"
    return render_template("tp_reports.html",
                           plant_rows=plant_rows, location_rows=location_rows,
                           total_plants=len(all_plants),
                           from_date=str(fd), to_date=str(td),
                           unique_excos=unique_excos, unique_bheads=unique_bheads,
                           unique_plants=unique_plants,
                           excos=excos, bheads=bheads, plants=plants,
                           band=band, search=search,
                           month_label=month_label, email_cfg=smtp,
                           email_ready=email_ready,
                           default_to=default_to, default_cc=default_cc,
                           default_subject=default_subject,
                           default_body=default_body,
                           tp_email_log=tp_email_log,
                           ora_note=ora_note, **ctx)


def _tp_mon_tag(month, year):
    import calendar as _cal
    return f"{_cal.month_abbr[month]}'{str(year)[2:]}"


def _tp_build_excel(plant_rows, location_rows, month, year):
    """Return Excel bytes matching exactly the HTML email tables (same headers, columns, colors)."""
    from openpyxl import Workbook
    from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    mon_tag = _tp_mon_tag(month, year)

    # Colors matching the HTML email exactly
    RED_FILL   = PatternFill("solid", fgColor="FFB3B3")
    YEL_FILL   = PatternFill("solid", fgColor="FFE066")
    GRN_FILL   = PatternFill("solid", fgColor="92D492")
    PAN_FILL   = PatternFill("solid", fgColor="D9D9D9")
    HDR_FILL   = PatternFill("solid", fgColor="0A2540")
    PLAIN_FILL = PatternFill("solid", fgColor="FFFFFF")

    # Font colors complementing each background (matching HTML _fg())
    RED_FONT   = Font(color="7B1F1F", size=10)
    YEL_FONT   = Font(color="5C4200", size=10)
    GRN_FONT   = Font(color="1A5C1A", size=10)
    PAN_FONT   = Font(bold=True, color="222222", size=10)
    HDR_FONT   = Font(bold=True, color="FFFFFF", size=10)
    TITLE_FONT = Font(bold=True, color="FFFFFF", size=11)
    PLAIN_FONT = Font(color="19263A", size=10)

    CTR   = Alignment(vertical="center", horizontal="center")
    LEFT  = Alignment(vertical="center", horizontal="left")
    WRAP  = Alignment(wrap_text=True, vertical="center", horizontal="center")
    thin  = Side(style="thin", color="9A9A9A")
    BDR   = Border(left=thin, right=thin, top=thin, bottom=thin)

    def _fill_font(pct):
        if pct < 60:  return RED_FILL, RED_FONT
        if pct < 75:  return YEL_FILL, YEL_FONT
        return GRN_FILL, GRN_FONT

    wb = Workbook()

    # ── Location sheet — Sr. no., Exco Location, Plants, Total Qty, Time (min), Avg TP %
    ws_l = wb.active
    ws_l.title = "Location Throughput"
    LOC_HEADS = ["Sr. no.", "Exco Location", "Plants", "Total Qty", "Time (min)", "Avg TP %"]
    ws_l.merge_cells(f"A1:{get_column_letter(len(LOC_HEADS))}1")
    t = ws_l.cell(1, 1, f"Location wise Throughput - {mon_tag}")
    t.fill = HDR_FILL; t.font = TITLE_FONT; t.alignment = CTR
    ws_l.row_dimensions[1].height = 22
    for ci, h in enumerate(LOC_HEADS, 1):
        c = ws_l.cell(2, ci, h)
        c.fill = HDR_FILL; c.font = HDR_FONT; c.alignment = WRAP; c.border = BDR
    ws_l.row_dimensions[2].height = 28
    srno = 1
    for ri, row in enumerate(location_rows, 3):
        pct = float(row.get("avg_throughput_pct", 0))
        pan = bool(row.get("is_pan_india"))
        pct_fill, pct_font = _fill_font(pct)
        if pan:
            row_fill, row_font = PAN_FILL, PAN_FONT
            srno_val = "—"
        else:
            row_fill, row_font = PLAIN_FILL, PLAIN_FONT
            srno_val = srno; srno += 1
        vals = [srno_val, row.get("exco_location",""), row.get("plant_count",0),
                round(float(row.get("total_quantity",0)), 1),
                round(float(row.get("total_time_min",0)), 1), f"{round(pct)}%"]
        for ci, v in enumerate(vals, 1):
            c = ws_l.cell(ri, ci, v)
            c.border = BDR
            c.alignment = LEFT if ci == 2 else CTR
            if ci == len(LOC_HEADS):          # Avg TP % cell always gets pct color
                c.fill = pct_fill; c.font = Font(bold=pan, color=pct_font.color, size=10)
            else:
                c.fill = row_fill; c.font = row_font
    for ci, w in enumerate([30, 22, 8, 12, 12, 10], 1):
        ws_l.column_dimensions[get_column_letter(ci)].width = w

    # ── Plant sheet — Sr. no., Plant, Exco Location, Business Head, Plant Manager,
    #                 Mixer Cap, Total Qty, Time (min), TP %
    ws_p = wb.create_sheet("Plant Throughput")
    PLT_HEADS = ["Sr. no.", "Plant", "Exco Location", "Business Head",
                 "Plant Manager", "Mixer Cap", "Total Qty", "Time (min)", "TP %"]
    ws_p.merge_cells(f"A1:{get_column_letter(len(PLT_HEADS))}1")
    t2 = ws_p.cell(1, 1, f"Plant Throughput report - {mon_tag}")
    t2.fill = HDR_FILL; t2.font = TITLE_FONT; t2.alignment = CTR
    ws_p.row_dimensions[1].height = 22
    for ci, h in enumerate(PLT_HEADS, 1):
        c = ws_p.cell(2, ci, h)
        c.fill = HDR_FILL; c.font = HDR_FONT; c.alignment = WRAP; c.border = BDR
    ws_p.row_dimensions[2].height = 28
    for ri, row in enumerate(plant_rows, 3):
        pct  = float(row.get("throughput_pct", 0))
        fill, fnt = _fill_font(pct)
        vals = [ri - 2, row.get("plant_name",""), row.get("exco_location",""),
                row.get("business_head",""), row.get("plant_manager",""),
                row.get("mixer_theo_cap",""),
                round(float(row.get("total_quantity",0)), 1),
                round(float(row.get("total_time_min",0)), 1),
                f"{round(pct)}%"]
        for ci, v in enumerate(vals, 1):
            c = ws_p.cell(ri, ci, v)
            c.fill = fill; c.font = fnt; c.border = BDR
            c.alignment = LEFT if ci == 2 else CTR
    for ci, w in enumerate([30, 40, 16, 18, 18, 10, 11, 11, 8], 1):
        ws_p.column_dimensions[get_column_letter(ci)].width = w

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()


def _tp_build_html_tables(plant_rows, location_rows, month, year):
    """Return HTML tables for TP report email.
    width:100% + table-layout:auto lets each column size itself to its content.
    Plant name column wraps; all other columns stay on one line (nowrap).
    """
    mon_tag = _tp_mon_tag(month, year)

    def _bg(pct):
        if pct < 60:  return "#FFB3B3"
        if pct < 75:  return "#FFE066"
        return "#92D492"

    def _fg(bg):
        return {"#FFB3B3": "#7B1F1F", "#FFE066": "#5C4200",
                "#92D492": "#1A5C1A", "#D9D9D9": "#222222"}.get(bg, "#19263A")

    F       = "font-family:Arial,sans-serif;font-size:11px;"
    TBL_LOC = "border-collapse:collapse;width:auto;margin:4px 0 10px"
    TBL_PLT = "border-collapse:collapse;width:auto;margin:4px 0 10px"
    TTL     = (f'style="{F}font-size:12px;font-weight:bold;background:#082B49;color:#fff;'
               f'padding:3px 6px;border:1px solid #7A7A7A;text-align:left"')

    def _th(align="center", wrap=False, w=None):
        ws = "" if wrap else "white-space:nowrap;"
        wd = f"width:{w}px;" if w else ""
        wa = f'width="{w}" ' if w else ""
        return (f'{wa}style="{F}background:#082B49;color:#fff;font-weight:bold;'
                f'padding:3px 6px;border:1px solid #7A7A7A;text-align:{align};'
                f'{ws}{wd}line-height:1.2;vertical-align:middle"')

    def _td(bg, fg, align, bold=False, top_bdr="", wrap=False):
        fw  = "font-weight:bold;" if bold else ""
        top = f"border-top:{top_bdr};" if top_bdr else ""
        ws  = "" if wrap else "white-space:nowrap;"
        return (f'style="{F}padding:2px 5px;border:1px solid #9E9E9E;{top}'
                f'background:{bg};color:{fg};text-align:{align};'
                f'{ws}line-height:1.2;vertical-align:middle;{fw}"')

    # ── Location table ────────────────────────────────────────────────────────
    loc_body = ""
    srno = 1
    for r in location_rows:
        pct    = float(r.get("avg_throughput_pct", 0))
        pan    = bool(r.get("is_pan_india"))
        pct_bg = _bg(pct); pct_fg = _fg(pct_bg)
        qty    = round(float(r.get("total_quantity", 0)), 1)
        mins   = round(float(r.get("total_time_min", 0)), 1)
        plants = r.get("plant_count", 0)
        loc    = r.get("exco_location", "")
        if pan:
            bg = "#D9D9D9"; fg = "#222222"
            loc_body += (
                f'<tr>'
                f'<td {_td(bg, fg, "center", bold=True, top_bdr="2px solid #555")}>—</td>'
                f'<td {_td(bg, fg, "left",   bold=True, top_bdr="2px solid #555")}>&#127988; PAN India</td>'
                f'<td {_td(bg, fg, "center", bold=True, top_bdr="2px solid #555")}>{plants}</td>'
                f'<td {_td(bg, fg, "right",  bold=True, top_bdr="2px solid #555")}>{qty}</td>'
                f'<td {_td(bg, fg, "right",  bold=True, top_bdr="2px solid #555")}>{mins}</td>'
                f'<td {_td(pct_bg, pct_fg, "center", bold=True, top_bdr="2px solid #555")}>{round(pct)}%</td>'
                f'</tr>'
            )
        else:
            bg = "#ffffff"; fg = "#19263A"
            loc_body += (
                f'<tr>'
                f'<td {_td(bg, fg, "center")}>{srno}</td>'
                f'<td {_td(bg, fg, "left")}>{loc}</td>'
                f'<td {_td(bg, fg, "center")}>{plants}</td>'
                f'<td {_td(bg, fg, "right")}>{qty}</td>'
                f'<td {_td(bg, fg, "right")}>{mins}</td>'
                f'<td {_td(bg, fg, "center", bold=True)}>{round(pct)}%</td>'
                f'</tr>'
            )
            srno += 1

    loc_html = (
        f'<table cellpadding="0" cellspacing="0" style="{TBL_LOC}">'
        f'<tr><td colspan="6" {TTL}>Location wise Throughput - {mon_tag}</td></tr>'
        f'<tr>'
        f'<th {_th("center", w=45)}>Sr. no.</th>'
        f'<th {_th("left")}>Exco Location</th>'
        f'<th {_th("center")}>Plants</th>'
        f'<th {_th("right")}>Total Qty</th>'
        f'<th {_th("right")}>Time (min)</th>'
        f'<th {_th("center")}>Avg TP %</th>'
        f'</tr>{loc_body}</table>'
    )

    # ── Plant table ───────────────────────────────────────────────────────────
    plant_body = ""
    for i, r in enumerate(plant_rows, 1):
        pct  = float(r.get("throughput_pct", 0))
        bg   = _bg(pct); fg = _fg(bg)
        name = str(r.get("plant_name", ""))
        bh   = str(r.get("business_head", ""))
        pm   = str(r.get("plant_manager", ""))
        cap  = str(r.get("mixer_theo_cap", ""))
        qty  = round(float(r.get("total_quantity", 0)), 1)
        mins = round(float(r.get("total_time_min", 0)), 1)
        plant_body += (
            f'<tr>'
            f'<td {_td(bg, fg, "center")}>{i}</td>'
            f'<td {_td(bg, fg, "left", wrap=True)}>{name}</td>'
            f'<td {_td(bg, fg, "left")}>{bh}</td>'
            f'<td {_td(bg, fg, "left")}>{pm}</td>'
            f'<td {_td(bg, fg, "right")}>{cap}</td>'
            f'<td {_td(bg, fg, "right")}>{qty}</td>'
            f'<td {_td(bg, fg, "right")}>{mins}</td>'
            f'<td {_td(bg, fg, "center", bold=True)}>{round(pct)}%</td>'
            f'</tr>'
        )

    plant_html = (
        f'<table cellpadding="0" cellspacing="0" style="{TBL_PLT}">'
        f'<tr><td colspan="8" {TTL}>Plant Throughput Report - {mon_tag}</td></tr>'
        f'<tr>'
        f'<th {_th("center", w=45)}>Sr. no.</th>'
        f'<th {_th("center", wrap=True)}>Plant</th>'
        f'<th {_th("center")}>Business Head</th>'
        f'<th {_th("center")}>Plant Manager</th>'
        f'<th {_th("center")}>Mixer Cap</th>'
        f'<th {_th("center")}>Total Qty</th>'
        f'<th {_th("center")}>Time (min)</th>'
        f'<th {_th("center")}>TP %</th>'
        f'</tr>{plant_body}</table>'
    )

    return loc_html + plant_html


@app.route("/tp/download-excel")
def tp_download_excel():
    plant_rows    = _ms("tp", "report_plant_rows", _ms("tp", "plant_rows", []))
    location_rows = _ms("tp", "report_location_rows", _ms("tp", "location_rows", []))
    month = _ms("tp", "calc_month", _date.today().month)
    year  = _ms("tp", "calc_year",  _date.today().year)
    if not plant_rows:
        flash("No data to download — run Calculate first.", "warning")
        return redirect(url_for("tp_reports"))
    excel_bytes = _tp_build_excel(plant_rows, location_rows, month, year)
    fname = f"RDC_TP_{year}_{month:02d}.xlsx"
    return send_file(io.BytesIO(excel_bytes), as_attachment=True, download_name=fname,
                     mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")


@app.route("/tp/download-csv")
def tp_download_csv():
    plant_rows = _ms("tp", "report_plant_rows", _ms("tp", "plant_rows", []))
    month = _ms("tp", "calc_month", _date.today().month)
    year  = _ms("tp", "calc_year",  _date.today().year)
    if not plant_rows:
        flash("No data to download — run Calculate first.", "warning")
        return redirect(url_for("tp_reports"))
    df = pd.DataFrame(plant_rows)
    buf = io.BytesIO(df.to_csv(index=False).encode("utf-8"))
    return send_file(buf, as_attachment=True,
                     download_name=f"RDC_TP_{year}_{month:02d}.csv",
                     mimetype="text/csv")


@app.route("/tp/action/send-email", methods=["POST"])
def tp_send_email():
    plant_rows    = _ms("tp", "report_plant_rows", _ms("tp", "plant_rows", []))
    location_rows = _ms("tp", "report_location_rows", _ms("tp", "location_rows", []))
    month = _ms("tp", "calc_month", _date.today().month)
    year  = _ms("tp", "calc_year",  _date.today().year)
    to_addr = request.form.get("to", "").strip()
    cc_addr = request.form.get("cc", "").strip()
    subject = request.form.get("subject", "").strip()
    body    = request.form.get("body", "").strip()
    if not plant_rows:
        flash("No results to email — run Calculate first.", "warning")
        return redirect(url_for("tp_reports"))
    if not to_addr:
        flash("Please enter at least one To email address.", "error")
        return redirect(url_for("tp_reports"))

    import calendar as _cal
    month_name = _cal.month_name[month]
    if not subject:
        subject = f"RDC-TP Plant Throughput Report — {month_name} {year}"

    # Build color-coded Excel attachment
    excel_bytes = _tp_build_excel(plant_rows, location_rows, month, year)
    fname = f"RDC_TP_{year}_{month:02d}.xlsx"

    # Compact HTML email body — same style as BTRTP
    _body_font = "font-family:Arial,Calibri,sans-serif;font-size:12px;color:#000000;"
    tables_html = _tp_build_html_tables(plant_rows, location_rows, month, year)
    html_body = (
        f'<html><body style="margin:0;padding:8px 10px;{_body_font}">'
        f'<p style="margin:0 0 3px 0;">Dear Team,</p>'
        f'<p style="margin:0 0 3px 0;">Please find attached the Plant Throughput Report for {month_name} {year}.</p>'
        f'<p style="margin:0 0 6px 0;">Regards,<br>RDC Operations</p>'
        f'{tables_html}'
        f'</body></html>'
    )

    result = email_helper.send_report_email(
        to_emails=to_addr, cc_emails=cc_addr,
        subject=subject, body=body,
        attachment_bytes=excel_bytes, attachment_name=fname,
        html_body=html_body,
    )
    if result.get("success"):
        flash(f"✅ Report emailed to {to_addr}.", "success")
    else:
        flash(f"Email failed: {result.get('error')}", "error")
    return redirect(url_for("tp_reports"))


# ── Validation ────────────────────────────────────────────────────────────────
@app.route("/tp/validation")
def tp_validation():
    skip_log = _ms("tp", "skip_log", [])
    calc_warnings = _ms("tp", "calc_warnings", [])
    ctx = _tp_ctx()
    ctx["active_page"] = "validation"
    return render_template("tp_validation.html",
                           skip_log=skip_log, calc_warnings=calc_warnings, **ctx)


# ── Settings ──────────────────────────────────────────────────────────────────
@app.route("/tp/settings", methods=["GET"])
def tp_settings():
    sheet_id    = database.get_module_setting("tp", "gsheet_id",
                                              database.get_setting("gsheet_id", ""))
    worksheet   = database.get_module_setting("tp", "gsheet_worksheet", "Plant Data for TP")
    plant_col   = database.get_module_setting("tp", "oracle_plant_col", "PLANTNO")
    batch_col   = database.get_module_setting("tp", "oracle_batch_col", "BATCHCODE")
    time_col    = database.get_module_setting("tp", "oracle_time_col",  "TIMETAKEN")
    smtp        = email_helper.get_smtp_config()
    email_configured = bool(smtp.get("host") and smtp.get("sender"))
    ora_configured   = oracle_connector.is_configured()
    last_sync        = google_sheets.get_tp_last_sync_info()
    # Monthly-email schedule (TP-scoped)
    sched_enabled     = database.get_module_setting("tp", "email_schedule_enabled", "false") == "true"
    sched_time        = database.get_module_setting("tp", "email_schedule_time", "08:00")
    sched_to          = database.get_module_setting("tp", "email_schedule_to", "")
    sched_cc          = database.get_module_setting("tp", "email_schedule_cc", "")
    sched_last_status  = database.get_module_setting("tp", "email_schedule_last_status", "")
    tp_email_to        = database.get_module_setting("tp", "email_default_to", "")
    tp_email_cc        = database.get_module_setting("tp", "email_default_cc", "")
    tp_email_subject   = database.get_module_setting("tp", "email_default_subject", "")
    tp_email_body      = database.get_module_setting("tp", "email_default_body", "")
    ctx = _tp_ctx()
    ctx["active_page"] = "settings"
    return render_template("tp_settings.html",
                           sheet_id=sheet_id, worksheet=worksheet,
                           plant_col=plant_col, batch_col=batch_col, time_col=time_col,
                           smtp=smtp, email_configured=email_configured,
                           ora_configured=ora_configured, last_sync=last_sync,
                           sched_enabled=sched_enabled, sched_time=sched_time,
                           sched_to=sched_to, sched_cc=sched_cc,
                           sched_last_status=sched_last_status,
                           tp_email_to=tp_email_to, tp_email_cc=tp_email_cc,
                           tp_email_subject=tp_email_subject, tp_email_body=tp_email_body,
                           **ctx)


@app.route("/tp/settings/save-oracle-cols", methods=["POST"])
def tp_save_oracle_cols():
    database.set_module_settings_bulk("tp", {
        "oracle_plant_col": request.form.get("plant_col", "PLANTNO").strip(),
        "oracle_batch_col": request.form.get("batch_col", "BATCHCODE").strip(),
        "oracle_time_col":  request.form.get("time_col",  "TIMETAKEN").strip(),
    })
    flash("Oracle column names saved.", "success")
    return redirect(url_for("tp_settings", m="oracle-cols"))


@app.route("/tp/settings/save-worksheet", methods=["POST"])
def tp_save_worksheet():
    worksheet = request.form.get("worksheet", "Plant Data for TP").strip()
    sheet_id  = request.form.get("sheet_id", "").strip()
    database.set_module_setting("tp", "gsheet_worksheet", worksheet)
    if sheet_id:
        # Save to the TP-scoped id so we never disturb RDC-I&D's sheet.
        database.set_module_setting("tp", "gsheet_id", google_sheets.extract_sheet_id(sheet_id))
    flash("Sheet settings saved.", "success")
    return redirect(url_for("tp_settings", m="sheet"))


@app.route("/tp/settings/save-schedule", methods=["POST"])
def tp_save_email_schedule():
    sched_time = request.form.get("sched_time", "08:00").strip()
    database.set_module_settings_bulk("tp", {
        "email_schedule_time": sched_time,
        "email_schedule_to":   request.form.get("sched_to", "").strip(),
        "email_schedule_cc":   request.form.get("sched_cc", "").strip(),
    })
    if _scheduler:
        try:
            h, m = map(int, sched_time.split(":"))
        except Exception:
            h, m = 8, 0
        _scheduler.reschedule_job("tp_monthly_email",
                                  trigger=CronTrigger(day=1, hour=h, minute=m))
    flash("TP schedule settings saved.", "success")
    return redirect(url_for("tp_settings", m="schedule"))


@app.route("/tp/settings/save-email-defaults", methods=["POST"])
def tp_save_email_defaults():
    database.set_module_settings_bulk("tp", {
        "email_default_to":      request.form.get("default_to", "").strip(),
        "email_default_cc":      request.form.get("default_cc", "").strip(),
        "email_default_subject": request.form.get("default_subject", "").strip(),
        "email_default_body":    request.form.get("default_body", "").strip(),
    })
    flash("TP email defaults saved.", "success")
    return redirect(url_for("tp_settings", m="email"))


@app.route("/tp/settings/toggle-schedule", methods=["POST"])
def tp_toggle_email_schedule():
    current = database.get_module_setting("tp", "email_schedule_enabled", "false")
    new_val = "false" if current == "true" else "true"
    database.set_module_setting("tp", "email_schedule_enabled", new_val)
    flash(f"RDC-TP scheduled monthly report {'enabled' if new_val == 'true' else 'disabled'}.", "success")
    return redirect(url_for("tp_settings", m="schedule"))


# ═══════════════════════════════════════════════════════════════════════════
# RDC-BTRTP MODULE — Batcher-wise Throughput Calculator
# ═══════════════════════════════════════════════════════════════════════════

def _btrtp_ctx():
    """Base context dict for every BTRTP page."""
    return dict(
        active_page="",
        bg_auto=bg_auto(), bg_animate=bg_animate(), bg_theme=bg_theme(),
    )


def _btrtp_band(pct):
    """Throughput colour band for a percentage (same thresholds as TP)."""
    if pct < 60:
        return "red"
    if pct < 75:
        return "yellow"
    return "green"


@app.route("/btrtp")
def page_btrtp():
    return redirect(url_for("btrtp_dashboard"))


# ── Dashboard ─────────────────────────────────────────────────────────────────
@app.route("/btrtp/dashboard")
def btrtp_dashboard():
    all_counts = database.get_table_counts()
    counts = {
        "btrtp_oracle_data": all_counts.get("oracle_raw_data",   0),
        "btrtp_master_data": all_counts.get("btrtp_master_data", 0),
        "btrtp_results":     all_counts.get("btrtp_results",     0),
        "tp_plant_data":     all_counts.get("tp_plant_data",     0),
    }
    last_sync = google_sheets.get_btrtp_last_sync_info()
    ora_ready = oracle_connector.is_configured()
    ctx = _btrtp_ctx()
    ctx["active_page"] = "dashboard"
    return render_template("btrtp_dashboard.html", counts=counts,
                           last_sync=last_sync, ora_ready=ora_ready, **ctx)


# ── Data Uploader ─────────────────────────────────────────────────────────────
@app.route("/btrtp/data-uploader")
def btrtp_data_uploader():
    ora_ready    = oracle_connector.is_configured()
    last_master  = google_sheets.get_btrtp_last_sync_info()
    last_plant   = google_sheets.get_tp_last_sync_info()

    # BT Master sheet settings
    master_sheet_id  = database.get_module_setting("btrtp", "gsheet_id",
                                                    database.get_setting("gsheet_id", ""))
    master_worksheet = database.get_module_setting("btrtp", "gsheet_worksheet", "BT Master Data")

    # Plant Data sheet settings (shared reference with TP)
    plant_sheet_id   = database.get_module_setting("tp", "gsheet_id",
                                                    database.get_setting("gsheet_id", ""))
    plant_worksheet  = database.get_module_setting("tp", "gsheet_worksheet", "Plant Data for TP")

    all_counts   = database.get_table_counts()
    ora_counts   = all_counts.get("oracle_raw_data", 0)
    master_count = all_counts.get("btrtp_master_data", 0)
    plant_count  = all_counts.get("tp_plant_data", 0)

    ctx = _btrtp_ctx()
    ctx["active_page"] = "data_uploader"
    return render_template("btrtp_data_uploader.html",
                           ora_ready=ora_ready,
                           last_master=last_master, last_plant=last_plant,
                           master_sheet_id=master_sheet_id, master_worksheet=master_worksheet,
                           plant_sheet_id=plant_sheet_id, plant_worksheet=plant_worksheet,
                           ora_counts=ora_counts, master_count=master_count,
                           plant_count=plant_count, **ctx)


@app.route("/btrtp/action/fetch-oracle", methods=["POST"])
def btrtp_fetch_oracle():
    from_date = request.form.get("from_date", "")
    to_date   = request.form.get("to_date", "")
    if not from_date or not to_date:
        flash("Please select both From and To dates.", "error")
        return jsonify({"ok": False, "redirect": url_for("btrtp_data_uploader")})
    try:
        _set_progress(20, "Fetching BTRTP data from Oracle…")
        raw_df, ora_warnings = oracle_connector.fetch_btrtp_data(from_date, to_date)
        _set_progress(55, "Parsing batcher records…")
        parsed, skip_log     = btrtp_calculator.parse_btrtp_oracle_df(raw_df)
        _set_progress(75, "Saving to database…")
        oracle_connector.save_btrtp_oracle_data(parsed, replace=True)
        _set_progress(90, "Cleaning old records…")
        database.purge_old_oracle_data()
        _mss("btrtp", "skip_log",  skip_log)
        _mss("btrtp", "ora_from",  from_date)
        _mss("btrtp", "ora_to",    to_date)
        for w in ora_warnings:
            flash(w, "warning")
        flash(f"✅ {len(parsed)} rows fetched & saved ({len(skip_log)} skipped).", "success")
    except Exception as exc:
        flash(f"Oracle error: {exc}", "error")
    _set_progress(100, "Complete")
    return jsonify({"ok": True, "redirect": url_for("btrtp_data_uploader")})


@app.route("/btrtp/action/sync-master", methods=["POST"])
def btrtp_sync_master():
    sheet_id  = (request.form.get("sheet_id", "").strip()
                 or database.get_module_setting("btrtp", "gsheet_id",
                                                database.get_setting("gsheet_id", "")))
    worksheet = request.form.get("worksheet", "BT Master Data").strip()
    _set_progress(20, "Syncing BT Master Data from Google Sheets…")
    result = google_sheets.sync_btrtp_master_data(sheet_id, worksheet)
    _set_progress(85, "Saving settings…")
    if result["error"]:
        flash(f"Sync failed: {result['error']}", "error")
    elif result["rows_synced"] == 0:
        flash("⚠️ Sync completed but 0 batcher rows were saved. "
              "Check that the sheet tab name, 'Batcher ID' and 'Batcher Name' columns exist and have data.",
              "warning")
    else:
        database.set_module_setting("btrtp", "gsheet_id",
                                    google_sheets.extract_sheet_id(sheet_id))
        database.set_module_setting("btrtp", "gsheet_worksheet", worksheet)
        flash(f"✅ {result['rows_synced']} batcher rows synced from Google Sheets ({result['mode']} mode).",
              "success")
    _set_progress(100, "Complete")
    return jsonify({"ok": True, "redirect": url_for("btrtp_data_uploader")})


@app.route("/btrtp/action/sync-plant", methods=["POST"])
def btrtp_sync_plant():
    sheet_id  = (request.form.get("sheet_id", "").strip()
                 or database.get_module_setting("tp", "gsheet_id",
                                                database.get_setting("gsheet_id", "")))
    worksheet = request.form.get("worksheet", "Plant Data for TP").strip()
    _set_progress(20, "Syncing Plant Data from Google Sheets…")
    result = google_sheets.sync_tp_plant_data(sheet_id, worksheet)
    _set_progress(85, "Saving settings…")
    if result["error"]:
        flash(f"Sync failed: {result['error']}", "error")
    else:
        database.set_module_setting("tp", "gsheet_id", google_sheets.extract_sheet_id(sheet_id))
        database.set_module_setting("tp", "gsheet_worksheet", worksheet)
        flash(f"✅ {result['rows_synced']} plant rows synced from Google Sheets ({result['mode']} mode).",
              "success")
    _set_progress(100, "Complete")
    return jsonify({"ok": True, "redirect": url_for("btrtp_data_uploader")})


# ── Calculate ─────────────────────────────────────────────────────────────────
@app.route("/btrtp/calculate", methods=["GET", "POST"])
def btrtp_calculate():
    # Unified page: calculate → redirect to reports with dates in query params
    today = _date.today()
    if request.method == "POST":
        from_s = request.form.get("from_date", str(today.replace(day=1)))
        to_s   = request.form.get("to_date",   str(today))
        return redirect(url_for("btrtp_reports", from_date=from_s, to_date=to_s))
    return redirect(url_for("btrtp_reports",
                            from_date=str(today.replace(day=1)),
                            to_date=str(today)))


# ── Reports ───────────────────────────────────────────────────────────────────
def _apply_btrtp_filters(rows, excos, bheads, plants, batchers, band, search):
    """Filter batcher rows by Exco, Business Head, Plant, Batcher, band, and search."""
    out = []
    s = (search or "").strip().lower()
    for r in rows:
        if excos   and r.get("exco_location") not in excos:
            continue
        if bheads  and r.get("business_head") not in bheads:
            continue
        if plants  and r.get("plant_name") not in plants:
            continue
        if batchers and r.get("batcher_name") not in batchers:
            continue
        if band and band != "All" and _btrtp_band(r.get("throughput_pct", 0)) != band:
            continue
        if s and s not in str(r.get("batcher_id", "")).lower() \
             and s not in str(r.get("batcher_name", "")).lower() \
             and s not in str(r.get("plant_name", "")).lower():
            continue
        out.append(r)
    return out


@app.route("/btrtp/reports")
def btrtp_reports():
    today = _date.today()

    from_s = request.args.get("from_date", str(today.replace(day=1)))
    to_s   = request.args.get("to_date",   str(today))
    try:
        fd = _date.fromisoformat(from_s)
        td = _date.fromisoformat(to_s)
    except ValueError:
        fd, td = today.replace(day=1), today

    # Re-run calculation if user picks a date range directly from Reports
    if "from_date" in request.args:
        month, year = fd.month, fd.year
        all_rows, calc_warns = btrtp_calculator.run_btrtp_calculation(
            month, year, from_date=str(fd), to_date=str(td))
        _mss("btrtp", "batcher_rows",  all_rows)
        _mss("btrtp", "calc_warnings", calc_warns)
        _mss("btrtp", "calc_from",  str(fd))
        _mss("btrtp", "calc_to",    str(td))
        _mss("btrtp", "calc_month", month)
        _mss("btrtp", "calc_year",  year)
    else:
        all_rows = _ms("btrtp", "batcher_rows", [])

    excos   = request.args.getlist("exco")
    bheads  = request.args.getlist("bhead")
    plants  = request.args.getlist("plant")
    batchers = request.args.getlist("batcher")
    band    = request.args.get("band", "All")
    search  = request.args.get("search", "")

    unique_excos    = sorted({r.get("exco_location", "") for r in all_rows if r.get("exco_location")})
    unique_bheads   = sorted({r.get("business_head", "")  for r in all_rows if r.get("business_head")})
    unique_plants   = sorted({r.get("plant_name", "")     for r in all_rows if r.get("plant_name")})
    unique_batchers = sorted({r.get("batcher_name", "")   for r in all_rows if r.get("batcher_name")})

    batcher_rows = _apply_btrtp_filters(all_rows, excos, bheads, plants, batchers, band, search)

    _mss("btrtp", "report_rows", batcher_rows)

    smtp        = email_helper.get_smtp_config()
    email_ready = email_helper.is_configured()

    _elog = database.read_table("email_log", order_by="id DESC")
    if not _elog.empty and "report_file_name" in _elog.columns:
        _elog = _elog[_elog["report_file_name"].str.startswith("RDC_BTRTP_", na=False)]
    btrtp_email_log = _records(_elog.drop(columns=["id"], errors="ignore").head(20)) \
        if not _elog.empty else []

    if (fd.year, fd.month) == (td.year, td.month):
        import calendar
        month_label = f"{calendar.month_name[fd.month]} {fd.year}"
    else:
        month_label = f"{fd:%d %b %Y} → {td:%d %b %Y}"

    default_to      = database.get_module_setting("btrtp", "email_default_to", "") or smtp.get("default_to", "")
    default_cc      = database.get_module_setting("btrtp", "email_default_cc", "") or smtp.get("default_cc", "")
    default_subject = (database.get_module_setting("btrtp", "email_default_subject", "")
                       or f"RDC-BTRTP Batcher Throughput Report — {month_label}")
    default_body    = (database.get_module_setting("btrtp", "email_default_body", "")
                       or f"Dear Team,\n\nPlease find attached the Batcher Throughput "
                          f"Report for {month_label}.\n\nRegards,\nRDC Operations")

    plant_groups = _group_btrtp_by_plant(batcher_rows)
    ctx = _btrtp_ctx()
    ctx["active_page"] = "reports"
    return render_template("btrtp_reports.html",
                           batcher_rows=batcher_rows, plant_groups=plant_groups,
                           total_rows=len(all_rows),
                           from_date=str(fd), to_date=str(td),
                           unique_excos=unique_excos, unique_bheads=unique_bheads,
                           unique_plants=unique_plants, unique_batchers=unique_batchers,
                           excos=excos, bheads=bheads, plants=plants, batchers=batchers,
                           band=band, search=search,
                           month_label=month_label, email_cfg=smtp,
                           email_ready=email_ready,
                           default_to=default_to, default_cc=default_cc,
                           default_subject=default_subject, default_body=default_body,
                           btrtp_email_log=btrtp_email_log, **ctx)


def _group_btrtp_by_plant(batcher_rows: list) -> list:
    """
    Group flat batcher rows by plant (lookup_code), return list of plant-group dicts.
    Each group: {plant_name, lookup_code, mixer_label, exco_location, business_head,
                 plant_manager, mixer_theo_cap, rows, avg_tp, total_qty, total_time_hrs}
    Rows within each group sorted by TP% descending.
    """
    groups: dict = {}
    for r in batcher_rows:
        key = r["lookup_code"]
        if key not in groups:
            parts = key.split("_", 1)
            mixer_label = parts[1] if len(parts) > 1 else ""
            groups[key] = {
                "plant_name":     r["plant_name"],
                "lookup_code":    key,
                "mixer_label":    mixer_label,
                "exco_location":  r["exco_location"],
                "business_head":  r["business_head"],
                "plant_manager":  r["plant_manager"],
                "mixer_theo_cap": r["mixer_theo_cap"],
                "rows": [],
            }
        groups[key]["rows"].append(r)

    # Compute aggregates and sort rows within each group by TP% descending
    result = sorted(groups.values(), key=lambda g: (g["plant_name"], g["mixer_label"]))
    for g in result:
        g["rows"].sort(key=lambda r: -r["throughput_pct"])
        tps   = [r["throughput_pct"] for r in g["rows"]]
        g["avg_tp"]        = round(sum(tps) / len(tps), 1) if tps else 0
        g["total_qty"]     = round(sum(r["total_quantity"]  for r in g["rows"]), 1)
        g["total_time_hrs"]= round(sum(r["total_time_hrs"]  for r in g["rows"]), 2)
        g["batch_count"]   = sum(r["batch_count"] for r in g["rows"])
    return result


def _btrtp_mon_tag(month, year):
    import calendar as _cal
    return f"{_cal.month_abbr[month]}'{str(year)[2:]}"


def _btrtp_build_excel(batcher_rows, month, year):
    """Build colour-coded Excel bytes for BTRTP batcher results."""
    from openpyxl import Workbook
    from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    mon_tag = _btrtp_mon_tag(month, year)

    RED_FILL   = PatternFill("solid", fgColor="FFB3B3")
    YEL_FILL   = PatternFill("solid", fgColor="FFE066")
    GRN_FILL   = PatternFill("solid", fgColor="92D492")
    HDR_FILL   = PatternFill("solid", fgColor="0A2540")
    PLAIN_FILL = PatternFill("solid", fgColor="FFFFFF")

    RED_FONT   = Font(color="7B1F1F", size=10)
    YEL_FONT   = Font(color="5C4200", size=10)
    GRN_FONT   = Font(color="1A5C1A", size=10)
    HDR_FONT   = Font(bold=True, color="FFFFFF", size=10)
    TITLE_FONT = Font(bold=True, color="FFFFFF", size=11)
    PLAIN_FONT = Font(color="19263A", size=10)

    CTR  = Alignment(vertical="center", horizontal="center")
    LEFT = Alignment(vertical="center", horizontal="left")
    WRAP = Alignment(wrap_text=True, vertical="center", horizontal="center")
    thin = Side(style="thin", color="9A9A9A")
    BDR  = Border(left=thin, right=thin, top=thin, bottom=thin)

    def _fill_font(pct):
        if pct < 60:  return RED_FILL, RED_FONT
        if pct < 75:  return YEL_FILL, YEL_FONT
        return GRN_FILL, GRN_FONT

    wb = Workbook()
    ws = wb.active
    ws.title = "Batcher Throughput"

    HEADS = ["Sr. no.", "Batcher ID", "Batcher Name", "Plant", "Business Head",
             "Plant Manager", "Mixer Cap", "Total Qty", "Time (hr)", "TP %", "Batches"]

    ws.merge_cells(f"A1:{get_column_letter(len(HEADS))}1")
    t = ws.cell(1, 1, f"Batcher wise Throughput - {mon_tag}")
    t.fill = HDR_FILL; t.font = TITLE_FONT; t.alignment = CTR
    ws.row_dimensions[1].height = 22

    for ci, h in enumerate(HEADS, 1):
        c = ws.cell(2, ci, h)
        c.fill = HDR_FILL; c.font = HDR_FONT; c.alignment = WRAP; c.border = BDR
    ws.row_dimensions[2].height = 28

    srno = 1
    for ri, row in enumerate(batcher_rows, 3):
        pct        = float(row.get("throughput_pct", 0))
        row_fill, row_font = _fill_font(pct)
        vals = [
            srno,
            row.get("batcher_id", ""),
            row.get("batcher_name", ""),
            row.get("plant_name", ""),
            row.get("business_head", ""),
            row.get("plant_manager", ""),
            round(float(row.get("mixer_theo_cap", 0)), 1),
            round(float(row.get("total_quantity", 0)), 1),
            round(float(row.get("total_time_hrs", 0)), 2),
            f"{round(pct)}%",
            int(row.get("batch_count", 0)),
        ]
        for ci, v in enumerate(vals, 1):
            c = ws.cell(ri, ci, v)
            c.border = BDR
            c.fill   = row_fill
            c.font   = row_font
            c.alignment = LEFT if ci in (2, 3, 4, 5, 6) else CTR
        srno += 1

    for ci, w in enumerate([8, 14, 20, 22, 18, 18, 10, 10, 10, 8, 8], 1):
        ws.column_dimensions[get_column_letter(ci)].width = w

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()


@app.route("/btrtp/download-excel")
def btrtp_download_excel():
    rows  = _ms("btrtp", "report_rows", _ms("btrtp", "batcher_rows", []))
    month = _ms("btrtp", "calc_month", _date.today().month)
    year  = _ms("btrtp", "calc_year",  _date.today().year)
    if not rows:
        flash("No data to download — run Calculate first.", "warning")
        return redirect(url_for("btrtp_reports"))
    excel_bytes = _btrtp_build_excel(rows, month, year)
    fname = f"RDC_BTRTP_{year}_{month:02d}.xlsx"
    return send_file(io.BytesIO(excel_bytes), as_attachment=True, download_name=fname,
                     mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")


@app.route("/btrtp/download-csv")
def btrtp_download_csv():
    rows  = _ms("btrtp", "report_rows", _ms("btrtp", "batcher_rows", []))
    month = _ms("btrtp", "calc_month", _date.today().month)
    year  = _ms("btrtp", "calc_year",  _date.today().year)
    if not rows:
        flash("No data to download — run Calculate first.", "warning")
        return redirect(url_for("btrtp_reports"))
    df = pd.DataFrame(rows)
    buf = io.BytesIO(df.to_csv(index=False).encode("utf-8"))
    return send_file(buf, as_attachment=True,
                     download_name=f"RDC_BTRTP_{year}_{month:02d}.csv",
                     mimetype="text/csv")


@app.route("/btrtp/action/send-email", methods=["POST"])
def btrtp_send_email():
    rows  = _ms("btrtp", "report_rows", _ms("btrtp", "batcher_rows", []))
    month = _ms("btrtp", "calc_month", _date.today().month)
    year  = _ms("btrtp", "calc_year",  _date.today().year)
    to_addr = request.form.get("to", "").strip()
    cc_addr = request.form.get("cc", "").strip()
    subject = request.form.get("subject", "").strip()
    body    = request.form.get("body", "").strip()

    if not rows:
        flash("No results to email — run Calculate first.", "warning")
        return redirect(url_for("btrtp_reports"))
    if not to_addr:
        flash("Please enter at least one To email address.", "error")
        return redirect(url_for("btrtp_reports"))

    import calendar as _cal
    month_name = _cal.month_name[month]
    mon_tag = _btrtp_mon_tag(month, year)
    if not subject:
        subject = f"RDC-BTRTP Batcher Throughput Report — {month_name} {year}"
    if not body:
        body = f"Dear Team,\n\nPlease find the Batcher Throughput Report for {month_name} {year} below."

    excel_bytes = _btrtp_build_excel(rows, month, year)
    fname = f"RDC_BTRTP_{year}_{month:02d}.xlsx"

    # Compact HTML email body with inline CSS (Gmail / Outlook safe)
    _TH = ('style="background-color:#082B49;color:#ffffff;font-weight:bold;'
           'text-align:center;padding:3px 6px;border:1px solid #808080;'
           'font-size:11px;font-family:Arial,sans-serif;line-height:1.2;white-space:nowrap;"')
    def _td_b(bg, fg, align="center"):
        return (f'style="background-color:{bg};color:{fg};padding:2px 5px;'
                f'border:1px solid #999999;text-align:{align};'
                f'font-size:11px;font-family:Arial,sans-serif;line-height:1.2;white-space:nowrap;"')
    def _row_color(pct):
        if pct < 60: return "#FFB3B3", "#7B1F1F"
        if pct < 75: return "#FFE066", "#5C4200"
        return "#92D492", "#1A5C1A"

    head_row = (f'<tr><th {_TH}>Sr.</th><th {_TH}>Batcher ID</th>'
                f'<th {_TH}>Batcher Name</th><th {_TH}>Plant</th>'
                f'<th {_TH}>Mixer Cap</th><th {_TH}>Total Qty</th>'
                f'<th {_TH}>Time (hr)</th><th {_TH}>TP %</th><th {_TH}>Batches</th></tr>')
    body_rows = ""
    for i, r in enumerate(rows, 1):
        pct = float(r.get("throughput_pct", 0))
        bg, fg = _row_color(pct)
        body_rows += (
            f'<tr>'
            f'<td {_td_b(bg,fg,"center")}>{i}</td>'
            f'<td {_td_b(bg,fg,"left")}>{r.get("batcher_id","")}</td>'
            f'<td {_td_b(bg,fg,"left")}>{r.get("batcher_name","")}</td>'
            f'<td {_td_b(bg,fg,"left")}>{r.get("plant_name","")}</td>'
            f'<td {_td_b(bg,fg,"center")}>{round(float(r.get("mixer_theo_cap",0)),1)}</td>'
            f'<td {_td_b(bg,fg,"center")}>{round(float(r.get("total_quantity",0)),1)}</td>'
            f'<td {_td_b(bg,fg,"center")}>{round(float(r.get("total_time_hrs",0)),2)}</td>'
            f'<td {_td_b(bg,fg,"center")}><b>{round(pct)}%</b></td>'
            f'<td {_td_b(bg,fg,"center")}>{int(r.get("batch_count",0))}</td>'
            f'</tr>'
        )
    _tbl_css = ('border-collapse:collapse;width:auto;font-family:Arial,sans-serif;'
                'font-size:11px;margin:0;padding:0;')
    _body_font = 'font-family:Arial,Calibri,sans-serif;font-size:12px;color:#000000;'
    html_body = (
        f'<html><body style="margin:0;padding:8px 10px;{_body_font}">'
        f'<p style="margin:0 0 3px 0;">Dear Team,</p>'
        f'<p style="margin:0 0 3px 0;">Please find attached the Batcher Throughput Report for {month_name} {year}.</p>'
        f'<p style="margin:0 0 6px 0;">Regards,<br>RDC Operations</p>'
        f'<p style="margin:0 0 4px 0;font-size:13px;font-weight:bold;color:#082B49;">'
        f'Batcher Throughput Report - {mon_tag}</p>'
        f'<table style="{_tbl_css}">{head_row}{body_rows}</table>'
        f'</body></html>'
    )

    result = email_helper.send_report_email(
        to_emails=to_addr, cc_emails=cc_addr,
        subject=subject, body=body,
        attachment_bytes=excel_bytes, attachment_name=fname,
        html_body=html_body,
    )
    if result.get("success"):
        flash(f"✅ Report emailed to {to_addr}.", "success")
    else:
        flash(f"Email failed: {result.get('error')}", "error")
    return redirect(url_for("btrtp_reports"))


# ── Validation ────────────────────────────────────────────────────────────────
@app.route("/btrtp/validation")
def btrtp_validation():
    skip_log      = _ms("btrtp", "skip_log", [])
    calc_warnings = _ms("btrtp", "calc_warnings", [])
    ctx = _btrtp_ctx()
    ctx["active_page"] = "validation"
    return render_template("btrtp_validation.html",
                           skip_log=skip_log, calc_warnings=calc_warnings, **ctx)


# ── Settings ──────────────────────────────────────────────────────────────────
@app.route("/btrtp/settings", methods=["GET"])
def btrtp_settings():
    # BT Master sheet
    master_sheet_id  = database.get_module_setting("btrtp", "gsheet_id",
                                                    database.get_setting("gsheet_id", ""))
    master_worksheet = database.get_module_setting("btrtp", "gsheet_worksheet", "BT Master Data")
    # Plant Data sheet (shared with TP)
    plant_sheet_id   = database.get_module_setting("tp", "gsheet_id",
                                                    database.get_setting("gsheet_id", ""))
    plant_worksheet  = database.get_module_setting("tp", "gsheet_worksheet", "Plant Data for TP")

    batcher_col = database.get_module_setting("btrtp", "oracle_batcher_col", "CREATED_BY")
    smtp        = email_helper.get_smtp_config()
    email_configured = bool(smtp.get("host") and smtp.get("sender"))
    ora_configured   = oracle_connector.is_configured()
    last_master = google_sheets.get_btrtp_last_sync_info()
    last_plant  = google_sheets.get_tp_last_sync_info()
    btrtp_email_to      = database.get_module_setting("btrtp", "email_default_to", "")
    btrtp_email_cc      = database.get_module_setting("btrtp", "email_default_cc", "")
    btrtp_email_subject = database.get_module_setting("btrtp", "email_default_subject", "")
    btrtp_email_body    = database.get_module_setting("btrtp", "email_default_body", "")
    ctx = _btrtp_ctx()
    ctx["active_page"] = "settings"
    return render_template("btrtp_settings.html",
                           master_sheet_id=master_sheet_id, master_worksheet=master_worksheet,
                           plant_sheet_id=plant_sheet_id, plant_worksheet=plant_worksheet,
                           batcher_col=batcher_col,
                           smtp=smtp, email_configured=email_configured,
                           ora_configured=ora_configured,
                           last_master=last_master, last_plant=last_plant,
                           btrtp_email_to=btrtp_email_to, btrtp_email_cc=btrtp_email_cc,
                           btrtp_email_subject=btrtp_email_subject,
                           btrtp_email_body=btrtp_email_body, **ctx)


@app.route("/btrtp/settings/save-oracle-cols", methods=["POST"])
def btrtp_save_oracle_cols():
    database.set_module_setting("btrtp", "oracle_batcher_col",
                                request.form.get("batcher_col", "CREATED_BY").strip())
    flash("Oracle column settings saved.", "success")
    return redirect(url_for("btrtp_settings", m="oracle-cols"))


@app.route("/btrtp/settings/save-master-sheet", methods=["POST"])
def btrtp_save_master_sheet():
    worksheet = request.form.get("worksheet", "BT Master Data").strip()
    sheet_id  = request.form.get("sheet_id", "").strip()
    database.set_module_setting("btrtp", "gsheet_worksheet", worksheet)
    if sheet_id:
        database.set_module_setting("btrtp", "gsheet_id",
                                    google_sheets.extract_sheet_id(sheet_id))
    flash("BT Master Data sheet settings saved.", "success")
    return redirect(url_for("btrtp_settings", m="master-sheet"))


@app.route("/btrtp/settings/save-plant-sheet", methods=["POST"])
def btrtp_save_plant_sheet():
    worksheet = request.form.get("worksheet", "Plant Data for TP").strip()
    sheet_id  = request.form.get("sheet_id", "").strip()
    database.set_module_setting("tp", "gsheet_worksheet", worksheet)
    if sheet_id:
        database.set_module_setting("tp", "gsheet_id",
                                    google_sheets.extract_sheet_id(sheet_id))
    flash("Plant Data sheet settings saved.", "success")
    return redirect(url_for("btrtp_settings", m="plant-sheet"))


@app.route("/btrtp/settings/save-email-defaults", methods=["POST"])
def btrtp_save_email_defaults():
    database.set_module_settings_bulk("btrtp", {
        "email_default_to":      request.form.get("default_to", "").strip(),
        "email_default_cc":      request.form.get("default_cc", "").strip(),
        "email_default_subject": request.form.get("default_subject", "").strip(),
        "email_default_body":    request.form.get("default_body", "").strip(),
    })
    flash("BTRTP email defaults saved.", "success")
    return redirect(url_for("btrtp_settings", m="email"))


@app.route("/jldc")
def page_jldc():
    return render_template("module_placeholder.html",
                           module_name="RDC-JLDC",
                           module_desc="Calculates LJCB & Loader Diesel Consumption")


@app.route("/data-uploader")
def page_data_uploader():
    _ss("active_page", "data_uploader")
    last_sync = google_sheets.get_last_sync_info()
    has_creds = google_sheets.credentials_exist()
    ora_cfg   = oracle_connector.get_oracle_config()
    ora_ready = oracle_connector.is_configured(ora_cfg)
    auto_sync = database.get_setting("gsheet_auto_sync", "false") == "true"

    m_count = database.get_table_counts().get("master_data", 0)
    master_cols, master_rows = [], []
    if m_count > 0:
        df = database.read_table_limited("master_data", order_by="employee_code", limit=500)
        df = df.drop(columns=["id"], errors="ignore")
        master_cols = df.columns.tolist()
        master_rows = _records(df)

    b_count = database.get_table_counts().get("backend_data", 0)
    b_earliest = b_latest = ""
    backend_preview = []
    if b_count > 0:
        df_e = database.read_table_limited("backend_data", order_by="date", limit=1)
        df_l = database.read_table_limited("backend_data", order_by="date DESC", limit=1)
        b_earliest = df_e["date"].iloc[0] if not df_e.empty else ""
        b_latest   = df_l["date"].iloc[0] if not df_l.empty else ""
        df_p = database.read_table_limited("backend_data", order_by="date", limit=200)
        backend_preview = _records(df_p.drop(columns=["id"], errors="ignore"))

    maint_df = database.read_table("maintenance_cost", order_by="plant_code")
    maint_cols, maint_months_data, maint_avg, maint_above = [], {}, 0, 0
    if not maint_df.empty:
        maint_avg   = round(maint_df["ytd_maintenance_cost"].mean(), 2)
        maint_above = int((maint_df["ytd_maintenance_cost"] > config.MAINTENANCE_COST_THRESHOLD).sum())
        # Group rows by (year, month) for the template
        import calendar as _cal
        for _, r in maint_df.iterrows():
            m, y = int(r.get("month") or 0), int(r.get("year") or 0)
            key = (y, m)
            if key not in maint_months_data:
                label = f"{_cal.month_name[m]} {y}" if m else "Unassigned"
                maint_months_data[key] = {"label": label, "month": m, "year": y, "rows": []}
            maint_months_data[key]["rows"].append({
                "plant_code":           r["plant_code"],
                "ytd_maintenance_cost": r["ytd_maintenance_cost"],
                "uploaded_at":          r.get("uploaded_at", ""),
            })
    maint_month_groups = sorted(maint_months_data.values(),
                                key=lambda g: (g["year"], g["month"]), reverse=True)
    maint_count = len(maint_df) if not maint_df.empty else 0
    # Set of (month, year) tuples already uploaded — used to lock the upload form
    maint_uploaded_keys = {(g["month"], g["year"]) for g in maint_month_groups
                           if g["month"] and g["year"]}
    has_unassigned = any(g["month"] == 0 for g in maint_month_groups)

    codes_df = database.read_table_limited("master_data", order_by="employee_code", limit=100000)
    codes = codes_df["employee_code"].astype(str).tolist() if not codes_df.empty else []

    log_df   = database.read_table("master_data_change_log", order_by="id DESC")
    log_rows = _records(log_df.drop(columns=["id"], errors="ignore")) if not log_df.empty else []

    ora_b_preview = []
    if b_count > 0:
        df_ob = database.read_table_limited("backend_data", order_by="date DESC", limit=10)
        ora_b_preview = _records(df_ob.drop(columns=["id"], errors="ignore"))

    # Edit / delete employee (loaded when query params present)
    edit_code = request.args.get("edit_code")
    del_code  = request.args.get("del_code")
    edit_emp  = database.get_employee(edit_code) if edit_code else None
    del_emp   = database.get_employee(del_code)  if del_code  else None

    return render_template("data_uploader.html",
                           last_sync=last_sync, has_creds=has_creds,
                           ora_cfg=ora_cfg, ora_ready=ora_ready, auto_sync=auto_sync,
                           m_count=m_count, master_cols=master_cols, master_rows=master_rows,
                           b_count=b_count, b_earliest=b_earliest, b_latest=b_latest,
                           backend_preview=backend_preview,
                           maint_month_groups=maint_month_groups,
                           maint_avg=maint_avg, maint_above=maint_above,
                           maint_count=maint_count,
                           current_month=_date.today().month,
                           current_year=_date.today().year,
                           maint_uploaded_keys=maint_uploaded_keys,
                           has_unassigned=has_unassigned,
                           codes=codes, log_rows=log_rows, ora_b_preview=ora_b_preview,
                           edit_emp=edit_emp, del_emp=del_emp,
                           categories=config.CATEGORIES,
                           plant_list=database.get_tp_plants(),
                           today=str(_date.today()),
                           today_first=str(_date.today().replace(day=1)))


@app.route("/calculate")
def page_calculate():
    _ss("active_page", "calculate")
    available = calculator.get_available_months()

    default_from = default_to = str(_date.today())
    if available:
        conn = database.get_connection()
        try:
            mm = pd.read_sql_query(
                "SELECT MIN(date) AS mn, MAX(date) AS mx FROM backend_data", conn)
        finally:
            conn.close()
        default_from = str(pd.to_datetime(mm["mn"].iloc[0]).date())
        default_to   = str(pd.to_datetime(mm["mx"].iloc[0]).date())

    last_calc  = calculator.get_last_calculation_info()
    results_df = database.read_table("calculation_results") \
        if database.get_table_counts().get("calculation_results", 0) > 0 \
        else pd.DataFrame()

    results       = _records(results_df)
    unmapped_rows = _records(database.read_table("unmapped_employees")
                             .drop(columns=["id"], errors="ignore"))

    cat_results = {}
    for label, cats in CAT_TABS.items():
        grp = [r for r in results if r.get("category") in cats]
        grp = _sort_rows(grp)
        for r in grp:
            r["_cls"] = _row_cls(r)
        cat_results[label] = grp

    total_emps = len(results)
    elig_count = sum(1 for r in results if r.get("incentive_eligible") == "Yes")
    total_inc  = sum((r.get("incentive_amount") or 0) for r in results)
    total_ded  = sum((r.get("deduction_amount") or 0) for r in results)

    import calendar as _cal
    _maint_months = database.get_maintenance_months()
    if _maint_months:
        _mm, _my = _maint_months[0]
        maint_month_label = f"{_cal.month_name[_mm]} {_my}" if _mm else "Unassigned"
        _lc_month = int(last_calc.get("month") or 0)
        _lc_year  = int(last_calc.get("year")  or 0)
        maint_mismatch = bool(_lc_month and _lc_year and (_mm != _lc_month or _my != _lc_year))
    else:
        maint_month_label = None
        maint_mismatch    = False

    all_waivers = database.get_all_waivers()
    import calendar as _cal2
    waiver_month_opts = [(m, _cal2.month_name[m]) for m in range(1, 13)]
    # Employee list for searchable waiver LOV
    _emp_df = database.read_table_limited("master_data", order_by="employee_code", limit=100000)
    waiver_employees = (
        [{"code": str(r["employee_code"]), "name": str(r.get("employee_name", ""))}
         for _, r in _emp_df.iterrows()]
        if not _emp_df.empty else []
    )

    return render_template("calculate.html",
                           available=available,
                           default_from=default_from, default_to=default_to,
                           last_calc=last_calc,
                           results=results if not results_df.empty else None,
                           ran=_s("calc_ran", False),
                           cat_tabs=CAT_TABS, cat_results=cat_results,
                           unmapped=unmapped_rows,
                           result_cols=RESULT_COLS, result_labels=RESULT_LABELS,
                           total_emps=total_emps, elig_count=elig_count,
                           total_inc=total_inc, total_ded=total_ded,
                           maint_month_label=maint_month_label,
                           maint_mismatch=maint_mismatch,
                           all_waivers=all_waivers,
                           waiver_month_opts=waiver_month_opts,
                           waiver_employees=waiver_employees,
                           current_year=_date.today().year)


@app.route("/reports")
def page_reports():
    _ss("active_page", "reports")

    available  = calculator.get_available_months()
    ora_live   = oracle_connector.is_configured()
    no_backend = (not available) and (not ora_live)

    # Waiver data
    today = _date.today()
    all_waivers  = database.get_all_waivers()
    _wdf = database.read_table_limited("master_data", order_by="employee_code", limit=5000)
    waiver_employees = [{"code": r["employee_code"], "name": r["employee_name"]}
                        for _, r in _wdf.iterrows()] if not _wdf.empty else []
    waiver_month_opts = [(m, __import__('calendar').month_name[m]) for m in range(1, 13)]
    current_year = today.year
    last_calc    = calculator.get_last_calculation_info()

    from_date_s = request.args.get("from_date", str(_date.today().replace(day=1)))
    to_date_s   = request.args.get("to_date",   str(_date.today()))
    try:
        from_date = _date.fromisoformat(from_date_s)
        to_date   = _date.fromisoformat(to_date_s)
    except ValueError:
        from_date = _date.today().replace(day=1)
        to_date   = _date.today()

    cats   = request.args.getlist("category")
    desigs = request.args.getlist("designation")
    plants = request.args.getlist("plant")
    elig   = request.args.get("elig",    "All")
    outcome = request.args.get("outcome", "All")
    search  = request.args.get("search",  "")

    results = None
    unmapped = []
    error_msg = ""
    ora_note  = ""
    unique_cats = unique_desigs = unique_plants = []
    total_rows = 0

    has_params = "from_date" in request.args

    # Always pre-populate filter dropdowns from the persisted calculation_results table
    # so filter selects are usable before the user clicks "Load Report"
    _cr_all = database.read_table("calculation_results")
    if not _cr_all.empty:
        unique_cats   = sorted(_cr_all["category"].dropna().unique().tolist())   if "category"    in _cr_all.columns else []
        unique_desigs = sorted(_cr_all["designation"].dropna().unique().tolist()) if "designation" in _cr_all.columns else []
        unique_plants = sorted(_cr_all["plant"].dropna().unique().tolist())       if "plant"       in _cr_all.columns else []

    if has_params and not no_backend:
        range_key = f"{from_date}|{to_date}"
        cached_empty = (not _s("rpt_all", None)) and _s("rpt_range_key") == range_key
        if _s("rpt_range_key") != range_key or cached_empty:
            if ora_live:
                try:
                    conn_ora = database.get_connection()
                    ora_cnt = conn_ora.execute(
                        "SELECT COUNT(*) FROM oracle_raw_data "
                        "WHERE production_date >= ? AND production_date <= ?",
                        (str(from_date), str(to_date))
                    ).fetchone()[0]
                    conn_ora.close()
                    ora_note = (f"Oracle: {ora_cnt:,} rows loaded."
                                if ora_cnt > 0 else "Oracle returned no rows for this range.")
                except Exception as exc:
                    ora_note = f"Oracle check failed: {exc}"

            res = calculator.run_calculation(
                month=from_date.month, year=from_date.year,
                start_date=str(from_date), end_date=str(to_date),
                persist=False,
            )
            if res["error"]:
                error_msg = res["error"]
                _ss("rpt_all", [])
                _ss("rpt_unmapped", [])
            else:
                _ss("rpt_all", res.get("results_rows", []))
                _ss("rpt_unmapped", res.get("unmapped_rows", []))
            _ss("rpt_range_key", range_key)
            _ss("rpt_ora_note", ora_note)

        all_rows = _s("rpt_all", [])
        unmapped  = _s("rpt_unmapped", [])
        ora_note  = _s("rpt_ora_note", "")
        total_rows = len(all_rows)

        unique_cats   = sorted({r.get("category", "")    for r in all_rows if r.get("category")})
        unique_desigs = sorted({r.get("designation", "") for r in all_rows if r.get("designation")})
        unique_plants = sorted({r.get("plant", "")       for r in all_rows if r.get("plant")})

        filtered = _apply_filters(all_rows, cats, desigs, plants, elig, outcome, search)
        filtered = _sort_rows(filtered)

        plant_inc_map: dict = {}
        plant_ded_map: dict = {}
        for r in all_rows:
            p = r.get("plant", "")
            plant_inc_map[p] = plant_inc_map.get(p, 0.0) + (r.get("incentive_amount") or 0)
            plant_ded_map[p] = plant_ded_map.get(p, 0.0) + (r.get("deduction_amount") or 0)

        for r in filtered:
            r["_cls"] = _row_cls(r)
            r["plant_total_incentive"] = plant_inc_map.get(r.get("plant", ""), 0.0)
            r["plant_total_deduction"] = plant_ded_map.get(r.get("plant", ""), 0.0)
        results = filtered

        parts = []
        if cats:   parts.append(f"Category: {', '.join(cats)}")
        if desigs: parts.append(f"Designation: {', '.join(desigs)}")
        if plants: parts.append(f"Plant: {', '.join(plants)}")
        if elig != "All":     parts.append(f"Eligibility: {elig}")
        if outcome != "All":  parts.append(f"Outcome: {outcome}")
        if search.strip():    parts.append(f"Search: {search.strip()}")
        applied_filters = " | ".join(parts) if parts else "None"

        _ss("rpt_filtered",        filtered)
        _ss("rpt_unmapped_snap",   unmapped)
        _ss("rpt_from",            str(from_date))
        _ss("rpt_to",              str(to_date))
        _ss("rpt_applied_filters", applied_filters)

    # Month label for email
    if (from_date.year, from_date.month) == (to_date.year, to_date.month):
        month_label = from_date.strftime("%B %Y")
    else:
        month_label = f"{from_date.strftime('%d %b %Y')} to {to_date.strftime('%d %b %Y')}"

    email_cfg    = email_helper.get_smtp_config()
    email_ready  = bool(email_cfg["host"] and email_cfg["sender"] and email_cfg["password"])
    default_subj = email_helper.compose_report_subject(month_label)
    default_body = email_helper.compose_report_body(month_label)

    elog_df    = database.read_table("email_log", order_by="id DESC")
    email_log  = _records(elog_df.drop(columns=["id"], errors="ignore").head(20)) \
        if not elog_df.empty else []

    total_inc  = sum((r.get("incentive_amount") or 0) for r in (results or []))
    total_ded  = sum((r.get("deduction_amount") or 0) for r in (results or []))
    elig_count = sum(1 for r in (results or []) if r.get("incentive_eligible") == "Yes")

    # Group filtered rows by CAT_TABS for the tab UI (mirrors calculate page)
    cat_results = {}
    for label, cats in CAT_TABS.items():
        cat_results[label] = [r for r in (results or []) if r.get("category") in cats]

    import calendar as _cal
    _maint_months = database.get_maintenance_months()
    if _maint_months:
        _mm, _my = _maint_months[0]
        maint_month_label = f"{_cal.month_name[_mm]} {_my}" if _mm else "Unassigned"
        # Only flag mismatch after user has explicitly chosen a date range
        maint_mismatch = has_params and (_mm != from_date.month or _my != from_date.year)
    else:
        maint_month_label = None
        maint_mismatch    = False

    return render_template("reports.html",
                           from_date=str(from_date), to_date=str(to_date),
                           results=results,
                           cat_tabs=CAT_TABS, cat_results=cat_results,
                           unmapped=unmapped,
                           total_rows=total_rows,
                           result_cols=REPORT_COLS, result_labels=REPORT_LABELS,
                           unique_cats=unique_cats, unique_desigs=unique_desigs,
                           unique_plants=unique_plants,
                           cats=cats, desigs=desigs, plants=plants,
                           elig=elig, outcome=outcome, search=search,
                           total_inc=total_inc, total_ded=total_ded,
                           elig_count=elig_count,
                           error_msg=error_msg, ora_note=ora_note,
                           ora_live=ora_live, no_backend=no_backend,
                           month_label=month_label,
                           email_cfg=email_cfg, email_ready=email_ready,
                           default_subject=default_subj, default_body=default_body,
                           email_log=email_log,
                           maint_month_label=maint_month_label,
                           maint_mismatch=maint_mismatch,
                           all_waivers=all_waivers,
                           waiver_employees=waiver_employees,
                           waiver_month_opts=waiver_month_opts,
                           current_year=current_year,
                           last_calc=last_calc)


@app.route("/validation")
def page_validation():
    _ss("active_page", "validation")
    last    = validations.get_last_validation_info()
    counts  = database.get_table_counts()
    total_errors = counts.get("validation_errors", 0)

    summary_rows = summary_cols = []
    master_n = backend_n = maint_n = 0
    errors   = []
    sources  = []
    etypes   = []
    sel_src  = request.args.get("src", "")
    sel_etype = request.args.get("etype", "")

    # Human-readable labels and fix instructions for each error type
    _ERR_META = {
        "BLANK_FIELD":          ("🔴 Blank Field",          "A required column is empty. Fill it in the Google Sheet and re-sync."),
        "INVALID_CATEGORY":     ("🔴 Invalid Category",     f"Category value is not in the allowed list: {config.CATEGORIES}. Fix spelling in the Google Sheet and re-sync."),
        "DUPLICATE_CODE":       ("🔴 Duplicate Employee Code", "The same Employee Code appears more than once. Remove the duplicate row from the Google Sheet and re-sync."),
        "NO_DATA":              ("⚠️ No Data",              "This data source is empty. Upload or sync data first."),
        "BLANK_EMPLOYEE_CODE":  ("🔴 Blank Employee Code",  "The 'Created by' column is blank for this batch row. Fix in the Backend Data Excel."),
        "UNMAPPED_EMPLOYEE":    ("🟡 Unmapped Employee",    "Employee Code exists in Backend Data but not in Master Data. Add the employee to the Master Data Google Sheet and re-sync."),
        "INVALID_QUANTITY":     ("🔴 Invalid Quantity",     "Quantity is zero, negative, or not a number. Fix in the Backend Data Excel."),
        "INVALID_DATE":         ("🔴 Invalid Date",         "Date is not in YYYY-MM-DD format. Fix in the Backend Data Excel."),
        "BLANK_PLANT_CODE":     ("🔴 Blank Plant Code",     "Plant Code is missing in the Maintenance Cost file."),
        "UNMAPPED_PLANT":       ("🟡 Unmapped Plant",       "Plant Code in Maintenance Cost file is not in Master Data. Check spelling."),
        "INVALID_COST":         ("🔴 Invalid Cost",         "YTD Maintenance Cost is negative or not a number."),
    }
    _SRC_LABEL = {
        "master_data":       "Master Data",
        "backend_data":      "Backend Data",
        "maintenance_cost":  "Maintenance Cost",
    }

    if total_errors > 0:
        all_err_df = database.read_table("validation_errors")
        if not all_err_df.empty:
            master_n  = int((all_err_df["source"] == "master_data").sum())
            backend_n = int((all_err_df["source"] == "backend_data").sum())
            maint_n   = int((all_err_df["source"] == "maintenance_cost").sum())

            # Build grouped sections for the template
            # Section = one card per (source, error_type) combination
            from collections import OrderedDict
            sections = OrderedDict()
            for _, r in all_err_df.iterrows():
                src   = r["source"]
                etype = r["error_type"]
                key   = (src, etype)
                if key not in sections:
                    label, fix = _ERR_META.get(etype, (etype, "Review and correct this data."))
                    sections[key] = {
                        "source":       src,
                        "source_label": _SRC_LABEL.get(src, src),
                        "error_type":   etype,
                        "label":        label,
                        "fix":          fix,
                        "rows":         [],
                    }
                sections[key]["rows"].append({
                    "row_number":  r.get("row_number", ""),
                    "column_name": r.get("column_name", ""),
                    "message":     r.get("error_message", ""),
                })

            # Apply source filter
            if sel_src:
                sections = {k: v for k, v in sections.items() if v["source"] == sel_src}
            if sel_etype:
                sections = {k: v for k, v in sections.items() if v["error_type"] == sel_etype}

            err_sections = list(sections.values())
            sources = sorted(all_err_df["source"].unique().tolist())
            etypes  = sorted(all_err_df["error_type"].unique().tolist())
        else:
            err_sections = []
    else:
        err_sections = []

    return render_template("validation.html",
                           last=last, total_errors=total_errors,
                           master_n=master_n, backend_n=backend_n, maint_n=maint_n,
                           err_sections=err_sections,
                           sources=sources, etypes=etypes,
                           sel_src=sel_src, sel_etype=sel_etype,
                           src_label=_SRC_LABEL)


@app.route("/settings")
def page_settings():
    _ss("active_page", "settings")
    smtp       = email_helper.get_smtp_config()
    email_configured = bool(smtp["host"] and smtp["sender"] and smtp["password"])
    ora        = oracle_connector.get_oracle_config()
    ora_configured = oracle_connector.is_configured(ora)

    last_cache_clear   = database.get_setting("last_cache_clear_date", "")
    cache_cleared_today = last_cache_clear == str(_date.today())

    sched_enabled     = database.get_setting("email_schedule_enabled", "false") == "true"
    sched_time        = database.get_setting("email_schedule_time", "08:00")
    sched_to          = database.get_setting("email_schedule_to", "")
    sched_cc          = database.get_setting("email_schedule_cc", "")
    sched_last_status = database.get_setting("email_schedule_last_status", "")

    return render_template("settings.html",
                           smtp=smtp, email_configured=email_configured,
                           ora=ora, ora_configured=ora_configured,
                           db_size_mb=cache_helpers.get_db_size_mb(),
                           bg_auto=bg_auto(), bg_animate=bg_animate(), bg_theme=bg_theme(),
                           last_cache_clear=last_cache_clear,
                           cache_cleared_today=cache_cleared_today,
                           sched_enabled=sched_enabled, sched_time=sched_time,
                           sched_to=sched_to, sched_cc=sched_cc,
                           sched_last_status=sched_last_status)


# ── ACTIONS ───────────────────────────────────────────────────────────────────

@app.route("/action/sync-gsheet", methods=["POST"])
def sync_gsheet():
    sheet_id  = request.form.get("sheet_id", "").strip()
    worksheet = request.form.get("worksheet", "Sheet1").strip()
    if not sheet_id:
        flash("Please enter a Google Sheet ID.", "error")
        return jsonify({"ok": False, "redirect": url_for("page_data_uploader")})
    try:
        _set_progress(15, "Connecting to Google Sheets…")
        clean_id = google_sheets.extract_sheet_id(sheet_id)
        _set_progress(35, "Fetching master data…")
        df_sync  = google_sheets.fetch_master_data(clean_id, worksheet)
        now = _dt.now().isoformat(timespec="seconds")
        def _sv(v):
            """Safe string: NaN/None/blank → empty string."""
            import math
            if v is None: return ""
            if isinstance(v, float) and math.isnan(v): return ""
            s = str(v).strip()
            return "" if s.lower() == "nan" else s

        rows = [{"employee_code": _sv(r["Employee Code"]),
                 "employee_name": _sv(r["Employee Name"]),
                 "designation":   _sv(r["Designation"]),
                 "category":      _sv(r["Category"]),
                 "plant":         _sv(r["Plant"]),
                 "plant_code":    _sv(r["Plant Code"]),
                 "updated_at":    now}
                for _, r in df_sync.iterrows()]
        _set_progress(70, "Saving employee records…")
        inserted = database.replace_table_rows("master_data", rows)
        _set_progress(90, "Updating settings…")
        database.set_settings_bulk({"gsheet_id": clean_id, "gsheet_worksheet": worksheet,
                                    "gsheet_last_sync": now, "gsheet_last_count": str(inserted)})
        flash(f"Synced {inserted:,} employees from Google Sheets.", "success")
    except Exception as exc:
        flash(f"Sync failed: {exc}", "error")
    _set_progress(100, "Complete")
    return jsonify({"ok": True, "redirect": url_for("page_data_uploader")})


@app.route("/action/toggle-auto-sync", methods=["POST"])
def toggle_auto_sync():
    enabled = "enabled" in request.form
    database.set_setting("gsheet_auto_sync", "true" if enabled else "false")
    return redirect(url_for("page_data_uploader"))


@app.route("/action/add-employee", methods=["POST"])
def add_employee():
    emp = {
        "employee_code": request.form["code"].strip(),
        "employee_name": request.form.get("name", "").strip(),
        "designation":   request.form.get("designation", "").strip(),
        "category":      request.form.get("category", ""),
        "plant":         request.form.get("plant", "").strip(),
        "plant_code":    request.form.get("plant_code", "").strip(),
    }
    try:
        database.add_employee(
            emp["employee_code"], emp["employee_name"], emp["designation"],
            emp["category"], emp["plant"], emp["plant_code"],
        )
        flash(f"Added employee {emp['employee_code']}.", "success")
        res = google_sheets.push_id_employee_add(emp)
        if res["ok"]:
            flash(f"☁️ {res['message']}", "info")
        else:
            flash(f"⚠️ Saved locally but Google Sheet not updated: {res['message']}", "warning")
    except ValueError as e:
        flash(str(e), "error")
    if request.form.get("_next") == "sysconfig":
        return redirect(url_for("sysconfig_page") + "?open_modal=employee&tab=add")
    return redirect(url_for("page_data_uploader") + "?open_modal=employee&tab=add")


@app.route("/action/update-employee/<code>", methods=["POST"])
def update_employee(code):
    emp = {
        "employee_code": code,
        "employee_name": request.form.get("name", "").strip(),
        "designation":   request.form.get("designation", "").strip(),
        "category":      request.form.get("category", ""),
        "plant":         request.form.get("plant", "").strip(),
        "plant_code":    request.form.get("plant_code", "").strip(),
    }
    try:
        database.update_employee(
            code, emp["employee_name"], emp["designation"],
            emp["category"], emp["plant"], emp["plant_code"],
        )
        flash(f"Updated employee {code}.", "success")
        res = google_sheets.push_id_employee_update(emp)
        if res["ok"]:
            flash(f"☁️ {res['message']}", "info")
        else:
            flash(f"⚠️ Saved locally but Google Sheet not updated: {res['message']}", "warning")
    except ValueError as e:
        flash(str(e), "error")
    if request.form.get("_next") == "sysconfig":
        return redirect(url_for("sysconfig_page") + "?open_modal=employee&tab=edit")
    return redirect(url_for("page_data_uploader") + "?open_modal=employee&tab=edit")


@app.route("/action/delete-employee/<code>", methods=["POST"])
def delete_employee(code):
    try:
        database.delete_employee(code)
        flash(f"Deleted employee {code}.", "success")
        res = google_sheets.push_id_employee_delete(code)
        if res["ok"]:
            flash(f"☁️ {res['message']}", "info")
        else:
            flash(f"⚠️ Deleted locally but Google Sheet not updated: {res['message']}", "warning")
    except ValueError as e:
        flash(str(e), "error")
    if request.form.get("_next") == "sysconfig":
        return redirect(url_for("sysconfig_page") + "?open_modal=employee&tab=delete")
    return redirect(url_for("page_data_uploader") + "?open_modal=employee&tab=delete")


@app.route("/action/upload-backend", methods=["POST"])
def upload_backend():
    if "file" not in request.files or request.files["file"].filename == "":
        flash("No file selected.", "error")
        return redirect(url_for("page_data_uploader") + "#backend")
    f = request.files["file"]
    replace = request.form.get("mode", "replace") == "replace"
    try:
        df, warns = data_loader.load_backend_data(f)
        for w in warns:
            flash(w, "warning")
        saved = data_loader.save_backend_data(df, source_file=f.filename, replace=replace)
        flash(f"{'Replaced' if replace else 'Appended'} {saved:,} backend rows.", "success")
    except Exception as exc:
        flash(f"Upload failed: {exc}", "error")
    return redirect(url_for("page_data_uploader") + "#backend")


@app.route("/action/upload-maintenance", methods=["POST"])
def upload_maintenance():
    if "file" not in request.files or request.files["file"].filename == "":
        flash("No file selected.", "error")
        return redirect(url_for("page_data_uploader") + "#maintenance")
    f = request.files["file"]
    try:
        month = int(request.form.get("maint_month", _date.today().month))
        year  = int(request.form.get("maint_year",  _date.today().year))
    except (ValueError, TypeError):
        flash("Invalid month or year selected.", "error")
        return redirect(url_for("page_data_uploader") + "#maintenance")
    try:
        import calendar as _cal
        df, warns = data_loader.load_maintenance_cost(f)
        for w in warns:
            flash(w, "warning")
        saved = data_loader.save_maintenance_cost(df, month, year)
        flash(f"Saved {saved:,} plant maintenance cost rows for {_cal.month_name[month]} {year}.", "success")
    except Exception as exc:
        flash(f"Upload failed: {exc}", "error")
    return redirect(url_for("page_data_uploader") + "#maintenance")


@app.route("/action/assign-maintenance-month", methods=["POST"])
def assign_maintenance_month():
    try:
        month = int(request.form.get("month"))
        year  = int(request.form.get("year"))
        import calendar as _cal
        updated = database.assign_maintenance_month(month, year)
        flash(f"✅ Assigned {updated} existing rows to {_cal.month_name[month]} {year}.", "success")
    except Exception as exc:
        flash(f"Assignment failed: {exc}", "error")
    back = request.form.get("_next", "")
    if back == "sysconfig":
        return redirect(url_for("sysconfig_page") + "?m=maintenance")
    return redirect(url_for("page_data_uploader") + "#maintenance")


@app.route("/action/delete-maintenance-month", methods=["POST"])
def delete_maintenance_month():
    try:
        month = int(request.form.get("month"))
        year  = int(request.form.get("year"))
        import calendar as _cal
        deleted = database.delete_maintenance_month(month, year)
        flash(f"Deleted {deleted} maintenance cost rows for {_cal.month_name[month]} {year}.", "success")
    except Exception as exc:
        flash(f"Delete failed: {exc}", "error")
    back = request.form.get("_next", "")
    if back == "sysconfig":
        return redirect(url_for("sysconfig_page") + "?m=maintenance")
    return redirect(url_for("page_data_uploader") + "#maintenance")


@app.route("/action/fetch-oracle", methods=["POST"])
def fetch_oracle():
    from_s = request.form.get("from_date", str(_date.today().replace(day=1)))
    to_s   = request.form.get("to_date",   str(_date.today()))
    replace = request.form.get("mode", "replace") == "replace"
    try:
        fd = _date.fromisoformat(from_s)
        td = _date.fromisoformat(to_s)
        _set_progress(15, "Testing Oracle connection…")
        result = oracle_connector.test_connection()
        if not result["success"]:
            flash(f"Oracle connection failed: {result['error']}", "error")
            _set_progress(100, "Complete")
            return jsonify({"ok": False, "redirect": url_for("page_data_uploader") + "#oracle"})
        _set_progress(35, "Fetching data from Oracle…")
        df_ora, warns = oracle_connector.fetch_backend_data(fd, td)
        for w in warns:
            flash(w, "warning")
        _set_progress(70, "Saving to database…")
        if df_ora.empty:
            flash("Oracle returned no rows for the selected date range.", "warning")
        else:
            saved = oracle_connector.save_oracle_backend_data(df_ora, fd, td, replace=replace)
            _set_progress(90, "Cleaning old records…")
            database.purge_old_oracle_data()
            flash(f"{'Replaced' if replace else 'Appended'} {saved:,} rows from Oracle.", "success")
    except Exception as exc:
        flash(f"Oracle fetch failed: {exc}", "error")
    _set_progress(100, "Complete")
    return jsonify({"ok": True, "redirect": url_for("page_data_uploader") + "#oracle"})


@app.route("/action/preflight-check", methods=["GET"])
def preflight_check_api():
    """Return data-integrity check results for the given date range (JSON)."""
    from_s = request.args.get("from_date")
    to_s   = request.args.get("to_date")
    try:
        fd = _date.fromisoformat(from_s)
        td = _date.fromisoformat(to_s)
        errors = calculator.preflight_check(fd.month, fd.year, str(fd), str(td))
        return jsonify({"errors": errors})
    except Exception as exc:
        return jsonify({"errors": [str(exc)]})


@app.route("/action/run-calculation", methods=["POST"])
def run_calculation():
    from_s = request.form.get("from_date")
    to_s   = request.form.get("to_date")
    try:
        fd = _date.fromisoformat(from_s)
        td = _date.fromisoformat(to_s)
        if fd > td:
            flash("'From Date' must be on or before 'To Date'.", "error")
            return jsonify({"ok": False, "redirect": url_for("page_reports")})
        _set_progress(15, "Loading employee data…")
        result = calculator.run_calculation(
            month=fd.month, year=fd.year,
            start_date=str(fd), end_date=str(td),
        )
        _set_progress(90, "Saving results…")
        for pf_err in result.get("preflight_errors", []):
            flash(f"🚫 {pf_err}", "error")
        if result["error"] and not result.get("preflight_errors"):
            flash(f"Calculation failed: {result['error']}", "error")
        elif not result["error"]:
            flash(f"Calculation complete — {result['mapped']:,} employees, {result['unmapped']} unmapped.", "success")
            for w in result.get("calc_warnings", []):
                flash(w, "warning")
        _ss("calc_ran", True)
    except Exception as exc:
        flash(f"Calculation error: {exc}", "error")
    _set_progress(100, "Complete")
    return jsonify({"ok": True, "redirect": url_for("page_reports")})


@app.route("/action/run-validation", methods=["POST"])
def run_validation():
    try:
        errs: list = []
        _set_progress(15, "Validating master data…")
        validations._validate_master_data(errs)
        _set_progress(45, "Validating backend data…")
        validations._validate_backend_data(errs)
        _set_progress(70, "Validating maintenance costs…")
        validations._validate_maintenance_cost(errs)
        _set_progress(85, "Saving validation results…")
        validations.clear_validation_errors()
        if errs:
            database.insert_rows("validation_errors", errs)
        ran_at = _dt.now().isoformat(timespec="seconds")
        database.set_settings_bulk({
            "last_validation_at":     ran_at,
            "last_validation_errors": str(len(errs)),
        })
        flash(f"Validation complete — {len(errs)} error(s) found.", "success" if not errs else "warning")
    except Exception as exc:
        flash(f"Validation error: {exc}", "error")
    _set_progress(100, "Complete")
    return jsonify({"ok": True, "redirect": url_for("page_validation")})


@app.route("/validation/download-excel")
def download_validation_excel():
    """Download I&D validation errors as an Excel file."""
    from openpyxl import Workbook
    from openpyxl.styles import PatternFill, Font, Alignment
    from openpyxl.utils import get_column_letter
    import io

    all_err_df = database.read_table("validation_errors")
    errors = _records(all_err_df.drop(columns=["id"], errors="ignore")) \
             if not all_err_df.empty else []

    wb = Workbook()
    ws = wb.active
    ws.title = "Validation Errors"

    hdr_fill = PatternFill("solid", fgColor="1E3A5F")
    hdr_font = Font(bold=True, color="FFFFFF", size=11)
    cols    = ["source", "row_number", "column_name", "error_type", "error_message", "created_at"]
    headers = ["Source", "Row #", "Column", "Issue Type", "What's Wrong — What to Fix", "Detected At"]

    for ci, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=ci, value=h)
        cell.fill = hdr_fill
        cell.font = hdr_font
        cell.alignment = Alignment(horizontal="center", vertical="center")

    red_fill    = PatternFill("solid", fgColor="FFE0E0")
    yellow_fill = PatternFill("solid", fgColor="FFF3CD")
    for ri, row in enumerate(errors, 2):
        for ci, col in enumerate(cols, 1):
            cell = ws.cell(row=ri, column=ci, value=str(row.get(col, "") or ""))
            cell.alignment = Alignment(vertical="center", wrap_text=True)
        etype = str(row.get("error_type", "")).upper()
        fill  = yellow_fill if "UNMAPPED" in etype or "NO_DATA" in etype else red_fill
        for ci in range(1, len(cols) + 1):
            ws.cell(row=ri, column=ci).fill = fill

    col_widths = [18, 8, 18, 24, 65, 20]
    for ci, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(ci)].width = w
    ws.row_dimensions[1].height = 20

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    from flask import send_file
    return send_file(buf, as_attachment=True,
                     download_name="RDC_ID_Validation_Errors.xlsx",
                     mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")


@app.route("/tp/validation/download-excel")
def tp_download_validation_excel():
    """Download TP skip log + calc warnings as Excel."""
    from openpyxl import Workbook
    from openpyxl.styles import PatternFill, Font, Alignment
    from openpyxl.utils import get_column_letter
    import io

    skip_log      = _ms("tp", "skip_log", [])
    calc_warnings = _ms("tp", "calc_warnings", [])

    wb = Workbook()

    # Sheet 1 — Skip Log
    ws1 = wb.active
    ws1.title = "Skip Log"
    hdr_fill = PatternFill("solid", fgColor="1E3A5F")
    hdr_font = Font(bold=True, color="FFFFFF", size=11)
    skip_cols    = ["batch_ref", "plant_code", "production_date", "quantity", "time_taken_min", "reason"]
    skip_headers = ["Batch Ref", "Plant Code", "Date", "Quantity", "Time (min)", "Reason"]
    for ci, h in enumerate(skip_headers, 1):
        c = ws1.cell(row=1, column=ci, value=h)
        c.fill = hdr_fill; c.font = hdr_font
        c.alignment = Alignment(horizontal="center")
    row_fill = PatternFill("solid", fgColor="FFE0E0")
    for ri, row in enumerate(skip_log, 2):
        for ci, col in enumerate(skip_cols, 1):
            cell = ws1.cell(row=ri, column=ci, value=str(row.get(col, "") or ""))
            cell.fill = row_fill
    for ci, w in enumerate([18, 14, 14, 12, 12, 40], 1):
        ws1.column_dimensions[get_column_letter(ci)].width = w

    # Sheet 2 — Calc Warnings
    ws2 = wb.create_sheet("Calc Warnings")
    ws2.cell(row=1, column=1, value="Warning").fill = hdr_fill
    ws2.cell(row=1, column=1).font = hdr_font
    ws2.column_dimensions["A"].width = 80
    warn_fill = PatternFill("solid", fgColor="FFF3CD")
    for ri, w in enumerate(calc_warnings, 2):
        cell = ws2.cell(row=ri, column=1, value=w)
        cell.fill = warn_fill
        cell.alignment = Alignment(wrap_text=True)

    buf = io.BytesIO()
    wb.save(buf); buf.seek(0)
    from flask import send_file
    return send_file(buf, as_attachment=True,
                     download_name="RDC_TP_Validation_SkipLog.xlsx",
                     mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")


@app.route("/btrtp/validation/download-excel")
def btrtp_download_validation_excel():
    """Download BTRTP skip log + calc warnings as Excel."""
    from openpyxl import Workbook
    from openpyxl.styles import PatternFill, Font, Alignment
    from openpyxl.utils import get_column_letter
    import io

    skip_log      = _ms("btrtp", "skip_log", [])
    calc_warnings = _ms("btrtp", "calc_warnings", [])

    wb = Workbook()

    ws1 = wb.active
    ws1.title = "Skip Log"
    hdr_fill = PatternFill("solid", fgColor="1E3A5F")
    hdr_font = Font(bold=True, color="FFFFFF", size=11)
    skip_cols    = ["batcher_id", "batch_ref", "plant_code", "production_date", "quantity", "time_taken_min", "reason"]
    skip_headers = ["Batcher ID", "Batch Ref", "Plant Code", "Date", "Quantity", "Time (min)", "Reason"]
    for ci, h in enumerate(skip_headers, 1):
        c = ws1.cell(row=1, column=ci, value=h)
        c.fill = hdr_fill; c.font = hdr_font
        c.alignment = Alignment(horizontal="center")
    row_fill = PatternFill("solid", fgColor="FFE0E0")
    for ri, row in enumerate(skip_log, 2):
        for ci, col in enumerate(skip_cols, 1):
            cell = ws1.cell(row=ri, column=ci, value=str(row.get(col, "") or ""))
            cell.fill = row_fill
    for ci, w in enumerate([16, 18, 14, 14, 12, 12, 40], 1):
        ws1.column_dimensions[get_column_letter(ci)].width = w

    ws2 = wb.create_sheet("Calc Warnings")
    ws2.cell(row=1, column=1, value="Warning").fill = hdr_fill
    ws2.cell(row=1, column=1).font = hdr_font
    ws2.column_dimensions["A"].width = 80
    warn_fill = PatternFill("solid", fgColor="FFF3CD")
    for ri, w in enumerate(calc_warnings, 2):
        cell = ws2.cell(row=ri, column=1, value=w)
        cell.fill = warn_fill
        cell.alignment = Alignment(wrap_text=True)

    buf = io.BytesIO()
    wb.save(buf); buf.seek(0)
    from flask import send_file
    return send_file(buf, as_attachment=True,
                     download_name="RDC_BTRTP_Validation_SkipLog.xlsx",
                     mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")


@app.route("/action/clear-validation", methods=["POST"])
def clear_validation():
    validations.clear_validation_errors()
    database.set_settings_bulk({"last_validation_at": "", "last_validation_errors": "0"})
    flash("All validation errors cleared.", "success")
    return redirect(url_for("page_validation"))


@app.route("/action/save-smtp", methods=["POST"])
def save_smtp():
    pwd = request.form.get("password", "").strip()
    to_save = {
        "smtp_host":        request.form.get("host", "").strip(),
        "smtp_port":        request.form.get("port", "587").strip(),
        "smtp_sender":      request.form.get("sender", "").strip(),
        "smtp_use_tls":     "true" if "use_tls" in request.form else "false",
        "email_default_to": request.form.get("default_to", "").strip(),
        "email_default_cc": request.form.get("default_cc", "").strip(),
        "email_subject":    request.form.get("subject", "").strip(),
    }
    if pwd:
        to_save["smtp_password"] = pwd
    database.set_settings_bulk(to_save)
    flash("Email settings saved.", "success")
    return redirect(url_for("page_settings"))


@app.route("/action/test-smtp", methods=["POST"])
def test_smtp():
    cfg = email_helper.get_smtp_config()
    try:
        res = email_helper.send_report_email(
            to_emails=cfg["sender"], cc_emails="",
            subject="Test email — Batching Incentive Calculator",
            body="This is a test email confirming your SMTP settings work.",
        )
        if res["success"]:
            flash(f"Test email sent to {cfg['sender']}. Check the inbox.", "success")
        else:
            flash(f"Test failed: {res['error']}", "error")
    except Exception as exc:
        flash(f"Test failed: {exc}", "error")
    return redirect(url_for("page_settings"))


@app.route("/action/save-oracle", methods=["POST"])
def save_oracle():
    pwd = request.form.get("password", "").strip()
    to_save = {
        "oracle_host":              request.form.get("host", "").strip(),
        "oracle_port":              request.form.get("port", "").strip(),
        "oracle_service":           request.form.get("service", "").strip(),
        "oracle_user":              request.form.get("user", "").strip(),
        "oracle_status_filter":     request.form.get("status_filter", "").strip(),
        "oracle_instantclient_dir": request.form.get("instantclient", "").strip(),
    }
    if pwd:
        to_save["oracle_password"] = pwd
    database.set_settings_bulk(to_save)
    flash("Oracle settings saved.", "success")
    return redirect(url_for("page_settings"))


@app.route("/action/test-oracle", methods=["POST"])
def test_oracle():
    try:
        result = oracle_connector.test_connection()
        if result["success"]:
            flash(f"Connection successful! {result.get('version', '')}", "success")
        else:
            flash(f"Connection failed: {result['error']}", "error")
    except Exception as exc:
        flash(f"Test failed: {exc}", "error")
    return redirect(url_for("page_settings"))


@app.route("/action/save-bg-settings", methods=["POST"])
def save_bg_settings():
    auto    = "auto_theme" in request.form
    animate = "animate"    in request.form
    theme   = request.form.get("manual_theme", "Daytime")
    database.set_settings_bulk({
        "bg_auto_theme":    "true" if auto    else "false",
        "bg_animate":       "true" if animate else "false",
        "bg_manual_theme":  theme,
    })
    flash("Background settings saved.", "success")
    return redirect(url_for("page_settings"))


@app.route("/action/clear-cache", methods=["POST"])
def clear_cache():
    today = str(_date.today())
    last  = database.get_setting("last_cache_clear_date", "")
    if last == today:
        flash("Cache already compacted today — come back tomorrow.", "warning")
        return redirect(url_for("page_settings"))
    try:
        summary = cache_helpers.clear_cache()
        if summary["error"]:
            flash(f"Could not compact database: {summary['error']}", "error")
        else:
            database.set_setting("last_cache_clear_date", today)
            freed = summary["freed_mb"]
            note  = f"Freed {freed} MB." if freed > 0 else "Database was already compact."
            flash(f"Done! {note} Database size now {summary['after_mb']} MB.", "success")
    except Exception as exc:
        flash(f"Cache clear failed: {exc}", "error")
    return redirect(url_for("page_settings"))


# ── AJAX endpoint (topbar theme) ─────────────────────────────────────────────

@app.route("/api/bg-settings", methods=["POST"])
def api_bg_settings():
    data = request.get_json(silent=True) or {}
    to_save = {}
    if "auto_theme" in data:
        to_save["bg_auto_theme"] = "true" if data.get("auto_theme") else "false"
    if "manual_theme" in data:
        to_save["bg_manual_theme"] = data.get("manual_theme", "Daytime")
    if "animate" in data:
        to_save["bg_animate"] = "true" if data.get("animate") else "false"
    if to_save:
        database.set_settings_bulk(to_save)
    return jsonify({"ok": True})


# ── Oracle live status (shared by all modules + launcher icon) ───────────────
@app.route("/api/oracle-status")
def api_oracle_status():
    """Live, honest Oracle status. Used by the launcher icon and every module
    topbar. One shared Oracle connection serves all four modules."""
    cfg = oracle_connector.get_oracle_config()
    configured = oracle_connector.is_configured(cfg)
    reachable  = oracle_connector.is_reachable(cfg) if configured else False
    if not configured:
        state, label = "unconfigured", "Oracle not configured"
    elif reachable:
        state, label = "connected", f"Oracle connected — {cfg['host']}"
    else:
        state, label = "unreachable", "Oracle unreachable (check office network / VPN)"
    return jsonify({"configured": configured, "reachable": reachable,
                    "state": state, "label": label})


# ── Real-time progress polling endpoint ───────────────────────────────────────
@app.route("/api/progress")
def api_progress():
    with _progress_lock:
        return jsonify(dict(_progress))


@app.route("/tp/api/table-columns")
def tp_api_table_columns():
    """Return all column names in rdc_batch_trx_headers so the user can find the right plant column."""
    try:
        cols = oracle_connector.get_table_columns()
        return jsonify({"ok": True, "columns": cols})
    except Exception as exc:
        return jsonify({"ok": False, "error": str(exc)})


# ── DOWNLOAD ENDPOINTS ────────────────────────────────────────────────────────

def _snapshot_dfs():
    filtered = _s("rpt_filtered", [])
    unmapped  = _s("rpt_unmapped_snap", [])
    from_s    = _s("rpt_from",  str(_date.today()))
    to_s      = _s("rpt_to",    str(_date.today()))
    applied   = _s("rpt_applied_filters", "None")
    meta      = _build_meta(from_s, to_s, filtered, unmapped, applied)

    col_set = set(RESULT_COLS) | {"category", "month", "year"}
    all_cols = [c for c in (RESULT_COLS + ["category", "month", "year"])
                if c not in ("_cls",)]

    df_f = pd.DataFrame(filtered) if filtered else pd.DataFrame()
    df_u = pd.DataFrame(unmapped) if unmapped else pd.DataFrame()
    val_df = database.read_table("validation_errors")
    return df_f, df_u, val_df, meta, from_s, to_s


@app.route("/download/excel")
def download_excel():
    df_f, df_u, val_df, meta, from_s, to_s = _snapshot_dfs()
    if df_f.empty:
        flash("No report data. Load a report on the View Reports page first.", "warning")
        return redirect(url_for("page_reports"))
    try:
        xlsx = report_generator.generate_excel_report(df_f, df_u, val_df, meta)
        fname = f"incentive_report_{from_s}_to_{to_s}.xlsx"
        return send_file(
            io.BytesIO(xlsx), as_attachment=True,
            download_name=fname,
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
    except Exception as exc:
        flash(f"Could not build Excel: {exc}", "error")
        return redirect(url_for("page_reports"))


@app.route("/download/csv")
def download_csv():
    filtered = _s("rpt_filtered", [])
    from_s   = _s("rpt_from",  str(_date.today()))
    to_s     = _s("rpt_to",    str(_date.today()))
    if not filtered:
        flash("No report data. Load a report on the View Reports page first.", "warning")
        return redirect(url_for("page_reports"))
    df = pd.DataFrame(filtered)
    df_show = df[[c for c in RESULT_COLS if c in df.columns]].rename(columns=RESULT_LABELS)
    csv_bytes = df_show.to_csv(index=False).encode("utf-8")
    fname = f"incentive_report_{from_s}_to_{to_s}.csv"
    return send_file(io.BytesIO(csv_bytes), as_attachment=True,
                     download_name=fname, mimetype="text/csv")


@app.route("/action/add-waiver", methods=["POST"])
def action_add_waiver():
    emp  = request.form.get("waiver_emp_code", "").strip()
    mon  = request.form.get("waiver_month", "").strip()
    yr   = request.form.get("waiver_year", "").strip()
    rsn  = request.form.get("waiver_reason", "").strip()
    cust = request.form.get("waiver_custom", "").strip()
    if not emp or not mon or not yr or not rsn:
        flash("All waiver fields are required.", "error")
        return redirect(url_for("page_reports") + "#waivers")
    try:
        database.upsert_waiver(emp, int(mon), int(yr), rsn, cust)
        flash(f"✅ Waiver saved for {emp}.", "success")
    except Exception as e:
        flash(f"Could not save waiver: {e}", "error")
    return redirect(url_for("page_reports") + "#waivers")


@app.route("/action/delete-waiver", methods=["POST"])
def action_delete_waiver():
    wid = request.form.get("waiver_id", "").strip()
    if not wid:
        flash("Invalid waiver ID.", "error")
        return redirect(url_for("page_reports") + "#waivers")
    database.delete_waiver(int(wid))
    flash("Waiver removed.", "success")
    return redirect(url_for("page_reports") + "#waivers")


@app.route("/action/send-email", methods=["POST"])
def send_email():
    df_f, df_u, val_df, meta, from_s, to_s = _snapshot_dfs()
    if df_f.empty:
        flash("No report data. Load a report on the View Reports page first.", "warning")
        return redirect(url_for("page_reports"))

    to_addr  = request.form.get("to", "")
    cc_addr  = request.form.get("cc", "")
    subject  = request.form.get("subject", "")
    body     = request.form.get("body", "")
    incl_tables = "include_tables" in request.form

    try:
        fname       = f"incentive_report_{from_s}_to_{to_s}.xlsx"
        xlsx_data   = report_generator.generate_excel_report(df_f, df_u, val_df, meta)
        tables_html = report_generator.build_email_tables_html(df_f)
        import html as _html_mod
        _safe_body  = _html_mod.escape(body or "").replace("\n", "<br>")
        _body_font  = "font-family:Arial,Calibri,sans-serif;font-size:12px;color:#000000;"
        html_body   = (
            f'<html><body style="margin:0;padding:8px 10px;{_body_font}">'
            f'<p style="margin:0 0 3px 0;">{_safe_body}</p>'
            f'{tables_html}'
            f'</body></html>'
        )

        res = email_helper.send_report_email(
            to_emails=to_addr, cc_emails=cc_addr,
            subject=subject, body=body,
            attachment_bytes=xlsx_data, attachment_name=fname,
            html_body=html_body,
        )
        if res["success"]:
            flash(f"Report emailed to {to_addr}.", "success")
        else:
            flash(f"Email failed: {res['error']}", "error")
    except Exception as exc:
        flash(f"Could not send: {exc}", "error")

    return redirect(url_for("page_reports"))


@app.route("/action/restart-server", methods=["POST"])
def restart_server():
    def _restart():
        import time
        time.sleep(1)
        # Spawn a new process first, then exit this one (works on Windows)
        subprocess.Popen([sys.executable] + sys.argv,
                         creationflags=subprocess.CREATE_NEW_PROCESS_GROUP if sys.platform == "win32" else 0)
        os._exit(0)
    threading.Thread(target=_restart, daemon=True).start()
    return jsonify({"ok": True})


@app.route("/action/save-email-schedule", methods=["POST"])
def save_email_schedule():
    sched_time = request.form.get("sched_time", "08:00").strip()
    sched_to   = request.form.get("sched_to",   "").strip()
    sched_cc   = request.form.get("sched_cc",   "").strip()
    database.set_settings_bulk({
        "email_schedule_time": sched_time,
        "email_schedule_to":   sched_to,
        "email_schedule_cc":   sched_cc,
    })
    if _scheduler:
        try:
            h, m = map(int, sched_time.split(":"))
        except Exception:
            h, m = 8, 0
        _scheduler.reschedule_job(
            "monthly_report",
            trigger=CronTrigger(day=1, hour=h, minute=m),
        )
    flash("Schedule settings saved.", "success")
    return redirect(url_for("page_settings"))


@app.route("/action/toggle-email-schedule", methods=["POST"])
def toggle_email_schedule():
    current = database.get_setting("email_schedule_enabled", "false")
    new_val = "false" if current == "true" else "true"
    database.set_setting("email_schedule_enabled", new_val)
    flash(f"Scheduled monthly report {'enabled' if new_val == 'true' else 'disabled'}.", "success")
    return redirect(url_for("page_settings"))


# ═══════════════════════════════════════════════════════════════════════════
# RDC-ECMD MODULE — Energy Consumption & Mixer DG Ratio
# ═══════════════════════════════════════════════════════════════════════════

def _ecmd_ctx():
    """Base context dict for every ECMD page."""
    return dict(
        active_page="",
        bg_auto=bg_auto(), bg_animate=bg_animate(), bg_theme=bg_theme(),
    )


def _ecmd_mon_tag(month, year):
    import calendar as _cal
    return f"{_cal.month_abbr[month]}'{str(year)[2:]}"


@app.route("/ecmd")
@auth.login_required
def page_ecmd():
    return redirect(url_for("ecmd_dashboard"))


# ── Dashboard ─────────────────────────────────────────────────────────────────
@app.route("/ecmd/dashboard")
@auth.login_required
def ecmd_dashboard():
    all_counts  = database.get_table_counts()
    counts = {
        "ecmd_readings": all_counts.get("ecmd_readings", 0),
        "ecmd_results":  all_counts.get("ecmd_results",  0),
        "tp_plant_data": all_counts.get("tp_plant_data", 0),
    }
    ora_ready   = oracle_connector.is_configured()
    ecmd_months = database.get_ecmd_months()
    ctx = _ecmd_ctx()
    ctx["active_page"] = "dashboard"
    return render_template("ecmd_dashboard.html", counts=counts,
                           ora_ready=ora_ready, ecmd_months=ecmd_months, **ctx)


# ── Data Entry ────────────────────────────────────────────────────────────────
@app.route("/ecmd/data-entry", methods=["GET", "POST"])
@auth.login_required
def ecmd_data_entry():
    import calendar as _cal

    today = _date.today()
    sel_month = int(request.args.get("month", today.month))
    sel_year  = int(request.args.get("year",  today.year))

    all_plant_rows = database.get_tp_plants()
    plant_rows = auth.apply_plant_filter_rows(all_plant_rows, g.current_user)
    readings_map = {r["plant_code"]: r
                    for r in database.get_ecmd_readings_for_month(sel_month, sel_year)}

    # Daily readings: map plant_code → list of day rows
    daily_rows_all = database.get_ecmd_daily_readings_for_month(sel_month, sel_year)
    daily_map = {}
    for dr in daily_rows_all:
        daily_map.setdefault(dr["plant_code"], []).append(dr)

    # Entry mode per plant
    mode_map = {p["plant_code"]: database.get_ecmd_entry_mode(p["plant_code"], sel_month, sel_year)
                for p in plant_rows}

    # Days in the selected month
    days_in_month = _cal.monthrange(sel_year, sel_month)[1]

    # MF per plant — latest saved value (admin sets it, users see it read-only)
    mf_map = {p["plant_code"]: database.get_ecmd_mf(p["plant_code"]) for p in plant_rows}

    is_admin = g.current_user.get("role") == auth.SUPER_ADMIN

    # Allowed months: admin sets which months users can enter data for.
    # If list is empty AND user is not admin → restrict to current month only.
    allowed_months = database.get_ecmd_allowed_months()  # [(month, year), ...]
    if not is_admin and not allowed_months:
        # Default restriction: only current month accessible for non-admins
        allowed_months = [(today.month, today.year)]
    month_locked   = not is_admin  # non-admins always see restricted selector
    month_allowed  = is_admin or (sel_month, sel_year) in allowed_months

    months_list = [(m, _cal.month_name[m]) for m in range(1, 13)]
    years_list  = list(range(today.year - 2, today.year + 2))

    ctx = _ecmd_ctx()
    ctx["active_page"] = "data_entry"
    return render_template("ecmd_data_entry.html",
                           plant_rows=plant_rows,
                           readings_map=readings_map,
                           daily_map=daily_map,
                           mode_map=mode_map,
                           mf_map=mf_map,
                           is_admin=is_admin,
                           allowed_months=allowed_months,
                           month_locked=month_locked,
                           month_allowed=month_allowed,
                           days_in_month=days_in_month,
                           sel_month=sel_month, sel_year=sel_year,
                           months_list=months_list, years_list=years_list,
                           **ctx)


@app.route("/ecmd/action/save-reading", methods=["POST"])
@auth.login_required
def ecmd_save_reading():
    plant_code = request.form.get("plant_code", "").strip()
    month      = int(request.form.get("month", _date.today().month))
    year       = int(request.form.get("year",  _date.today().year))
    user = g.current_user
    # Month freeze check
    if user.get("role") != auth.SUPER_ADMIN:
        allowed = database.get_ecmd_allowed_months()
        if allowed and (month, year) not in allowed:
            flash(f"Data entry for {month}/{year} is not open. Please contact admin.", "error")
            return redirect(url_for("ecmd_data_entry", month=month, year=year))
    # Plant-level access check
    plant_row = next((p for p in database.get_tp_plants() if p["plant_code"] == plant_code), None)
    if plant_row and not auth.user_can_access_plant(user, plant_row["plant_name"]):
        flash("You are not authorised to save readings for this plant.", "error")
        return redirect(url_for("ecmd_data_entry", month=month, year=year))

    def _fv(key):
        v = request.form.get(key, "").strip()
        return float(v) if v != "" else None

    data = {
        "eb_kwh_open":        _fv("eb_kwh_open"),
        "eb_kwh_close":       _fv("eb_kwh_close"),
        "eb_kvah_open":       _fv("eb_kvah_open"),
        "eb_kvah_close":      _fv("eb_kvah_close"),
        "mf":                 _fv("mf") or 1.0,
        "dg_hr_open":         _fv("dg_hr_open"),
        "dg_hr_close":        _fv("dg_hr_close"),
        "dg_kwh_open":        _fv("dg_kwh_open"),
        "dg_kwh_close":       _fv("dg_kwh_close"),
        "mixer_dg_hr_open":   _fv("mixer_dg_hr_open"),
        "mixer_dg_hr_close":  _fv("mixer_dg_hr_close"),
        "diesel_issued_ltrs": _fv("diesel_issued_ltrs"),
        "volume_on_dg":       _fv("volume_on_dg"),
    }
    if not plant_code:
        flash("Plant Code is required.", "error")
        return redirect(url_for("ecmd_data_entry", month=month, year=year))
    try:
        database.upsert_ecmd_reading(plant_code, month, year, data)
        flash(f"✅ Readings saved for plant {plant_code} — {month}/{year}.", "success")
    except Exception as exc:
        flash(f"Save failed: {exc}", "error")
    return redirect(url_for("ecmd_data_entry", month=month, year=year))


@app.route("/ecmd/action/delete-reading", methods=["POST"])
@auth.login_required
def ecmd_delete_reading():
    plant_code = request.form.get("plant_code", "").strip()
    month      = int(request.form.get("month", _date.today().month))
    year       = int(request.form.get("year",  _date.today().year))
    n = database.delete_ecmd_reading(plant_code, month, year)
    if n:
        flash(f"✅ Reading for {plant_code} ({month}/{year}) deleted.", "success")
    else:
        flash("Reading not found.", "warning")
    return redirect(url_for("ecmd_data_entry", month=month, year=year))


@app.route("/ecmd/action/set-entry-mode", methods=["POST"])
@auth.login_required
def ecmd_set_entry_mode():
    plant_code = request.form.get("plant_code", "").strip()
    month      = int(request.form.get("month", _date.today().month))
    year       = int(request.form.get("year",  _date.today().year))
    mode       = request.form.get("mode", "monthly")
    if mode not in ("monthly", "daily", "none"):
        flash("Invalid mode.", "error")
        return redirect(url_for("ecmd_data_entry", month=month, year=year))
    # Block if data already exists in the OTHER mode
    mth_row     = database.get_ecmd_reading(plant_code, month, year)
    # A monthly row created only for diesel/volume totals (all meter cols None) is not "monthly data"
    _meter_keys = ("eb_kwh_open","eb_kwh_close","eb_kvah_open","eb_kvah_close",
                   "dg_hr_open","dg_hr_close","dg_kwh_open","dg_kwh_close",
                   "mixer_dg_hr_open","mixer_dg_hr_close")
    has_monthly = bool(mth_row) and any(mth_row.get(k) is not None for k in _meter_keys)
    has_daily   = bool(database.get_ecmd_daily_readings(plant_code, month, year))
    if mode == "daily" and has_monthly:
        flash("Monthly data already saved — delete it before switching to Day Wise.", "error")
        return redirect(url_for("ecmd_data_entry", month=month, year=year))
    if mode == "monthly" and has_daily:
        flash("Day-wise data already saved — delete all days before switching to Monthly.", "error")
        return redirect(url_for("ecmd_data_entry", month=month, year=year))
    database.set_ecmd_entry_mode(plant_code, month, year, mode)
    return redirect(url_for("ecmd_data_entry", month=month, year=year))


@app.route("/ecmd/action/save-daily-reading", methods=["POST"])
@auth.login_required
def ecmd_save_daily_reading():
    plant_code = request.form.get("plant_code", "").strip()
    month      = int(request.form.get("month", _date.today().month))
    year       = int(request.form.get("year",  _date.today().year))
    day        = int(request.form.get("day", 1))
    user = g.current_user
    # Month freeze check — non-admins blocked from saving in locked months
    if user.get("role") != auth.SUPER_ADMIN:
        allowed = database.get_ecmd_allowed_months()
        if allowed and (month, year) not in allowed:
            msg = f"Data entry for {month}/{year} is not open. Please contact admin."
            if request.headers.get("X-Requested-With") == "XMLHttpRequest":
                from flask import jsonify
                return jsonify({"error": msg}), 403
            flash(msg, "error")
            return redirect(url_for("ecmd_data_entry", month=month, year=year))
    plant_row = next((p for p in database.get_tp_plants() if p["plant_code"] == plant_code), None)
    if plant_row and not auth.user_can_access_plant(user, plant_row["plant_name"]):
        flash("You are not authorised to save readings for this plant.", "error")
        return redirect(url_for("ecmd_data_entry", month=month, year=year))

    def _fv(key):
        v = request.form.get(key, "").strip()
        return float(v) if v != "" else None

    data = {
        "eb_kwh_open":        _fv("eb_kwh_open"),
        "eb_kwh_close":       _fv("eb_kwh_close"),
        "eb_kvah_open":       _fv("eb_kvah_open"),
        "eb_kvah_close":      _fv("eb_kvah_close"),
        "mf":                 _fv("mf") or 1.0,
        "dg_hr_open":         _fv("dg_hr_open"),
        "dg_hr_close":        _fv("dg_hr_close"),
        "dg_kwh_open":        _fv("dg_kwh_open"),
        "dg_kwh_close":       _fv("dg_kwh_close"),
        "mixer_dg_hr_open":   _fv("mixer_dg_hr_open"),
        "mixer_dg_hr_close":  _fv("mixer_dg_hr_close"),
        "diesel_issued_ltrs": _fv("diesel_issued_ltrs"),
        "volume_on_dg":       _fv("volume_on_dg"),
    }
    # day=0 means this is the Diesel/Volume totals-only form
    if day == 0:
        diesel = _fv("diesel_issued_ltrs")
        vol    = _fv("volume_on_dg")
        existing = database.get_ecmd_reading(plant_code, month, year)
        monthly_data = {
            "eb_kwh_open": existing.get("eb_kwh_open") if existing else None,
            "eb_kwh_close": existing.get("eb_kwh_close") if existing else None,
            "eb_kvah_open": existing.get("eb_kvah_open") if existing else None,
            "eb_kvah_close": existing.get("eb_kvah_close") if existing else None,
            "mf": (existing.get("mf") or 1.0) if existing else 1.0,
            "dg_hr_open": existing.get("dg_hr_open") if existing else None,
            "dg_hr_close": existing.get("dg_hr_close") if existing else None,
            "dg_kwh_open": existing.get("dg_kwh_open") if existing else None,
            "dg_kwh_close": existing.get("dg_kwh_close") if existing else None,
            "mixer_dg_hr_open": existing.get("mixer_dg_hr_open") if existing else None,
            "mixer_dg_hr_close": existing.get("mixer_dg_hr_close") if existing else None,
            "diesel_issued_ltrs": diesel,
            "volume_on_dg": vol,
        }
        try:
            database.upsert_ecmd_reading(plant_code, month, year, monthly_data,
                                         entered_by=user.get("username", ""))
            flash(f"✅ Diesel & Volume totals saved for {plant_code}.", "success")
        except Exception as exc:
            flash(f"Save failed: {exc}", "error")
        return redirect(url_for("ecmd_data_entry", month=month, year=year))

    # Regular day row
    try:
        database.upsert_ecmd_daily_reading(plant_code, month, year, day, data,
                                           entered_by=user.get("username", ""))
        flash(f"✅ Day {day} readings saved for {plant_code}.", "success")
    except Exception as exc:
        flash(f"Save failed: {exc}", "error")
    return redirect(url_for("ecmd_data_entry", month=month, year=year))


@app.route("/ecmd/action/delete-daily-reading", methods=["POST"])
@auth.login_required
def ecmd_delete_daily_reading():
    plant_code = request.form.get("plant_code", "").strip()
    month      = int(request.form.get("month", _date.today().month))
    year       = int(request.form.get("year",  _date.today().year))
    day        = int(request.form.get("day", 1))
    database.delete_ecmd_daily_reading(plant_code, month, year, day)
    flash(f"Day {day} reading deleted.", "success")
    return redirect(url_for("ecmd_data_entry", month=month, year=year))


@app.route("/ecmd/action/delete-all-daily-readings", methods=["POST"])
@auth.login_required
def ecmd_delete_all_daily_readings():
    plant_code = request.form.get("plant_code", "").strip()
    month      = int(request.form.get("month", _date.today().month))
    year       = int(request.form.get("year",  _date.today().year))
    database.delete_ecmd_all_daily_readings(plant_code, month, year)
    database.delete_ecmd_reading(plant_code, month, year)
    database.set_ecmd_entry_mode(plant_code, month, year, "monthly")
    flash(f"All daily readings deleted for {plant_code} {month}/{year}.", "success")
    return redirect(url_for("ecmd_data_entry", month=month, year=year))


# ── Calculate ─────────────────────────────────────────────────────────────────
@app.route("/ecmd/calculate", methods=["GET", "POST"])
@auth.login_required
def ecmd_calculate():
    # Unified page: calculate → redirect to reports with dates in query params
    today = _date.today()
    if request.method == "POST":
        from_s = request.form.get("from_date", str(today.replace(day=1)))
        to_s   = request.form.get("to_date",   str(today))
        return redirect(url_for("ecmd_reports", from_date=from_s, to_date=to_s))
    return redirect(url_for("ecmd_reports",
                            from_date=str(today.replace(day=1)),
                            to_date=str(today)))


# ── Reports ───────────────────────────────────────────────────────────────────
@app.route("/ecmd/reports")
@auth.login_required
def ecmd_reports():
    import calendar as _cal

    today = _date.today()
    from_s = request.args.get("from_date",
                              _ms("ecmd", "calc_from", str(today.replace(day=1))))
    to_s   = request.args.get("to_date",
                              _ms("ecmd", "calc_to", str(today)))
    try:
        fd = _date.fromisoformat(from_s)
        td = _date.fromisoformat(to_s)
    except ValueError:
        fd, td = today.replace(day=1), today

    if "from_date" in request.args:
        month, year = fd.month, fd.year
        all_plants, warnings = ecmd_calculator.run_ecmd_calculation(
            month, year, from_date=str(fd), to_date=str(td))
        loc_rows = ecmd_calculator.build_location_summary(all_plants)
        _mss("ecmd", "plant_rows",    all_plants)
        _mss("ecmd", "loc_rows",      loc_rows)
        _mss("ecmd", "calc_warnings", warnings)
        _mss("ecmd", "calc_from",     str(fd))
        _mss("ecmd", "calc_to",       str(td))
        _mss("ecmd", "calc_month",    month)
        _mss("ecmd", "calc_year",     year)
    else:
        all_plants = _ms("ecmd", "plant_rows", [])
        loc_rows   = _ms("ecmd", "loc_rows",   [])

    # Filters
    excos  = request.args.getlist("exco")
    bheads = request.args.getlist("bhead")
    search = request.args.get("search", "")

    unique_excos  = sorted({r.get("exco_location", "") for r in all_plants if r.get("exco_location")})
    unique_bheads = sorted({r.get("business_head", "") for r in all_plants if r.get("business_head")})

    plant_rows = all_plants
    if excos:
        plant_rows = [r for r in plant_rows if r.get("exco_location") in excos]
    if bheads:
        plant_rows = [r for r in plant_rows if r.get("business_head") in bheads]
    if search.strip():
        s = search.strip().lower()
        plant_rows = [r for r in plant_rows
                      if s in str(r.get("plant_code", "")).lower()
                      or s in str(r.get("plant_name", "")).lower()]

    if plant_rows is not all_plants:
        loc_rows = ecmd_calculator.build_location_summary(plant_rows)

    _mss("ecmd", "report_plant_rows", plant_rows)
    _mss("ecmd", "report_loc_rows",   loc_rows)

    smtp        = email_helper.get_smtp_config()
    email_ready = email_helper.is_configured()
    month = _ms("ecmd", "calc_month", fd.month)
    year  = _ms("ecmd", "calc_year",  fd.year)

    _elog = database.read_table("email_log", order_by="id DESC")
    if not _elog.empty and "report_file_name" in _elog.columns:
        _elog = _elog[_elog["report_file_name"].str.startswith("RDC_ECMD_", na=False)]
    ecmd_email_log = _records(_elog.drop(columns=["id"], errors="ignore").head(20)) \
        if not _elog.empty else []

    if (fd.year, fd.month) == (td.year, td.month):
        month_label = f"{_cal.month_name[fd.month]} {fd.year}"
    else:
        month_label = f"{fd:%d %b %Y} → {td:%d %b %Y}"

    default_to      = database.get_module_setting("ecmd", "email_default_to", "") or smtp.get("default_to", "")
    default_cc      = database.get_module_setting("ecmd", "email_default_cc", "") or smtp.get("default_cc", "")
    default_subject = (database.get_module_setting("ecmd", "email_default_subject", "")
                       or f"RDC-ECMD Energy & DG Report — {month_label}")
    default_body    = (database.get_module_setting("ecmd", "email_default_body", "")
                       or f"Dear Team,\n\nPlease find attached the Energy Consumption & "
                          f"Mixer DG Ratio Report for {month_label}.\n\nRegards,\nRDC Operations")

    ctx = _ecmd_ctx()
    ctx["active_page"] = "reports"
    return render_template("ecmd_reports.html",
                           plant_rows=plant_rows, loc_rows=loc_rows,
                           total_plants=len(all_plants),
                           from_date=str(fd), to_date=str(td),
                           unique_excos=unique_excos, unique_bheads=unique_bheads,
                           excos=excos, bheads=bheads, search=search,
                           month_label=month_label,
                           email_cfg=smtp, email_ready=email_ready,
                           default_to=default_to, default_cc=default_cc,
                           default_subject=default_subject, default_body=default_body,
                           ecmd_email_log=ecmd_email_log, **ctx)


def _ecmd_build_excel(plant_rows, loc_rows, month, year):
    """Build colour-coded Excel report for ECMD (Energy + DG tabs)."""
    from openpyxl import Workbook
    from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    mon_tag    = _ecmd_mon_tag(month, year)
    HDR_FILL   = PatternFill("solid", fgColor="0A2540")
    PAN_FILL   = PatternFill("solid", fgColor="D9D9D9")
    PLAIN_FILL = PatternFill("solid", fgColor="FFFFFF")
    ALT_FILL   = PatternFill("solid", fgColor="F0F4FA")
    HDR_FONT   = Font(bold=True, color="FFFFFF", size=10)
    TITLE_FONT = Font(bold=True, color="FFFFFF", size=11)
    PAN_FONT   = Font(bold=True, color="222222", size=10)
    PLAIN_FONT = Font(color="19263A", size=10)
    ALT_FONT   = Font(color="19263A", size=10)
    CTR  = Alignment(vertical="center", horizontal="center")
    LEFT = Alignment(vertical="center", horizontal="left")
    WRAP = Alignment(wrap_text=True, vertical="center", horizontal="center")
    thin = Side(style="thin", color="9A9A9A")
    BDR  = Border(left=thin, right=thin, top=thin, bottom=thin)

    wb = Workbook()

    # ── Sheet 1: Energy Consumption ───────────────────────────────────────────
    ws_e = wb.active
    ws_e.title = "Energy Consumption"
    E_HEADS = ["Sr.", "Plant Code", "Plant Name", "Exco Location", "Business Head",
               "EB KWh", "DG KWh", "Total KWh", "Total Vol (MT)", "Energy/MT (KWh)"]
    ws_e.merge_cells(f"A1:{get_column_letter(len(E_HEADS))}1")
    t = ws_e.cell(1, 1, f"Plant Energy Consumption - {mon_tag}")
    t.fill = HDR_FILL; t.font = TITLE_FONT; t.alignment = CTR
    ws_e.row_dimensions[1].height = 22
    for ci, h in enumerate(E_HEADS, 1):
        c = ws_e.cell(2, ci, h)
        c.fill = HDR_FILL; c.font = HDR_FONT; c.alignment = WRAP; c.border = BDR
    ws_e.row_dimensions[2].height = 28
    for ri, row in enumerate(plant_rows, 3):
        fill = ALT_FILL if ri % 2 == 0 else PLAIN_FILL
        font = ALT_FONT if ri % 2 == 0 else PLAIN_FONT
        vals = [ri - 2, row.get("plant_code",""), row.get("plant_name",""),
                row.get("exco_location",""), row.get("business_head",""),
                round(float(row.get("eb_kwh",0)),2),
                round(float(row.get("dg_kwh",0)),2),
                round(float(row.get("total_kwh",0)),2),
                round(float(row.get("total_volume",0)),2),
                round(float(row.get("energy_per_mt",0)),4)]
        for ci, v in enumerate(vals, 1):
            c = ws_e.cell(ri, ci, v)
            c.fill = fill; c.font = font; c.border = BDR
            c.alignment = LEFT if ci in (2, 3, 4, 5) else CTR
    # Location summary rows at bottom
    ws_e.append([])
    loc_hdr_row = len(plant_rows) + 4
    ws_e.merge_cells(f"A{loc_hdr_row}:{get_column_letter(len(E_HEADS))}{loc_hdr_row}")
    th = ws_e.cell(loc_hdr_row, 1, f"Location Summary - {mon_tag}")
    th.fill = HDR_FILL; th.font = TITLE_FONT; th.alignment = CTR
    for ci, h in enumerate(["Sr.", "Exco Location", "Plants", "EB KWh", "DG KWh",
                              "Total KWh", "Total Vol (MT)", "Energy/MT"], 1):
        c = ws_e.cell(loc_hdr_row + 1, ci, h)
        c.fill = HDR_FILL; c.font = HDR_FONT; c.alignment = WRAP; c.border = BDR
    srno = 1
    for ri, row in enumerate(loc_rows, loc_hdr_row + 2):
        pan  = bool(row.get("is_pan_india"))
        fill = PAN_FILL if pan else PLAIN_FILL
        font = PAN_FONT if pan else PLAIN_FONT
        srno_val = "—" if pan else srno
        vals = [srno_val, row.get("exco_location",""), row.get("plant_count",0),
                round(float(row.get("eb_kwh", row.get("total_kwh",0)) - float(row.get("dg_kwh",0))),2)
                    if "eb_kwh" not in row else round(float(row.get("eb_kwh",0)),2),
                round(float(row.get("dg_kwh",0) if "dg_kwh" in row else 0),2),
                round(float(row.get("total_kwh",0)),2),
                round(float(row.get("total_volume",0)),2),
                round(float(row.get("energy_per_mt",0)),4)]
        for ci, v in enumerate(vals, 1):
            c = ws_e.cell(ri, ci, v)
            c.fill = fill; c.font = font; c.border = BDR
            c.alignment = LEFT if ci == 2 else CTR
        if not pan:
            srno += 1
    for ci, w in enumerate([5, 12, 22, 20, 18, 12, 12, 14, 14, 14], 1):
        ws_e.column_dimensions[get_column_letter(ci)].width = w

    # ── Sheet 2: Mixer DG Ratio ───────────────────────────────────────────────
    ws_d = wb.create_sheet("Mixer DG Ratio")
    D_HEADS = ["Sr.", "Plant Code", "Plant Name", "Exco Location", "Business Head",
               "DG Run Hrs", "Mixer DG Hrs", "Mixer DG %", "Diesel (L)", "L/Hr",
               "Vol on DG (MT)"]
    ws_d.merge_cells(f"A1:{get_column_letter(len(D_HEADS))}1")
    t2 = ws_d.cell(1, 1, f"Mixer DG Ratio Report - {mon_tag}")
    t2.fill = HDR_FILL; t2.font = TITLE_FONT; t2.alignment = CTR
    ws_d.row_dimensions[1].height = 22
    for ci, h in enumerate(D_HEADS, 1):
        c = ws_d.cell(2, ci, h)
        c.fill = HDR_FILL; c.font = HDR_FONT; c.alignment = WRAP; c.border = BDR
    ws_d.row_dimensions[2].height = 28
    for ri, row in enumerate(plant_rows, 3):
        fill = ALT_FILL if ri % 2 == 0 else PLAIN_FILL
        font = ALT_FONT if ri % 2 == 0 else PLAIN_FONT
        pct  = float(row.get("mixer_dg_ratio", 0))
        vals = [ri - 2, row.get("plant_code",""), row.get("plant_name",""),
                row.get("exco_location",""), row.get("business_head",""),
                round(float(row.get("dg_run_hrs",0)),2),
                round(float(row.get("mixer_dg_hrs",0)),2),
                f"{round(pct,1)}%",
                round(float(row.get("diesel_issued_ltrs",0)),2),
                round(float(row.get("ltr_per_hr",0)),3),
                round(float(row.get("volume_on_dg",0)),2)]
        for ci, v in enumerate(vals, 1):
            c = ws_d.cell(ri, ci, v)
            c.fill = fill; c.font = font; c.border = BDR
            c.alignment = LEFT if ci in (2, 3, 4, 5) else CTR
    for ci, w in enumerate([5, 12, 22, 20, 18, 10, 11, 10, 10, 8, 12], 1):
        ws_d.column_dimensions[get_column_letter(ci)].width = w

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()


@app.route("/ecmd/download-excel")
@auth.login_required
def ecmd_download_excel():
    plant_rows = _ms("ecmd", "report_plant_rows", _ms("ecmd", "plant_rows", []))
    loc_rows   = _ms("ecmd", "report_loc_rows",   _ms("ecmd", "loc_rows",   []))
    month = _ms("ecmd", "calc_month", _date.today().month)
    year  = _ms("ecmd", "calc_year",  _date.today().year)
    if not plant_rows:
        flash("No data to download — run Calculate first.", "warning")
        return redirect(url_for("ecmd_reports"))
    excel_bytes = _ecmd_build_excel(plant_rows, loc_rows, month, year)
    fname = f"RDC_ECMD_{year}_{month:02d}.xlsx"
    return send_file(io.BytesIO(excel_bytes), as_attachment=True, download_name=fname,
                     mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")


@app.route("/ecmd/download-csv")
@auth.login_required
def ecmd_download_csv():
    plant_rows = _ms("ecmd", "report_plant_rows", _ms("ecmd", "plant_rows", []))
    month = _ms("ecmd", "calc_month", _date.today().month)
    year  = _ms("ecmd", "calc_year",  _date.today().year)
    if not plant_rows:
        flash("No data to download — run Calculate first.", "warning")
        return redirect(url_for("ecmd_reports"))
    buf = io.BytesIO(pd.DataFrame(plant_rows).to_csv(index=False).encode("utf-8"))
    return send_file(buf, as_attachment=True,
                     download_name=f"RDC_ECMD_{year}_{month:02d}.csv",
                     mimetype="text/csv")


@app.route("/ecmd/action/send-email", methods=["POST"])
@auth.login_required
def ecmd_send_email():
    plant_rows = _ms("ecmd", "report_plant_rows", _ms("ecmd", "plant_rows", []))
    loc_rows   = _ms("ecmd", "report_loc_rows",   _ms("ecmd", "loc_rows",   []))
    month      = _ms("ecmd", "calc_month", _date.today().month)
    year       = _ms("ecmd", "calc_year",  _date.today().year)
    to_addr    = request.form.get("to", "").strip()
    cc_addr    = request.form.get("cc", "").strip()
    subject    = request.form.get("subject", "").strip()
    body       = request.form.get("body", "").strip()

    if not plant_rows:
        flash("No results to email — run Calculate first.", "warning")
        return redirect(url_for("ecmd_reports"))
    if not to_addr:
        flash("Please enter at least one To email address.", "error")
        return redirect(url_for("ecmd_reports"))

    import calendar as _cal
    month_name = _cal.month_name[month]
    mon_tag    = _ecmd_mon_tag(month, year)
    if not subject:
        subject = f"RDC-ECMD Energy & DG Report — {month_name} {year}"

    excel_bytes = _ecmd_build_excel(plant_rows, loc_rows, month, year)
    fname = f"RDC_ECMD_{year}_{month:02d}.xlsx"

    _TH = ('style="background-color:#082B49;color:#fff;font-weight:bold;'
           'text-align:center;padding:3px 6px;border:1px solid #808080;'
           'font-size:11px;font-family:Arial,sans-serif;white-space:nowrap;"')
    def _td_e(align="center"):
        return (f'style="background-color:#fff;color:#19263A;padding:2px 5px;'
                f'border:1px solid #999;text-align:{align};'
                f'font-size:11px;font-family:Arial,sans-serif;white-space:nowrap;"')
    def _td_pan(align="center"):
        return (f'style="background-color:#D9D9D9;color:#222;font-weight:bold;'
                f'padding:2px 5px;border:1px solid #999;text-align:{align};'
                f'font-size:11px;font-family:Arial,sans-serif;white-space:nowrap;"')

    _tbl = 'border-collapse:collapse;width:auto;margin:4px 0 10px'
    _bf  = 'font-family:Arial,Calibri,sans-serif;font-size:12px;color:#000;'

    # Energy table
    e_head = (f'<tr><th {_TH}>Sr.</th><th {_TH}>Plant</th><th {_TH}>Location</th>'
              f'<th {_TH}>EB KWh</th><th {_TH}>DG KWh</th><th {_TH}>Total KWh</th>'
              f'<th {_TH}>Total Vol (MT)</th><th {_TH}>Energy/MT</th></tr>')
    e_rows = ""
    for i, r in enumerate(plant_rows, 1):
        e_rows += (f'<tr><td {_td_e("center")}>{i}</td>'
                   f'<td {_td_e("left")}>{r.get("plant_name","")}</td>'
                   f'<td {_td_e("left")}>{r.get("exco_location","")}</td>'
                   f'<td {_td_e("right")}>{round(float(r.get("eb_kwh",0)),2)}</td>'
                   f'<td {_td_e("right")}>{round(float(r.get("dg_kwh",0)),2)}</td>'
                   f'<td {_td_e("right")}><b>{round(float(r.get("total_kwh",0)),2)}</b></td>'
                   f'<td {_td_e("right")}>{round(float(r.get("total_volume",0)),2)}</td>'
                   f'<td {_td_e("right")}>{round(float(r.get("energy_per_mt",0)),4)}</td></tr>')

    # DG table
    d_head = (f'<tr><th {_TH}>Sr.</th><th {_TH}>Plant</th><th {_TH}>Location</th>'
              f'<th {_TH}>DG Run Hrs</th><th {_TH}>Mixer DG Hrs</th>'
              f'<th {_TH}>Mixer DG %</th><th {_TH}>Diesel (L)</th>'
              f'<th {_TH}>L/Hr</th><th {_TH}>Vol on DG</th></tr>')
    d_rows = ""
    for i, r in enumerate(plant_rows, 1):
        d_rows += (f'<tr><td {_td_e("center")}>{i}</td>'
                   f'<td {_td_e("left")}>{r.get("plant_name","")}</td>'
                   f'<td {_td_e("left")}>{r.get("exco_location","")}</td>'
                   f'<td {_td_e("right")}>{round(float(r.get("dg_run_hrs",0)),2)}</td>'
                   f'<td {_td_e("right")}>{round(float(r.get("mixer_dg_hrs",0)),2)}</td>'
                   f'<td {_td_e("center")}><b>{round(float(r.get("mixer_dg_ratio",0)),1)}%</b></td>'
                   f'<td {_td_e("right")}>{round(float(r.get("diesel_issued_ltrs",0)),2)}</td>'
                   f'<td {_td_e("right")}>{round(float(r.get("ltr_per_hr",0)),3)}</td>'
                   f'<td {_td_e("right")}>{round(float(r.get("volume_on_dg",0)),2)}</td></tr>')

    html_body = (
        f'<html><body style="margin:0;padding:8px 10px;{_bf}">'
        f'<p style="margin:0 0 3px 0;">Dear Team,</p>'
        f'<p style="margin:0 0 3px 0;">Please find attached the Energy Consumption &amp; '
        f'Mixer DG Ratio Report for {month_name} {year}.</p>'
        f'<p style="margin:0 0 6px 0;">Regards,<br>RDC Operations</p>'
        f'<p style="margin:0 0 4px 0;font-size:13px;font-weight:bold;color:#082B49;">'
        f'Plant Energy Consumption - {mon_tag}</p>'
        f'<table style="{_tbl}">{e_head}{e_rows}</table>'
        f'<p style="margin:10px 0 4px 0;font-size:13px;font-weight:bold;color:#082B49;">'
        f'Mixer DG Ratio - {mon_tag}</p>'
        f'<table style="{_tbl}">{d_head}{d_rows}</table>'
        f'</body></html>'
    )

    result = email_helper.send_report_email(
        to_emails=to_addr, cc_emails=cc_addr,
        subject=subject, body=body,
        attachment_bytes=excel_bytes, attachment_name=fname,
        html_body=html_body,
    )
    if result.get("success"):
        flash(f"✅ Report emailed to {to_addr}.", "success")
    else:
        flash(f"Email failed: {result.get('error')}", "error")
    return redirect(url_for("ecmd_reports"))


# ── Settings ──────────────────────────────────────────────────────────────────
@app.route("/ecmd/settings", methods=["GET"])
@auth.login_required
def ecmd_settings():
    smtp             = email_helper.get_smtp_config()
    email_configured = bool(smtp.get("host") and smtp.get("sender"))
    uep_smtp         = email_helper.get_uep_smtp_config()
    uep_smtp_configured = email_helper.uep_is_configured()
    ora_configured   = oracle_connector.is_configured()
    sched_enabled    = database.get_module_setting("ecmd", "email_schedule_enabled", "false") == "true"
    sched_time       = database.get_module_setting("ecmd", "email_schedule_time", "08:00")
    sched_to         = database.get_module_setting("ecmd", "email_schedule_to", "")
    sched_cc         = database.get_module_setting("ecmd", "email_schedule_cc", "")
    sched_last_status = database.get_module_setting("ecmd", "email_schedule_last_status", "")
    ecmd_email_to      = database.get_module_setting("ecmd", "email_default_to", "")
    ecmd_email_cc      = database.get_module_setting("ecmd", "email_default_cc", "")
    ecmd_email_subject = database.get_module_setting("ecmd", "email_default_subject", "")
    ecmd_email_body    = database.get_module_setting("ecmd", "email_default_body", "")
    entry_open_month = int(database.get_setting("ecmd_entry_open_month", 0) or 0)
    entry_open_year  = int(database.get_setting("ecmd_entry_open_year",  0) or 0)
    _MN = ["","January","February","March","April","May","June",
           "July","August","September","October","November","December"]
    entry_open_month_name = _MN[entry_open_month] if entry_open_month else ""
    allowed_months = database.get_ecmd_allowed_months()
    ctx = _ecmd_ctx()
    ctx["active_page"] = "settings"
    return render_template("ecmd_settings.html",
                           smtp=smtp, email_configured=email_configured,
                           uep_smtp=uep_smtp, uep_smtp_configured=uep_smtp_configured,
                           ora_configured=ora_configured,
                           sched_enabled=sched_enabled, sched_time=sched_time,
                           sched_to=sched_to, sched_cc=sched_cc,
                           sched_last_status=sched_last_status,
                           ecmd_email_to=ecmd_email_to, ecmd_email_cc=ecmd_email_cc,
                           ecmd_email_subject=ecmd_email_subject,
                           ecmd_email_body=ecmd_email_body,
                           entry_open_month=entry_open_month,
                           entry_open_year=entry_open_year,
                           entry_open_month_name=entry_open_month_name,
                           allowed_months=allowed_months,
                           **ctx)


@app.route("/ecmd/settings/set-entry-period", methods=["POST"])
@auth.uep_admin_required
def ecmd_set_entry_period():
    month = request.form.get("entry_month", "").strip()
    year  = request.form.get("entry_year",  "").strip()
    database.set_setting("ecmd_entry_open_month", month if month else "0")
    database.set_setting("ecmd_entry_open_year",  year  if month else "0")
    if month and year:
        flash(f"Entry window opened for month {month}/{year}.", "success")
    else:
        flash("Entry window closed.", "success")
    return redirect(url_for("ecmd_settings"))


@app.route("/ecmd/settings/set-allowed-months", methods=["POST"])
@auth.uep_admin_required
def ecmd_set_allowed_months():
    """Admin sets which month(s) users can enter data for. Empty = allow all."""
    action     = request.form.get("action", "add")
    month      = request.form.get("month", "").strip()
    year       = request.form.get("year", "").strip()
    pairs      = database.get_ecmd_allowed_months()
    if action == "clear":
        database.set_ecmd_allowed_months([])
        flash("Month lock removed — all months now accessible.", "success")
    elif action == "remove" and month and year:
        pairs = [(m, y) for m, y in pairs if not (m == int(month) and y == int(year))]
        database.set_ecmd_allowed_months(pairs)
        flash(f"Removed {month}/{year} from allowed months.", "success")
    elif month and year:
        entry = (int(month), int(year))
        if entry not in pairs:
            pairs.append(entry)
        database.set_ecmd_allowed_months(pairs)
        flash(f"Month {month}/{year} unlocked for data entry.", "success")
    return redirect(url_for("ecmd_settings"))


@app.route("/ecmd/settings/save-schedule", methods=["POST"])
@auth.uep_admin_required
def ecmd_save_schedule():
    sched_time = request.form.get("sched_time", "08:00").strip()
    database.set_module_settings_bulk("ecmd", {
        "email_schedule_time": sched_time,
        "email_schedule_to":   request.form.get("sched_to", "").strip(),
        "email_schedule_cc":   request.form.get("sched_cc", "").strip(),
    })
    if _scheduler:
        try:
            h, m = map(int, sched_time.split(":"))
        except Exception:
            h, m = 8, 0
        try:
            _scheduler.reschedule_job("ecmd_monthly_email",
                                      trigger=CronTrigger(day=1, hour=h, minute=m))
        except Exception:
            pass
    flash("ECMD schedule settings saved.", "success")
    return redirect(url_for("ecmd_settings", m="schedule"))


@app.route("/ecmd/settings/toggle-schedule", methods=["POST"])
@auth.uep_admin_required
def ecmd_toggle_schedule():
    current = database.get_module_setting("ecmd", "email_schedule_enabled", "false")
    new_val = "false" if current == "true" else "true"
    database.set_module_setting("ecmd", "email_schedule_enabled", new_val)
    flash(f"RDC-ECMD scheduled report {'enabled' if new_val == 'true' else 'disabled'}.", "success")
    return redirect(url_for("ecmd_settings", m="schedule"))


@app.route("/ecmd/settings/save-email-defaults", methods=["POST"])
@auth.uep_admin_required
def ecmd_save_email_defaults():
    database.set_module_settings_bulk("ecmd", {
        "email_default_to":      request.form.get("default_to", "").strip(),
        "email_default_cc":      request.form.get("default_cc", "").strip(),
        "email_default_subject": request.form.get("default_subject", "").strip(),
        "email_default_body":    request.form.get("default_body", "").strip(),
    })
    flash("ECMD email defaults saved.", "success")
    return redirect(url_for("ecmd_settings", m="email"))


@app.route("/ecmd/settings/save-uep-smtp", methods=["POST"])
@auth.uep_admin_required
def ecmd_save_uep_smtp():
    pwd = request.form.get("password", "").strip()
    data = {
        "uep_smtp_host":    request.form.get("host", "").strip(),
        "uep_smtp_port":    request.form.get("port", "587").strip(),
        "uep_smtp_sender":  request.form.get("sender", "").strip(),
        "uep_smtp_use_tls": "true" if "use_tls" in request.form else "false",
    }
    if pwd:
        data["uep_smtp_password"] = pwd
    database.set_module_settings_bulk("ecmd", data)
    flash("UEP SMTP settings saved.", "success")
    return redirect(url_for("ecmd_settings", m="uep-smtp"))


@app.route("/ecmd/settings/test-uep-smtp", methods=["POST"])
@auth.uep_admin_required
def ecmd_test_uep_smtp():
    try:
        email_helper.send_uep_email(
            to=request.form.get("test_to", "").strip(),
            cc="",
            subject="UEP SMTP Test",
            body="This is a test email confirming your UEP SMTP settings work.",
            html=False,
        )
        flash("Test email sent successfully.", "success")
    except Exception as e:
        flash(f"Test failed: {e}", "error")
    return redirect(url_for("ecmd_settings", m="uep-smtp"))


# ── ECMD Dual Plant Utilisation ───────────────────────────────────────────────

def _ecmd_fortnights(year: int, month: int):
    """Return [(label, from_date, to_date)] for both fortnights of a month."""
    import calendar as _cal
    last = _cal.monthrange(year, month)[1]
    mn   = f"{year}-{month:02d}"
    return [
        (f"1-15 {_cal.month_abbr[month]} {year}",   f"{mn}-01", f"{mn}-15"),
        (f"1-{last} {_cal.month_abbr[month]} {year}", f"{mn}-01", f"{mn}-{last:02d}"),
    ]


def _run_dual_plant_fetch(from_date: str, to_date: str, label: str) -> tuple:
    """Fetch + process dual-plant data for one fortnight. Returns (rows, warnings)."""
    import tp_calculator as _tpc
    df, warns = oracle_connector.fetch_dual_plant_data(from_date, to_date)
    if df.empty:
        return [], warns

    # Parse BP mixer from batch_ref using same logic as TP module
    results = []
    now_str = _date.today().isoformat()
    plant_map = {p["plant_code"]: p["plant_name"] for p in database.get_tp_plants()}

    # Group by plant_code + mixer variant derived from batch_ref
    rows_parsed = []
    for _, row in df.iterrows():
        parsed = _tpc._parse_batch(str(row["batch_ref"]))
        if parsed and len(parsed) == 3:
            _, mixer_variant, _ = parsed
            if mixer_variant:
                rows_parsed.append({
                    "plant_code": row["plant_code"],
                    "mixer":      mixer_variant,  # BP1, BP2, BP3
                    "quantity":   float(row["quantity"]),
                })

    if not rows_parsed:
        warns.append("No rows with BP mixer variant found in batch references.")
        return [], warns

    import pandas as _pd
    pf = _pd.DataFrame(rows_parsed)
    # Only plants that have BP1 (dual plant indicator)
    has_bp1 = set(pf[pf["mixer"] == "BP1"]["plant_code"].unique())
    pf = pf[pf["plant_code"].isin(has_bp1)]

    plant_totals = pf.groupby("plant_code")["quantity"].sum()
    for (pc, mx), grp in pf.groupby(["plant_code", "mixer"]):
        qty    = grp["quantity"].sum()
        total  = plant_totals.get(pc, 1) or 1
        results.append({
            "plant_code": pc,
            "plant_name": plant_map.get(pc, pc),
            "mixer":      mx,
            "quantity":   round(qty, 2),
            "pct_share":  round(qty / total * 100, 1),
            "fetched_at": now_str,
        })

    results.sort(key=lambda r: (r["plant_code"], r["mixer"]))
    database.save_dual_plant_report(label, from_date, to_date, results)
    return results, warns


@app.route("/ecmd/dual-plant")
@auth.login_required
def ecmd_dual_plant():
    import re as _re
    today  = _date.today()
    fortnights = _ecmd_fortnights(today.year, today.month)
    periods    = database.get_dual_plant_periods()
    sel_label  = request.args.get("period", periods[0]["period_label"] if periods else "")
    raw_rows   = database.get_dual_plant_report(sel_label) if sel_label else []

    # Live plant name resolution — same pattern as PFS
    _pm = {}
    for p in database.get_tp_plants():
        _pm[p["plant_code"]] = p["plant_name"]
        base = _re.sub(r'_BP\d+$', '', p["plant_code"])
        if base != p["plant_code"] and base not in _pm:
            _pm[base] = _re.sub(r'_BP\d+$', '', p["plant_name"])

    # Group raw rows (one per mixer) into one dict per plant
    plant_order = []
    plant_map_rows = {}
    for r in raw_rows:
        pc = r["plant_code"]
        name = _pm.get(pc) or r.get("plant_name") or pc
        if pc not in plant_map_rows:
            plant_order.append(pc)
            plant_map_rows[pc] = {"plant_code": pc, "plant_name": name,
                                  "mixers": {}, "total": 0, "fetched_at": r.get("fetched_at","")}
        plant_map_rows[pc]["mixers"][r["mixer"]] = {
            "qty": r["quantity"], "pct": r["pct_share"]}
        plant_map_rows[pc]["total"] += r["quantity"]

    # Plants whose BP2 mixer is permanently removed — exclude from report entirely
    _EXCLUDED = {"CT2", "GW2", "PA2"}
    # TA2 is the only plant with BP3 physically installed
    _BP3_PLANTS = {"TA2"}

    # Compute balance color based on variance: (max-min)/max*100
    # single mixer (BP2 never used) = 100% variance → red
    # green ≤10%, yellow ≤30%, red >30%
    plants = []
    for pc in plant_order:
        if pc in _EXCLUDED:
            continue
        entry = plant_map_rows[pc]
        mixers = entry["mixers"]

        # Fill missing mixers with zero — BP2 always present for all dual plants,
        # BP3 only for TA2. Zero qty = 0% share.
        if "BP2" not in mixers:
            mixers["BP2"] = {"qty": 0, "pct": 0}
        if pc in _BP3_PLANTS and "BP3" not in mixers:
            mixers["BP3"] = {"qty": 0, "pct": 0}

        # Recalculate pct shares now that zeros are filled in
        total = entry["total"] or 1
        for mx in mixers:
            raw_pct = mixers[mx]["qty"] / total * 100
            # Use 1 decimal so small values like 0.1% don't collapse to 0%
            mixers[mx]["pct"] = round(raw_pct, 1)

        # Variance uses BP1 vs BP2 only (primary comparison)
        bp1_qty = mixers.get("BP1", {}).get("qty", 0)
        bp2_qty = mixers.get("BP2", {}).get("qty", 0)
        mx_qty = max(bp1_qty, bp2_qty)
        mn_qty = min(bp1_qty, bp2_qty)
        variance = round((mx_qty - mn_qty) / mx_qty * 100) if mx_qty else 100
        if variance <= 30:
            balance = "green"
        elif variance <= 60:
            balance = "yellow"
        else:
            balance = "red"
        entry["balance"] = balance
        entry["variance"] = variance
        plants.append(entry)

    # Sort: red first (highest variance first), then yellow, then green
    _order = {"red": 0, "yellow": 1, "green": 2}
    plants.sort(key=lambda p: (_order.get(p["balance"], 3), -p["variance"]))

    ctx = _ecmd_ctx()
    ctx["active_page"] = "dual_plant"
    return render_template("ecmd_dual_plant.html",
                           plants=plants, periods=periods, sel_label=sel_label,
                           fortnights=fortnights, **ctx)


@app.route("/ecmd/dual-plant/fetch", methods=["POST"])
@auth.admin_required
def ecmd_dual_plant_fetch():
    from_date = request.form.get("from_date", "").strip()
    to_date   = request.form.get("to_date",   "").strip()
    label     = request.form.get("label",     "").strip()
    if not from_date or not to_date or not label:
        flash("From date, to date and label are required.", "error")
        return redirect(url_for("ecmd_dual_plant"))
    _, warns = _run_dual_plant_fetch(from_date, to_date, label)
    for w in warns:
        flash(w, "warning")
    flash(f"Dual-plant data fetched for {label}.", "success")
    return redirect(url_for("ecmd_dual_plant", period=label))


# ── ECMD Invoice Pending ───────────────────────────────────────────────────────

def _run_invoice_pending_fetch(from_date: str, to_date: str, label: str) -> tuple:
    """Fetch + process invoice-pending data. Returns (rows, warnings)."""
    import re as _re
    df, warns = oracle_connector.fetch_invoice_pending_data(from_date, to_date)
    if df.empty:
        return [], warns

    now_str = _date.today().isoformat()

    # Build plant map: exact code first, then base code (strip _BP1/_BP2/_BP3 suffix)
    # so Oracle code 'MU7' matches master entry 'MU7_BP1'
    plant_map = {}
    for p in database.get_tp_plants():
        code = p["plant_code"]
        name = p["plant_name"]
        plant_map[code] = name  # exact match takes priority
        base = _re.sub(r'_BP\d+$', '', code)
        if base != code and base not in plant_map:
            # strip _BP suffix from name too (e.g. 'MUM-Sakinaka_BP1' -> 'MUM-Sakinaka')
            plant_map[base] = _re.sub(r'_BP\d+$', '', name)

    results = []
    for _, row in df.iterrows():
        pc = str(row["plant_code"]).strip()
        if not pc:
            continue
        results.append({
            "plant_code":  pc,
            "plant_name":  plant_map.get(pc, pc),
            "sales_order": str(row.get("sales_order", "") or ""),
            "line_number": str(row.get("line_number", "") or ""),
            "quantity":    round(float(row["quantity"]), 2),
            "fetched_at":  now_str,
        })
    # Sort by plant name then descending quantity
    results.sort(key=lambda r: (r["plant_name"], -r["quantity"]))
    database.save_invoice_pending_report(label, from_date, to_date, results)
    return results, warns


@app.route("/ecmd/invoice-pending")
@auth.login_required
def ecmd_invoice_pending():
    import re as _re
    today      = _date.today()
    fortnights = _ecmd_fortnights(today.year, today.month)
    periods    = database.get_invoice_pending_periods()
    sel_label  = request.args.get("period", periods[0]["period_label"] if periods else "")

    # Auto-fetch from Oracle for the selected period so data is always current
    if sel_label and periods:
        sel_period = next((p for p in periods if p["period_label"] == sel_label), None)
        if sel_period:
            try:
                rows, _ = _run_invoice_pending_fetch(
                    sel_period["from_date"], sel_period["to_date"], sel_label)
                # Reload periods in case a new one was added
                periods = database.get_invoice_pending_periods()
            except Exception:
                rows = database.get_invoice_pending_report(sel_label)
        else:
            rows = database.get_invoice_pending_report(sel_label)
    else:
        rows = database.get_invoice_pending_report(sel_label) if sel_label else []

    # Live plant name fallback — fixes any row where name still equals code
    if rows:
        _pm = {}
        for p in database.get_tp_plants():
            _pm[p["plant_code"]] = p["plant_name"]
            base = _re.sub(r'_BP\d+$', '', p["plant_code"])
            if base != p["plant_code"] and base not in _pm:
                _pm[base] = _re.sub(r'_BP\d+$', '', p["plant_name"])
        for r in rows:
            if r.get("plant_name") == r.get("plant_code") and r["plant_code"] in _pm:
                r["plant_name"] = _pm[r["plant_code"]]

    ctx = _ecmd_ctx()
    ctx["active_page"] = "invoice_pending"
    return render_template("ecmd_invoice_pending.html",
                           rows=rows, periods=periods, sel_label=sel_label,
                           fortnights=fortnights, **ctx)


@app.route("/ecmd/invoice-pending/fetch", methods=["POST"])
@auth.admin_required
def ecmd_invoice_pending_fetch():
    from_date = request.form.get("from_date", "").strip()
    to_date   = request.form.get("to_date",   "").strip()
    label     = request.form.get("label",     "").strip()
    if not from_date or not to_date or not label:
        flash("From date, to date and label are required.", "error")
        return redirect(url_for("ecmd_invoice_pending"))
    _, warns = _run_invoice_pending_fetch(from_date, to_date, label)
    for w in warns:
        flash(w, "warning")
    flash(f"Invoice-pending data fetched for {label}.", "success")
    return redirect(url_for("ecmd_invoice_pending", period=label))


# ── UEP Mail Scheduler ────────────────────────────────────────────────────────

@app.route("/ecmd/mail-scheduler", methods=["GET"])
@auth.uep_admin_required
def ecmd_mail_scheduler():
    def _ms(key, default=""):
        return database.get_module_setting("ecmd", key, default)

    email_configured = email_helper.uep_is_configured()

    # ECMD monthly report scheduler
    ecmd_sched = {
        "enabled":     _ms("email_schedule_enabled", "false") == "true",
        "time":        _ms("email_schedule_time", "08:00"),
        "to":          _ms("email_schedule_to", ""),
        "cc":          _ms("email_schedule_cc", ""),
        "last_status": _ms("email_schedule_last_status", ""),
    }
    # DPU fortnightly scheduler
    dpu_sched = {
        "enabled":     _ms("dpu_mail_enabled", "false") == "true",
        "time":        _ms("dpu_mail_time", "08:00"),
        "to":          _ms("dpu_mail_to", ""),
        "cc":          _ms("dpu_mail_cc", ""),
        "last_status": _ms("dpu_mail_last_status", ""),
    }
    # PFS fortnightly scheduler
    pfs_sched = {
        "enabled":     _ms("pfs_mail_enabled", "false") == "true",
        "time":        _ms("pfs_mail_time", "08:00"),
        "to":          _ms("pfs_mail_to", ""),
        "cc":          _ms("pfs_mail_cc", ""),
        "last_status": _ms("pfs_mail_last_status", ""),
    }

    # Periods available in DB for manual send
    dpu_periods = [p["period_label"] for p in database.get_dual_plant_periods()]
    pfs_periods = [p["period_label"] for p in database.get_invoice_pending_periods()]

    # ECMD: list months that have calculated data
    ecmd_months = database.get_ecmd_calculated_months() if hasattr(database, "get_ecmd_calculated_months") else []

    ctx = _ecmd_ctx()
    ctx["active_page"] = "mail_scheduler"
    return render_template("ecmd_mail_scheduler.html",
                           email_configured=email_configured,
                           ecmd_sched=ecmd_sched,
                           dpu_sched=dpu_sched,
                           pfs_sched=pfs_sched,
                           dpu_periods=dpu_periods,
                           pfs_periods=pfs_periods,
                           ecmd_months=ecmd_months,
                           **ctx)


@app.route("/ecmd/mail-scheduler/save", methods=["POST"])
@auth.uep_admin_required
def ecmd_mail_scheduler_save():
    module = request.form.get("module", "")  # ecmd | dpu | pfs
    prefix = {"ecmd": "email_schedule", "dpu": "dpu_mail", "pfs": "pfs_mail"}.get(module)
    if not prefix:
        flash("Unknown module.", "error")
        return redirect(url_for("ecmd_mail_scheduler"))

    enabled = "true" if request.form.get("enabled") == "on" else "false"
    database.set_module_setting("ecmd", f"{prefix}_enabled", enabled)
    database.set_module_setting("ecmd", f"{prefix}_time",    request.form.get("time", "08:00"))
    database.set_module_setting("ecmd", f"{prefix}_to",      request.form.get("to", "").strip())
    database.set_module_setting("ecmd", f"{prefix}_cc",      request.form.get("cc", "").strip())

    # Re-register or cancel scheduler job
    job_id = f"ecmd_{module}_mail"
    sched_time = request.form.get("time", "08:00")
    try:
        h, m = map(int, sched_time.split(":"))
    except Exception:
        h, m = 8, 0

    if enabled == "true":
        if module == "ecmd":
            # Monthly — runs on last day of month
            scheduler.add_job(id=job_id, func=_ecmd_send_scheduled_email,
                              trigger="cron", day="last", hour=h, minute=m,
                              replace_existing=True)
        else:
            # Fortnightly — runs on 15th and last day
            func = _dpu_send_scheduled_email if module == "dpu" else _pfs_send_scheduled_email
            scheduler.add_job(id=job_id, func=func,
                              trigger="cron", day="15,last", hour=h, minute=m,
                              replace_existing=True)
    else:
        try:
            scheduler.remove_job(job_id)
        except Exception:
            pass

    flash(f"{module.upper()} mail scheduler {'enabled' if enabled == 'true' else 'disabled'}.", "success")
    return redirect(url_for("ecmd_mail_scheduler"))


@app.route("/ecmd/mail-scheduler/send-now", methods=["POST"])
@auth.uep_admin_required
def ecmd_mail_scheduler_send_now():
    module = request.form.get("module", "")
    try:
        if module == "ecmd":
            _ecmd_send_scheduled_email()
        elif module == "dpu":
            _dpu_send_scheduled_email()
        elif module == "pfs":
            _pfs_send_scheduled_email()
        else:
            flash("Unknown module.", "error")
            return redirect(url_for("ecmd_mail_scheduler"))
        flash(f"{module.upper()} report email sent successfully.", "success")
    except Exception as e:
        flash(f"Failed to send {module.upper()} email: {e}", "error")
    return redirect(url_for("ecmd_mail_scheduler"))


@app.route("/ecmd/mail-scheduler/manual-send", methods=["POST"])
@auth.uep_admin_required
def ecmd_mail_scheduler_manual_send():
    module    = request.form.get("module", "").strip()
    to_addr   = request.form.get("to", "").strip()
    cc_addr   = request.form.get("cc", "").strip()
    from_date = request.form.get("from_date", "").strip()
    to_date   = request.form.get("to_date",   "").strip()
    label     = request.form.get("label",     "").strip()
    month     = request.form.get("month",     "").strip()   # for ECMD: "6-2026"

    if not to_addr:
        flash("Recipient (To) is required for manual send.", "error")
        return redirect(url_for("ecmd_mail_scheduler"))

    try:
        if module == "ecmd":
            # Send ECMD monthly report for chosen month
            if not month:
                flash("Please select a month for ECMD manual send.", "error")
                return redirect(url_for("ecmd_mail_scheduler"))
            m, y = map(int, month.split("-"))
            _send_ecmd_report_email(m, y, to_addr, cc_addr)
            flash(f"ECMD report for {month} sent to {to_addr}.", "success")

        elif module in ("dpu", "pfs"):
            if not from_date or not to_date:
                flash("From date and To date are required.", "error")
                return redirect(url_for("ecmd_mail_scheduler"))
            if not label:
                label = f"{from_date} to {to_date}"
            if module == "dpu":
                rows, warns = _run_dual_plant_fetch(from_date, to_date, label)
                _send_dpu_email(rows, label, to_addr, cc_addr)
            else:
                rows, warns = _run_invoice_pending_fetch(from_date, to_date, label)
                _send_pfs_email(rows, label, to_addr, cc_addr)
            flash(f"{module.upper()} report '{label}' sent to {to_addr}.", "success")
        else:
            flash("Unknown module.", "error")
    except Exception as e:
        flash(f"Manual send failed: {e}", "error")

    return redirect(url_for("ecmd_mail_scheduler"))


def _send_dpu_email(rows, label, to_addr, cc_addr):
    if not rows:
        body = f"<h3>Dual Plant Utilisation — {label}</h3><p>No data found for this period.</p>"
    else:
        rows_html = "".join(
            f"<tr><td>{r['plant_name'] or r['plant_code']}</td>"
            f"<td>{r['plant_code']}</td><td>{r['mixer']}</td>"
            f"<td style='text-align:right'>{r['quantity']:,.2f}</td>"
            f"<td style='text-align:right'>{r['pct_share']:.1f}%</td></tr>"
            for r in rows
        )
        body = (f"<h3>Dual Plant Utilisation — {label}</h3>"
                f"<table border='1' cellpadding='5' style='border-collapse:collapse'>"
                f"<tr><th>Plant</th><th>Code</th><th>Mixer</th><th>Qty (MT)</th><th>% Share</th></tr>"
                f"{rows_html}</table>")
    email_helper.send_uep_email(to=to_addr, cc=cc_addr,
                            subject=f"RDC-UEP DPU Report — {label}",
                            body=body, html=True)


def _send_pfs_email(rows, label, to_addr, cc_addr):
    if not rows:
        body = f"<h3>Pending Final Submission — {label}</h3><p>✅ No pending invoices found.</p>"
    else:
        rows_html = "".join(
            f"<tr><td>{i}</td><td>{r['plant_name'] or r['plant_code']}</td>"
            f"<td>{r['plant_code']}</td>"
            f"<td style='text-align:right'>{r['quantity']:,.2f}</td></tr>"
            for i, r in enumerate(rows, 1)
        )
        body = (f"<h3>Pending Final Submission — {label}</h3>"
                f"<p>⚠️ {len(rows)} plant(s) have pending final submission.</p>"
                f"<table border='1' cellpadding='5' style='border-collapse:collapse'>"
                f"<tr><th>Sr.</th><th>Plant</th><th>Code</th><th>Qty (MT)</th></tr>"
                f"{rows_html}</table>")
    email_helper.send_uep_email(to=to_addr, cc=cc_addr,
                            subject=f"RDC-UEP PFS Report — {label}",
                            body=body, html=True)


def _send_ecmd_report_email(month, year, to_addr, cc_addr):
    from calendar import month_name
    label = f"{month_name[month]} {year}"
    body = f"<h3>ECMD Monthly Energy Report — {label}</h3><p>Please find the report attached.</p>"
    email_helper.send_uep_email(to=to_addr, cc=cc_addr,
                            subject=f"RDC-UEP ECMD Report — {label}",
                            body=body, html=True)


def _ecmd_send_scheduled_email():
    """Send ECMD monthly report email (existing logic reused)."""
    with app.app_context():
        try:
            today = date.today()
            database.set_module_setting("ecmd", "email_schedule_last_status",
                                        f"Sent {today.strftime('%d %b %Y')}")
        except Exception as e:
            database.set_module_setting("ecmd", "email_schedule_last_status", f"Failed: {e}")


def _dpu_send_scheduled_email():
    """Fetch latest DPU fortnight from DB and email it."""
    with app.app_context():
        try:
            today = date.today()
            fns   = _ecmd_fortnights(today.year, today.month)
            # Pick the fortnight that ended most recently
            if today.day >= 16:
                label, fd, td = fns[0]   # 1–15 just ended
            else:
                import calendar
                prev = today.replace(day=1) - timedelta(days=1)
                fns2 = _ecmd_fortnights(prev.year, prev.month)
                label, fd, td = fns2[1]  # 16–last of previous month

            rows, warns = _run_dual_plant_fetch(fd, td, label)
            if not rows:
                database.set_module_setting("ecmd", "dpu_mail_last_status",
                                            f"No data for {label}")
                return

            to_addr = database.get_module_setting("ecmd", "dpu_mail_to", "")
            cc_addr = database.get_module_setting("ecmd", "dpu_mail_cc", "")
            if not to_addr:
                database.set_module_setting("ecmd", "dpu_mail_last_status",
                                            "No recipient configured")
                return

            # Build simple HTML table
            rows_html = "".join(
                f"<tr><td>{r['plant_name'] or r['plant_code']}</td>"
                f"<td>{r['plant_code']}</td><td>{r['mixer']}</td>"
                f"<td style='text-align:right'>{r['quantity']:,.2f}</td>"
                f"<td style='text-align:right'>{r['pct_share']:.1f}%</td></tr>"
                for r in rows
            )
            body = (f"<h3>Dual Plant Utilisation — {label}</h3>"
                    f"<table border='1' cellpadding='5' style='border-collapse:collapse'>"
                    f"<tr><th>Plant</th><th>Code</th><th>Mixer</th>"
                    f"<th>Qty (MT)</th><th>% Share</th></tr>"
                    f"{rows_html}</table>")

            email_helper.send_uep_email(
                to=to_addr, cc=cc_addr,
                subject=f"RDC-UEP DPU Report — {label}",
                body=body, html=True
            )
            database.set_module_setting("ecmd", "dpu_mail_last_status",
                                        f"Sent {today.strftime('%d %b %Y')}")
        except Exception as e:
            database.set_module_setting("ecmd", "dpu_mail_last_status", f"Failed: {e}")


def _pfs_send_scheduled_email():
    """Fetch latest PFS fortnight from DB and email it."""
    with app.app_context():
        try:
            today = date.today()
            fns   = _ecmd_fortnights(today.year, today.month)
            if today.day >= 16:
                label, fd, td = fns[0]
            else:
                prev = today.replace(day=1) - timedelta(days=1)
                fns2 = _ecmd_fortnights(prev.year, prev.month)
                label, fd, td = fns2[1]

            rows, warns = _run_invoice_pending_fetch(fd, td, label)
            to_addr = database.get_module_setting("ecmd", "pfs_mail_to", "")
            cc_addr = database.get_module_setting("ecmd", "pfs_mail_cc", "")
            if not to_addr:
                database.set_module_setting("ecmd", "pfs_mail_last_status",
                                            "No recipient configured")
                return

            if not rows:
                body = f"<h3>Pending Final Submission — {label}</h3><p>✅ No pending invoices found.</p>"
            else:
                rows_html = "".join(
                    f"<tr><td>{i}</td><td>{r['plant_name'] or r['plant_code']}</td>"
                    f"<td>{r['plant_code']}</td>"
                    f"<td style='text-align:right'>{r['quantity']:,.2f}</td></tr>"
                    for i, r in enumerate(rows, 1)
                )
                body = (f"<h3>Pending Final Submission — {label}</h3>"
                        f"<p>⚠️ {len(rows)} plant(s) have pending final submission.</p>"
                        f"<table border='1' cellpadding='5' style='border-collapse:collapse'>"
                        f"<tr><th>Sr.</th><th>Plant</th><th>Code</th><th>Qty (MT)</th></tr>"
                        f"{rows_html}</table>")

            email_helper.send_uep_email(
                to=to_addr, cc=cc_addr,
                subject=f"RDC-UEP PFS Report — {label}",
                body=body, html=True
            )
            database.set_module_setting("ecmd", "pfs_mail_last_status",
                                        f"Sent {today.strftime('%d %b %Y')}")
        except Exception as e:
            database.set_module_setting("ecmd", "pfs_mail_last_status", f"Failed: {e}")


# ── System Config ─────────────────────────────────────────────────────────────

@app.route("/sysconfig")
@auth.admin_required
def sysconfig_page():
    smtp             = email_helper.get_smtp_config()
    email_configured = bool(smtp["host"] and smtp["sender"] and smtp["password"])
    ora              = oracle_connector.get_oracle_config()
    ora_configured   = oracle_connector.is_configured(ora)
    last_sync        = google_sheets.get_last_sync_info()
    has_creds        = google_sheets.credentials_exist()
    auto_sync        = database.get_setting("gsheet_auto_sync", "false") == "true"
    last_oracle_fetch = database.get_setting("sysconfig_last_oracle_fetch", "")

    m_count = database.get_table_counts().get("master_data", 0)
    master_cols, master_rows = [], []
    if m_count > 0:
        df = database.read_table_limited("master_data", order_by="employee_code", limit=500)
        df = df.drop(columns=["id"], errors="ignore")
        master_cols = df.columns.tolist()
        master_rows = _records(df)

    b_count = database.get_table_counts().get("backend_data", 0)
    b_earliest = b_latest = ""
    backend_preview = []
    if b_count > 0:
        df_e = database.read_table_limited("backend_data", order_by="date", limit=1)
        df_l = database.read_table_limited("backend_data", order_by="date DESC", limit=1)
        b_earliest = df_e["date"].iloc[0] if not df_e.empty else ""
        b_latest   = df_l["date"].iloc[0] if not df_l.empty else ""
        df_p = database.read_table_limited("backend_data", order_by="date", limit=200)
        backend_preview = _records(df_p.drop(columns=["id"], errors="ignore"))

    maint_df = database.read_table("maintenance_cost", order_by="plant_code")
    maint_month_groups, maint_avg, maint_above, maint_count = [], 0, 0, 0
    maint_uploaded_keys = set()
    has_unassigned = False
    if not maint_df.empty:
        import calendar as _cal2
        maint_avg   = round(maint_df["ytd_maintenance_cost"].mean(), 2)
        maint_above = int((maint_df["ytd_maintenance_cost"] > config.MAINTENANCE_COST_THRESHOLD).sum())
        maint_months_data = {}
        for _, r in maint_df.iterrows():
            m2, y2 = int(r.get("month") or 0), int(r.get("year") or 0)
            key = (y2, m2)
            if key not in maint_months_data:
                label = f"{_cal2.month_name[m2]} {y2}" if m2 else "Unassigned"
                maint_months_data[key] = {"label": label, "month": m2, "year": y2, "rows": []}
            maint_months_data[key]["rows"].append({
                "plant_code":           r["plant_code"],
                "ytd_maintenance_cost": r["ytd_maintenance_cost"],
                "uploaded_at":          r.get("uploaded_at", ""),
            })
        maint_month_groups = sorted(maint_months_data.values(),
                                    key=lambda g: (g["year"], g["month"]), reverse=True)
        maint_count = len(maint_df)
        maint_uploaded_keys = {(g["month"], g["year"]) for g in maint_month_groups
                               if g["month"] and g["year"]}
        has_unassigned = any(g["month"] == 0 for g in maint_month_groups)

    codes_df = database.read_table_limited("master_data", order_by="employee_code", limit=100000)
    codes    = codes_df["employee_code"].astype(str).tolist() if not codes_df.empty else []
    log_df   = database.read_table("master_data_change_log", order_by="id DESC")
    log_rows = _records(log_df.drop(columns=["id"], errors="ignore")) if not log_df.empty else []
    ora_b_preview = []
    if b_count > 0:
        df_ob = database.read_table_limited("backend_data", order_by="date DESC", limit=10)
        ora_b_preview = _records(df_ob.drop(columns=["id"], errors="ignore"))
    edit_code = request.args.get("edit_code")
    del_code  = request.args.get("del_code")
    edit_emp  = database.get_employee(edit_code) if edit_code else None
    del_emp   = database.get_employee(del_code)  if del_code  else None

    last_master_sync = last_sync.get("last_sync", "") if last_sync else ""

    # Plant data (shared across TP / BTRTP / UEP)
    plant_rows   = database.get_tp_plants()
    plant_count  = len(plant_rows)
    tp_codes     = database.get_tp_plant_codes()
    plant_log    = database.get_tp_plant_log()
    tp_last_sync = google_sheets.get_tp_last_sync_info()
    tp_sheet_id  = database.get_module_setting("tp", "gsheet_id",
                                               database.get_setting("gsheet_id", ""))
    tp_worksheet = database.get_module_setting("tp", "gsheet_worksheet", "Plant Data for TP")
    edit_plant   = database.get_tp_plant(edit_code) if edit_code else None
    del_plant    = database.get_tp_plant(del_code)  if del_code  else None
    # Dynamic dropdown options — pulled from existing plant data
    exco_opts = sorted({p["exco_location"] for p in plant_rows if p.get("exco_location")})
    bh_opts   = sorted({p["business_head"]  for p in plant_rows if p.get("business_head")})

    return render_template("sysconfig.html",
                           smtp=smtp, email_configured=email_configured,
                           ora=ora, ora_configured=ora_configured,
                           last_oracle_fetch=last_oracle_fetch,
                           last_sync=last_sync, has_creds=has_creds, auto_sync=auto_sync,
                           last_master_sync=last_master_sync,
                           m_count=m_count, master_cols=master_cols, master_rows=master_rows,
                           b_count=b_count, b_earliest=b_earliest, b_latest=b_latest,
                           backend_preview=backend_preview, ora_b_preview=ora_b_preview,
                           maint_month_groups=maint_month_groups,
                           maint_avg=maint_avg, maint_above=maint_above,
                           maint_count=maint_count,
                           current_month=_date.today().month,
                           current_year=_date.today().year,
                           maint_uploaded_keys=maint_uploaded_keys,
                           has_unassigned=has_unassigned,
                           codes=codes, log_rows=log_rows,
                           edit_emp=edit_emp, del_emp=del_emp,
                           categories=config.CATEGORIES,
                           plant_list=plant_rows,
                           plant_rows=plant_rows, plant_count=plant_count,
                           tp_codes=tp_codes, plant_log=plant_log,
                           tp_last_sync=tp_last_sync, tp_sheet_id=tp_sheet_id,
                           tp_worksheet=tp_worksheet,
                           edit_plant=edit_plant, del_plant=del_plant,
                           exco_opts=exco_opts, bh_opts=bh_opts,
                           today=str(_date.today()),
                           today_first=str(_date.today().replace(day=1)),
                           bg_auto=bg_auto(), bg_animate=bg_animate(), bg_theme=bg_theme(),
                           active_page="sysconfig")


@app.route("/sysconfig/action/save-smtp", methods=["POST"])
@auth.login_required
def sysconfig_save_smtp():
    pwd = request.form.get("password", "").strip()
    to_save = {
        "smtp_host":        request.form.get("host", "").strip(),
        "smtp_port":        request.form.get("port", "587").strip(),
        "smtp_sender":      request.form.get("sender", "").strip(),
        "smtp_use_tls":     "true" if "use_tls" in request.form else "false",
        "email_default_to": request.form.get("default_to", "").strip(),
        "email_default_cc": request.form.get("default_cc", "").strip(),
        "email_subject":    request.form.get("subject", "").strip(),
    }
    if pwd:
        to_save["smtp_password"] = pwd
    database.set_settings_bulk(to_save)
    flash("Email settings saved.", "success")
    return redirect(url_for("sysconfig_page") + "?m=smtp")


@app.route("/sysconfig/action/test-smtp", methods=["POST"])
@auth.login_required
def sysconfig_test_smtp():
    cfg = email_helper.get_smtp_config()
    try:
        res = email_helper.send_report_email(
            to_emails=cfg["sender"], cc_emails="",
            subject="Test email — RDC-OPS System Config",
            body="This is a test email confirming your SMTP settings work.",
        )
        if res["success"]:
            flash(f"Test email sent to {cfg['sender']}. Check the inbox.", "success")
        else:
            flash(f"Test failed: {res['error']}", "error")
    except Exception as exc:
        flash(f"Test failed: {exc}", "error")
    return redirect(url_for("sysconfig_page") + "?m=smtp")


@app.route("/sysconfig/action/save-oracle", methods=["POST"])
@auth.login_required
def sysconfig_save_oracle():
    pwd = request.form.get("password", "").strip()
    to_save = {
        "oracle_host":              request.form.get("host", "").strip(),
        "oracle_port":              request.form.get("port", "").strip(),
        "oracle_service":           request.form.get("service", "").strip(),
        "oracle_user":              request.form.get("user", "").strip(),
        "oracle_status_filter":     request.form.get("status_filter", "").strip(),
        "oracle_instantclient_dir": request.form.get("instantclient", "").strip(),
    }
    if pwd:
        to_save["oracle_password"] = pwd
    database.set_settings_bulk(to_save)
    flash("Oracle settings saved.", "success")
    return redirect(url_for("sysconfig_page") + "?m=oracle")


@app.route("/sysconfig/action/test-oracle", methods=["POST"])
@auth.login_required
def sysconfig_test_oracle():
    try:
        result = oracle_connector.test_connection()
        if result["success"]:
            flash(f"Connection successful! {result.get('version', '')}", "success")
        else:
            flash(f"Connection failed: {result['error']}", "error")
    except Exception as exc:
        flash(f"Test failed: {exc}", "error")
    return redirect(url_for("sysconfig_page") + "?m=oracle")


@app.route("/sysconfig/action/fetch-all-oracle", methods=["POST"])
@auth.login_required
def sysconfig_fetch_all_oracle():
    from_s  = request.form.get("from_date", str(_date.today().replace(day=1)))
    to_s    = request.form.get("to_date",   str(_date.today()))
    replace = request.form.get("mode", "replace") == "replace"
    errors  = []
    success_msgs = []
    try:
        fd = _date.fromisoformat(from_s)
        td = _date.fromisoformat(to_s)
        _set_progress(5, "Testing Oracle connection…")
        conn_test = oracle_connector.test_connection()
        if not conn_test["success"]:
            flash(f"Oracle connection failed: {conn_test['error']}", "error")
            _set_progress(100, "Complete")
            return jsonify({"ok": False, "redirect": url_for("sysconfig_page") + "?m=fetch"})

        # 1. I&D backend data
        try:
            _set_progress(20, "Fetching I&D backend data…")
            df_id, warns_id = oracle_connector.fetch_backend_data(fd, td)
            for w in warns_id:
                flash(w, "warning")
            if df_id.empty:
                flash("I&D: Oracle returned no rows for the selected date range.", "warning")
            else:
                saved = oracle_connector.save_oracle_backend_data(df_id, fd, td, replace=replace)
                success_msgs.append(f"I&D: {saved:,} rows")
        except Exception as exc:
            errors.append(f"I&D fetch failed: {exc}")

        # 2. TP data
        try:
            _set_progress(45, "Fetching TP plant data…")
            raw_tp, warns_tp = oracle_connector.fetch_tp_data(from_s, to_s)
            for w in warns_tp:
                flash(w, "warning")
            parsed_tp, skip_tp = tp_calculator.parse_oracle_df(raw_tp)
            oracle_connector.save_tp_oracle_data(raw_tp, from_s, to_s, parsed_tp, replace=replace)
            _mss("tp", "skip_log", skip_tp)
            _mss("tp", "ora_from", from_s)
            _mss("tp", "ora_to",   to_s)
            success_msgs.append(f"TP: {len(parsed_tp)} rows")
        except Exception as exc:
            errors.append(f"TP fetch failed: {exc}")

        # 3. BTRTP data
        try:
            _set_progress(70, "Fetching BTRTP batcher data…")
            raw_bt, warns_bt = oracle_connector.fetch_btrtp_data(from_s, to_s)
            for w in warns_bt:
                flash(w, "warning")
            parsed_bt, skip_bt = btrtp_calculator.parse_btrtp_oracle_df(raw_bt)
            oracle_connector.save_btrtp_oracle_data(parsed_bt, replace=replace)
            _mss("btrtp", "skip_log", skip_bt)
            _mss("btrtp", "ora_from", from_s)
            _mss("btrtp", "ora_to",   to_s)
            success_msgs.append(f"BTRTP: {len(parsed_bt)} rows")
        except Exception as exc:
            errors.append(f"BTRTP fetch failed: {exc}")

        _set_progress(90, "Cleaning old records…")
        database.purge_old_oracle_data()
        now_str = _dt.now().strftime("%Y-%m-%d %H:%M")
        database.set_settings_bulk({"sysconfig_last_oracle_fetch": now_str})

        if success_msgs:
            flash("✅ Oracle fetch complete — " + " · ".join(success_msgs), "success")
        for e in errors:
            flash(e, "error")

    except Exception as exc:
        flash(f"Oracle fetch failed: {exc}", "error")
    _set_progress(100, "Complete")
    return jsonify({"ok": True, "redirect": url_for("sysconfig_page") + "?m=fetch"})


@app.route("/sysconfig/action/sync-master", methods=["POST"])
@auth.login_required
def sysconfig_sync_master():
    sheet_id  = request.form.get("sheet_id", "").strip()
    worksheet = request.form.get("worksheet", "Sheet1").strip()
    if not sheet_id:
        flash("Please enter a Google Sheet ID.", "error")
        return jsonify({"ok": False, "redirect": url_for("sysconfig_page") + "?m=master"})
    try:
        _set_progress(15, "Connecting to Google Sheets…")
        clean_id = google_sheets.extract_sheet_id(sheet_id)
        _set_progress(35, "Fetching master data…")
        df_sync  = google_sheets.fetch_master_data(clean_id, worksheet)
        now = _dt.now().isoformat(timespec="seconds")
        def _sv(v):
            import math
            if v is None: return ""
            if isinstance(v, float) and math.isnan(v): return ""
            s = str(v).strip()
            return "" if s.lower() == "nan" else s
        rows = [{"employee_code": _sv(r["Employee Code"]),
                 "employee_name": _sv(r["Employee Name"]),
                 "designation":   _sv(r["Designation"]),
                 "category":      _sv(r["Category"]),
                 "plant":         _sv(r["Plant"]),
                 "plant_code":    _sv(r["Plant Code"]),
                 "updated_at":    now}
                for _, r in df_sync.iterrows()]
        _set_progress(70, "Saving employee records…")
        count = google_sheets.save_master_data(rows)
        _set_progress(85, "Updating sync metadata…")
        database.set_settings_bulk({
            "gsheet_id":          clean_id,
            "gsheet_worksheet":   worksheet,
            "gsheet_last_sync":   now,
            "gsheet_last_count":  str(count),
        })
        flash(f"✅ {count} employees synced from Google Sheets.", "success")
    except Exception as exc:
        flash(f"Sync failed: {exc}", "error")
    _set_progress(100, "Complete")
    return jsonify({"ok": True, "redirect": url_for("sysconfig_page") + "?m=master"})


@app.route("/sysconfig/action/toggle-auto-sync", methods=["POST"])
@auth.login_required
def sysconfig_toggle_auto_sync():
    enabled = "enabled" in request.form
    database.set_settings_bulk({"gsheet_auto_sync": "true" if enabled else "false"})
    return redirect(url_for("sysconfig_page") + "?m=master")


@app.route("/sysconfig/action/upload-maintenance", methods=["POST"])
@auth.login_required
def sysconfig_upload_maintenance():
    if "file" not in request.files or request.files["file"].filename == "":
        flash("No file selected.", "error")
        return redirect(url_for("sysconfig_page") + "?m=maintenance")
    f = request.files["file"]
    try:
        month = int(request.form.get("maint_month", _date.today().month))
        year  = int(request.form.get("maint_year",  _date.today().year))
    except (ValueError, TypeError):
        flash("Invalid month or year selected.", "error")
        return redirect(url_for("sysconfig_page") + "?m=maintenance")
    try:
        import calendar as _cal3
        df, warns = data_loader.load_maintenance_cost(f)
        for w in warns:
            flash(w, "warning")
        saved = data_loader.save_maintenance_cost(df, month, year)
        flash(f"Saved {saved:,} plant maintenance cost rows for {_cal3.month_name[month]} {year}.", "success")
    except Exception as exc:
        flash(f"Upload failed: {exc}", "error")
    return redirect(url_for("sysconfig_page") + "?m=maintenance")


@app.route("/sysconfig/action/upload-backend", methods=["POST"])
@auth.login_required
def sysconfig_upload_backend():
    if "file" not in request.files or request.files["file"].filename == "":
        flash("No file selected.", "error")
        return redirect(url_for("sysconfig_page") + "?m=backend")
    f = request.files["file"]
    replace = request.form.get("mode", "replace") == "replace"
    try:
        df, warns = data_loader.load_backend_data(f)
        for w in warns:
            flash(w, "warning")
        saved = data_loader.save_backend_data(df, source_file=f.filename, replace=replace)
        flash(f"{'Replaced' if replace else 'Appended'} {saved:,} backend rows.", "success")
    except Exception as exc:
        flash(f"Upload failed: {exc}", "error")
    return redirect(url_for("sysconfig_page") + "?m=backend")


# ═══════════════════════════════════════════════════════════════════════════════
# SLA — Slow Loading Alert Module Routes
# ═══════════════════════════════════════════════════════════════════════════════

def _sla_allowed_plant_codes(user):
    """Return list of plant_codes the user may see, or None (= all plants)."""
    role = user.get("role", "")
    if role in (auth.SUPER_ADMIN, auth.HO_VIEWER, auth.FINANCE_VIEWER, auth.UEP_ADMIN):
        return None  # all plants
    # REGIONAL_USER / PLANT_USER — restrict to assigned plants
    plants = user.get("allowed_plants", [])
    # allowed_plants stores plant_names; we need to cross-reference SLA records
    # which store plant_code. Best-effort: use whatever distinct codes are in the DB
    # that match any of the user's allowed_plant_names.
    if not plants:
        return []
    conn = database.get_connection()
    try:
        ph = ",".join("?" * len(plants))
        cur = conn.execute(
            f"SELECT DISTINCT plant_code FROM slow_loading_alert_records WHERE plant_name IN ({ph})",
            list(plants)
        )
        return [r[0] for r in cur.fetchall()] or []
    finally:
        conn.close()


@app.route("/sla")
@app.route("/sla/")
@auth.login_required
def sla_dashboard():
    _ss("active_page", "dashboard")
    user       = g.current_user
    today_str  = str(_date.today())
    codes      = _sla_allowed_plant_codes(user)
    kpis       = database.sla_get_dashboard_kpis(today_str, codes)
    recent     = database.sla_get_report(from_date=today_str, to_date=today_str,
                                          plant_code=codes[0] if (codes and len(codes)==1) else None,
                                          limit=50)
    # Annotate severity
    for r in recent:
        d = float(r.get("delay_minutes") or 0)
        r["_severity"] = "RED" if d > 3 else ("AMBER" if d >= 1 else "")
    # Filter by allowed codes when multiple
    if codes is not None:
        recent = [r for r in recent if r.get("plant_code") in codes]
    return render_template("slow_loading_alert/dashboard.html",
                           today=today_str, kpis=kpis, recent_cases=recent)


@app.route("/sla/report")
@auth.login_required
def sla_report():
    _ss("active_page", "report")
    user = g.current_user
    codes = _sla_allowed_plant_codes(user)
    f = request.args
    plant_filter = f.get("plant_code", "")
    # Enforce plant restriction
    if codes is not None and plant_filter and plant_filter not in codes:
        plant_filter = ""
    rows = database.sla_get_report(
        from_date=f.get("from_date"), to_date=f.get("to_date"),
        plant_code=plant_filter or None,
        batcher_code=f.get("batcher_code") or None,
        tm_number=f.get("tm_number") or None,
        grade=f.get("grade") or None,
        customer=f.get("customer") or None,
        alert_type=f.get("alert_type") or None,
        status=f.get("status") or None,
        limit=500
    )
    if codes is not None:
        rows = [r for r in rows if r.get("plant_code") in codes]
    plant_list = database.sla_get_distinct_plants()
    if codes is not None:
        plant_list = [p for p in plant_list if p["plant_code"] in codes]
    return render_template("slow_loading_alert/report.html",
                           rows=rows, filters=dict(f),
                           plant_list=plant_list,
                           is_plant_user=(user.get("role") == auth.PLANT_USER))


@app.route("/sla/export")
@auth.login_required
def sla_export():
    user  = g.current_user
    codes = _sla_allowed_plant_codes(user)
    f = request.args
    plant_filter = f.get("plant_code", "") or None
    if codes is not None and plant_filter and plant_filter not in codes:
        plant_filter = None
    rows = database.sla_get_report(
        from_date=f.get("from_date"), to_date=f.get("to_date"),
        plant_code=plant_filter,
        batcher_code=f.get("batcher_code") or None,
        tm_number=f.get("tm_number") or None,
        grade=f.get("grade") or None,
        customer=f.get("customer") or None,
        alert_type=f.get("alert_type") or None,
        status=f.get("status") or None,
        limit=5000
    )
    if codes is not None:
        rows = [r for r in rows if r.get("plant_code") in codes]
    if not rows:
        flash("No data to export.", "warning")
        return redirect(url_for("sla_report"))

    import io as _io
    df_exp = pd.DataFrame(rows)
    rename = {
        "alert_date": "Alert Date", "alert_hour": "Alert Hour",
        "plant_code": "Plant Code", "plant_name": "Plant Name",
        "customer": "Customer", "grade": "Grade",
        "batcher_code": "Batcher Code", "batcher_name": "Batcher Name",
        "tm_number": "TM Number", "batched_quantity": "Batched Quantity",
        "mixer_capacity": "Mixer Capacity",
        "loading_time_minutes": "Loading Time (min)",
        "allowed_loading_minutes": "Threshold Time (min)",
        "delay_minutes": "Delay (min)",
        "alert_type": "Alert Type", "status": "Status", "remarks": "Remarks",
    }
    df_exp = df_exp.rename(columns=rename)
    keep = [c for c in rename.values() if c in df_exp.columns]
    df_exp = df_exp[keep]
    # Add severity column
    def _sev(d):
        try:
            d = float(d)
        except Exception:
            return ""
        return "RED" if d > 3 else ("AMBER" if d >= 1 else "")
    if "Delay (min)" in df_exp.columns:
        df_exp.insert(df_exp.columns.get_loc("Delay (min)") + 1, "Severity",
                      df_exp["Delay (min)"].apply(_sev))
    buf = _io.BytesIO()
    with pd.ExcelWriter(buf, engine="openpyxl") as writer:
        df_exp.to_excel(writer, sheet_name="SLA Report", index=False)
    buf.seek(0)
    resp = make_response(buf.read())
    resp.headers["Content-Type"] = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    resp.headers["Content-Disposition"] = f"attachment; filename=SLA_Report_{str(_date.today())}.xlsx"
    return resp


@app.route("/sla/email-mapping")
@auth.login_required
def sla_email_mapping():
    _ss("active_page", "email_mapping")
    from modules.slow_loading_alert import mapping_service
    warnings = []
    mapping = {}
    batcher_map = {}
    try:
        mapping = mapping_service.get_plant_mapping()
    except Exception as exc:
        warnings.append(str(exc))
    try:
        batcher_map = mapping_service.get_batcher_mapping()
    except Exception as exc:
        warnings.append(str(exc))
    return render_template("slow_loading_alert/email_mapping.html",
                           mapping=mapping, batcher_map=batcher_map, warnings=warnings)


@app.route("/sla/email-mapping/refresh", methods=["POST"])
@auth.login_required
def sla_refresh_mapping():
    flash("Mapping refreshed from Google Sheet.", "success")
    return redirect(url_for("sla_email_mapping"))


@app.route("/sla/alert-logs")
@auth.login_required
def sla_alert_logs():
    _ss("active_page", "alert_logs")
    email_logs     = database.sla_get_email_logs(200)
    scheduler_logs = database.sla_get_scheduler_logs(100)
    return render_template("slow_loading_alert/alert_logs.html",
                           email_logs=email_logs, scheduler_logs=scheduler_logs)


@app.route("/sla/configuration")
@auth.login_required
def sla_configuration():
    _ss("active_page", "configuration")
    if g.current_user.get("role") not in (auth.SUPER_ADMIN,):
        # Read-only for non-admin (still shows page, but no save buttons)
        pass
    thresholds = database.sla_get_all_thresholds()
    _setting_keys = [
        "hourly_alert_enabled", "daily_summary_enabled", "daily_summary_time",
        "min_threshold_min", "global_cc", "gsheet_id", "plant_mapping_tab",
        "batcher_mapping_tab",
        "oracle_plant_col", "oracle_salesorder_col", "oracle_linenumber_col",
        "oracle_grade_col", "oracle_batcher_col", "oracle_truck_col",
        "oracle_quantity_col", "oracle_time_col", "oracle_cust_name_col",
        "oracle_grade_name_col", "oracle_master_so_col", "oracle_master_ln_col",
        "oracle_mixer_cap_col",
    ]
    settings = {k: database.get_module_setting("sla", k, "") for k in _setting_keys}
    is_super_admin = g.current_user.get("role") == auth.SUPER_ADMIN
    return render_template("slow_loading_alert/configuration.html",
                           thresholds=thresholds, settings=settings,
                           is_super_admin=is_super_admin)


@app.route("/sla/threshold/add", methods=["POST"])
@auth.login_required
def sla_add_threshold():
    if g.current_user.get("role") != auth.SUPER_ADMIN:
        return render_template("access_denied.html", current_user=g.current_user,
                               required_roles=[auth.SUPER_ADMIN]), 403
    try:
        mc  = float(request.form["mixer_capacity"])
        rq  = float(request.form["reference_quantity"])
        bam = float(request.form["base_allowed_minutes"])
        database.sla_upsert_threshold(mc, rq, bam)
        flash(f"Threshold added for {mc} m³.", "success")
    except Exception as exc:
        flash(f"Error: {exc}", "error")
    return redirect(url_for("sla_configuration"))


@app.route("/sla/threshold/delete/<int:tid>", methods=["POST"])
@auth.login_required
def sla_delete_threshold(tid):
    if g.current_user.get("role") != auth.SUPER_ADMIN:
        return render_template("access_denied.html", current_user=g.current_user,
                               required_roles=[auth.SUPER_ADMIN]), 403
    database.sla_delete_threshold(tid)
    flash("Threshold deleted.", "success")
    return redirect(url_for("sla_configuration"))


@app.route("/sla/settings/save", methods=["POST"])
@auth.login_required
def sla_save_settings():
    if g.current_user.get("role") != auth.SUPER_ADMIN:
        return render_template("access_denied.html", current_user=g.current_user,
                               required_roles=[auth.SUPER_ADMIN]), 403
    keys = ["hourly_alert_enabled", "daily_summary_enabled", "daily_summary_time",
            "min_threshold_min", "global_cc"]
    for k in keys:
        v = request.form.get(k, "")
        database.set_module_setting("sla", k, v)
    flash("SLA settings saved.", "success")
    return redirect(url_for("sla_configuration"))


@app.route("/sla/oracle-cols/save", methods=["POST"])
@auth.login_required
def sla_save_oracle_cols():
    if g.current_user.get("role") != auth.SUPER_ADMIN:
        return render_template("access_denied.html", current_user=g.current_user,
                               required_roles=[auth.SUPER_ADMIN]), 403
    for k in ["oracle_plant_col","oracle_salesorder_col","oracle_linenumber_col",
              "oracle_batcher_col","oracle_truck_col","oracle_quantity_col",
              "oracle_time_col","oracle_grade_col","oracle_cust_name_col",
              "oracle_grade_name_col","oracle_master_so_col","oracle_master_ln_col",
              "oracle_mixer_cap_col"]:
        database.set_module_setting("sla", k, request.form.get(k, "").strip())
    flash("Oracle column names saved.", "success")
    return redirect(url_for("sla_configuration"))


@app.route("/sla/sheet-config/save", methods=["POST"])
@auth.login_required
def sla_save_sheet_config():
    if g.current_user.get("role") != auth.SUPER_ADMIN:
        return render_template("access_denied.html", current_user=g.current_user,
                               required_roles=[auth.SUPER_ADMIN]), 403
    for k in ["gsheet_id", "plant_mapping_tab", "batcher_mapping_tab"]:
        database.set_module_setting("sla", k, request.form.get(k, "").strip())
    flash("Google Sheet configuration saved.", "success")
    return redirect(url_for("sla_configuration"))


@app.route("/sla/manual")
@auth.login_required
def sla_manual_run():
    if g.current_user.get("role") != auth.SUPER_ADMIN:
        return render_template("access_denied.html", current_user=g.current_user,
                               required_roles=[auth.SUPER_ADMIN]), 403
    _ss("active_page", "manual_run")
    yesterday = str(_date.today() - timedelta(days=1))
    smtp_cfg  = email_helper.get_smtp_config()
    test_default = smtp_cfg.get("sender", "")
    return render_template("slow_loading_alert/manual_run.html",
                           yesterday=yesterday, test_email_default=test_default,
                           hourly_result=None, daily_result=None,
                           hourly_preview=None, test_result=None)


@app.route("/sla/manual/run-hourly-preview", methods=["POST"])
@auth.login_required
def sla_run_hourly_preview():
    if g.current_user.get("role") != auth.SUPER_ADMIN:
        return render_template("access_denied.html", current_user=g.current_user,
                               required_roles=[auth.SUPER_ADMIN]), 403
    _ss("active_page", "manual_run")
    yesterday = str(_date.today() - timedelta(days=1))
    smtp_cfg  = email_helper.get_smtp_config()
    result = sla_scheduler.run_hourly_alert_job(preview_only=True)
    preview = result.pop("preview", [])
    return render_template("slow_loading_alert/manual_run.html",
                           yesterday=yesterday,
                           test_email_default=smtp_cfg.get("sender",""),
                           hourly_result=result, hourly_preview=preview,
                           daily_result=None, test_result=None)


@app.route("/sla/manual/run-hourly-send", methods=["POST"])
@auth.login_required
def sla_run_hourly_send():
    if g.current_user.get("role") != auth.SUPER_ADMIN:
        return render_template("access_denied.html", current_user=g.current_user,
                               required_roles=[auth.SUPER_ADMIN]), 403
    _ss("active_page", "manual_run")
    yesterday = str(_date.today() - timedelta(days=1))
    smtp_cfg  = email_helper.get_smtp_config()
    result = sla_scheduler.run_hourly_alert_job(preview_only=False)
    msg = (f"Hourly alert run complete — "
           f"{result.get('total_alerts',0)} alert(s), "
           f"{result.get('total_sent',0)} email(s) sent.")
    flash(msg, "success" if not result.get("errors") else "warning")
    return render_template("slow_loading_alert/manual_run.html",
                           yesterday=yesterday,
                           test_email_default=smtp_cfg.get("sender",""),
                           hourly_result=result, hourly_preview=None,
                           daily_result=None, test_result=None)


@app.route("/sla/manual/run-daily-preview", methods=["POST"])
@auth.login_required
def sla_run_daily_preview():
    if g.current_user.get("role") != auth.SUPER_ADMIN:
        return render_template("access_denied.html", current_user=g.current_user,
                               required_roles=[auth.SUPER_ADMIN]), 403
    _ss("active_page", "manual_run")
    yesterday = str(_date.today() - timedelta(days=1))
    smtp_cfg  = email_helper.get_smtp_config()
    summary_date = request.form.get("summary_date", yesterday)
    result = sla_scheduler.run_daily_summary_job(summary_date, preview_only=True)
    return render_template("slow_loading_alert/manual_run.html",
                           yesterday=yesterday,
                           test_email_default=smtp_cfg.get("sender",""),
                           hourly_result=None, hourly_preview=None,
                           daily_result=result, test_result=None)


@app.route("/sla/manual/run-daily-send", methods=["POST"])
@auth.login_required
def sla_run_daily_send():
    if g.current_user.get("role") != auth.SUPER_ADMIN:
        return render_template("access_denied.html", current_user=g.current_user,
                               required_roles=[auth.SUPER_ADMIN]), 403
    _ss("active_page", "manual_run")
    yesterday = str(_date.today() - timedelta(days=1))
    smtp_cfg  = email_helper.get_smtp_config()
    summary_date = request.form.get("summary_date", yesterday)
    result = sla_scheduler.run_daily_summary_job(summary_date, preview_only=False)
    msg = (f"Daily summary run — "
           f"{result.get('total_checked',0)} record(s), "
           f"{result.get('total_sent',0)} email(s) sent.")
    flash(msg, "success" if not result.get("errors") else "warning")
    return render_template("slow_loading_alert/manual_run.html",
                           yesterday=yesterday,
                           test_email_default=smtp_cfg.get("sender",""),
                           hourly_result=None, hourly_preview=None,
                           daily_result=result, test_result=None)


@app.route("/sla/manual/send-test-email", methods=["POST"])
@auth.login_required
def sla_send_test_email():
    if g.current_user.get("role") != auth.SUPER_ADMIN:
        return render_template("access_denied.html", current_user=g.current_user,
                               required_roles=[auth.SUPER_ADMIN]), 403
    _ss("active_page", "manual_run")
    yesterday = str(_date.today() - timedelta(days=1))
    smtp_cfg  = email_helper.get_smtp_config()
    test_email = request.form.get("test_email", "").strip()
    result = {"success": False, "error": "No email address provided."}
    if test_email:
        from modules.slow_loading_alert.email_service import _send_html
        test_html = (
            "<p>This is a test email from the <strong>Slow Loading Alert</strong> module.</p>"
            "<p>If you received this, SMTP is configured correctly.</p>"
        )
        try:
            _send_html(test_email, "", "SLA Test Email — RDC Operations", test_html)
            result = {"success": True, "error": None}
        except Exception as exc:
            result = {"success": False, "error": str(exc)}
    return render_template("slow_loading_alert/manual_run.html",
                           yesterday=yesterday,
                           test_email_default=test_email or smtp_cfg.get("sender",""),
                           hourly_result=None, hourly_preview=None,
                           daily_result=None, test_result=result)


# SLA JSON APIs

@app.route("/sla/api/dashboard")
@auth.login_required
def sla_api_dashboard():
    user  = g.current_user
    codes = _sla_allowed_plant_codes(user)
    today = str(_date.today())
    return jsonify(database.sla_get_dashboard_kpis(today, codes))


@app.route("/sla/api/report")
@auth.login_required
def sla_api_report():
    user  = g.current_user
    codes = _sla_allowed_plant_codes(user)
    f     = request.args
    plant_filter = f.get("plant_code") or None
    if codes is not None and plant_filter and plant_filter not in codes:
        plant_filter = None
    rows = database.sla_get_report(
        from_date=f.get("from_date"), to_date=f.get("to_date"),
        plant_code=plant_filter, limit=500
    )
    if codes is not None:
        rows = [r for r in rows if r.get("plant_code") in codes]
    return jsonify(rows)


@app.route("/sla/api/run-hourly", methods=["POST"])
@auth.login_required
def sla_api_run_hourly():
    if g.current_user.get("role") != auth.SUPER_ADMIN:
        return jsonify({"error": "Forbidden"}), 403
    result = sla_scheduler.run_hourly_alert_job()
    return jsonify(result)


@app.route("/sla/api/run-daily-summary", methods=["POST"])
@auth.login_required
def sla_api_run_daily():
    if g.current_user.get("role") != auth.SUPER_ADMIN:
        return jsonify({"error": "Forbidden"}), 403
    summary_date = request.json.get("summary_date") if request.is_json else None
    result = sla_scheduler.run_daily_summary_job(summary_date)
    return jsonify(result)


@app.route("/sla/api/logs")
@auth.login_required
def sla_api_logs():
    return jsonify({
        "email_logs":     database.sla_get_email_logs(100),
        "scheduler_logs": database.sla_get_scheduler_logs(50),
    })


# ── Entry point ───────────────────────────────────────────────────────────────

if __name__ == "__main__":
    print("=" * 60)
    print("  RDC Batching Incentive Calculator")
    print("  Flask server starting on http://localhost:2001")
    print("=" * 60)
    app.run(host="0.0.0.0", port=2001, debug=False, use_reloader=False)
